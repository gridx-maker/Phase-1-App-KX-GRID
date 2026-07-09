# ============================================================
# KXGRID — Router: BRAND_HEAD
# ============================================================

from fastapi import APIRouter, Depends, HTTPException, Request, Response, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import uuid
import logging
import os
import base64
import io
import csv
import random
import string
import qrcode
from openpyxl import Workbook
import resend

# Core configurations
from core.database import db
from core.config import twilio_client, SENDER_EMAIL, resend as resend_client
from core.security import get_current_user, hash_password, verify_password, create_token, otp_store

# Shared models
from models import *

logger = logging.getLogger(__name__)

router = APIRouter()

@router.post("/brand-head/crew")
async def create_crew_account(data: CreateUserAccount, user: dict = Depends(get_current_user)):
    """Brand Head creates a Crew (Trainer) account for their own brand"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned to you")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Get brand info
    brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    # Create crew account
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    hashed_password = bcrypt.hashpw(data.password.encode('utf-8'), bcrypt.gensalt())
    
    user_doc = {
        "user_id": user_id,
        "email": data.email,
        "password": hashed_password.decode('utf-8'),
        "name": data.name,
        "role": "trainer",
        "assigned_brand_id": brand_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.users.insert_one(user_doc)
    
    # Send welcome email with credentials (non-blocking)
    email_sent = False
    try:
        result = await send_welcome_email(data.email, data.name, "trainer", data.password, brand.get("name"))
        email_sent = result is not None
    except Exception as e:
        logger.warning(f"Could not send welcome email: {e}")
    
    return {
        "message": "Crew member created successfully",
        "user_id": user_id,
        "email": data.email,
        "name": data.name,
        "role": "trainer",
        "assigned_brand_id": brand_id,
        "brand_name": brand.get("name"),
        "email_sent": email_sent
    }

@router.get("/brand-head/crew")
async def get_brand_head_crew(user: dict = Depends(get_current_user)):
    """Get all crew members for the Brand Head's brand"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    crew = await db.users.find(
        {"role": "trainer", "assigned_brand_id": brand_id},
        {"_id": 0, "password": 0}
    ).to_list(100)
    
    return crew

@router.delete("/brand-head/crew/{user_id}")
async def delete_crew_account(user_id: str, user: dict = Depends(get_current_user)):
    """Brand Head can delete crew members from their brand"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Verify the crew member belongs to this brand
    crew_member = await db.users.find_one({"user_id": user_id, "role": "trainer"})
    if not crew_member:
        raise HTTPException(status_code=404, detail="Crew member not found")
    
    if crew_member.get("assigned_brand_id") != brand_id:
        raise HTTPException(status_code=403, detail="This crew member does not belong to your brand")
    
    await db.users.delete_one({"user_id": user_id})
    
    return {"message": "Crew member deleted successfully"}

@router.get("/brand-head/dashboard")
async def get_brand_head_dashboard(user: dict = Depends(get_current_user)):
    """Get Brand Head dashboard data"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    # Check for brand_id (supports both field names)
    brand_id = user.get("assigned_brand_id") or user.get("brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Get brand details
    brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    # Get programs for this brand (or all programs if brand_id not set on programs)
    programs = await db.programs.find(
        {"$or": [{"brand_id": brand_id}, {"brand_id": {"$exists": False}}]},
        {"_id": 0}
    ).to_list(50)
    
    # Get crew/trainers assigned to this brand
    crew = await db.users.find(
        {"role": "trainer", "assigned_brand_id": brand_id},
        {"_id": 0, "password": 0}
    ).to_list(50)
    
    # If no crew assigned to brand, get all trainers
    if not crew:
        crew = await db.users.find(
            {"role": "trainer"},
            {"_id": 0, "password": 0}
        ).to_list(20)
    
    # Get students enrolled in this brand's programs
    # For now, get all students
    students = await db.students.find(
        {"status": "active"},
        {"_id": 0}
    ).to_list(100)
    
    # Get recent assessments
    assessments = await db.assessments.find(
        {},
        {"_id": 0}
    ).sort("timestamp", -1).to_list(20)
    
    return {
        "brand": brand,
        "programs": programs,
        "crew": crew,
        "students": students[:50],
        "assessments": assessments,
        "stats": {
            "total_programs": len(programs),
            "total_crew": len(crew),
            "total_students": len(students),
            "recent_assessments": len(assessments)
        }
    }

@router.get("/brand-head/brand")
async def get_brand_head_brand(user: dict = Depends(get_current_user)):
    """Get the brand assigned to the current brand head"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    return brand

@router.get("/brand-head/reports/incomplete-students")
async def get_incomplete_students_report(user: dict = Depends(get_current_user)):
    """Get students with incomplete units for the brand head's brand"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Get all units assigned to this brand
    units = await db.program_units.find(
        {"brand_id": brand_id},
        {"_id": 0}
    ).to_list(100)
    
    if not units:
        return {"students": [], "units": [], "message": "No units assigned to this brand"}
    
    unit_ids = [u["unit_id"] for u in units]
    
    # Get all students
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    
    # Get all progress records for these units
    progress_records = await db.student_progress.find(
        {"unit_id": {"$in": unit_ids}},
        {"_id": 0}
    ).to_list(5000)
    
    # Group progress by student
    student_progress = {}
    for record in progress_records:
        sid = record["student_id"]
        if sid not in student_progress:
            student_progress[sid] = {}
        student_progress[sid][record["unit_id"]] = record
    
    # Build report
    incomplete_students = []
    for student in students:
        sid = student["student_id"]
        prog = student_progress.get(sid, {})
        
        student_units = []
        incomplete_count = 0
        for unit in units:
            unit_status = prog.get(unit["unit_id"], {}).get("status", "not_started")
            if unit_status != "completed":
                incomplete_count += 1
            student_units.append({
                "unit_id": unit["unit_id"],
                "unit_name": unit["name"],
                "status": unit_status,
                "completion_date": prog.get(unit["unit_id"], {}).get("completion_date"),
                "score": prog.get(unit["unit_id"], {}).get("score")
            })
        
        if incomplete_count > 0:
            incomplete_students.append({
                "student_id": sid,
                "student_name": student.get("full_name", "Unknown"),
                "mobile": student.get("mobile", ""),
                "email": student.get("email", ""),
                "city": student.get("city", ""),
                "total_units": len(units),
                "completed_units": len(units) - incomplete_count,
                "incomplete_units": incomplete_count,
                "progress_percent": round(((len(units) - incomplete_count) / len(units)) * 100),
                "units": student_units
            })
    
    # Sort by incomplete count (most incomplete first)
    incomplete_students.sort(key=lambda x: x["incomplete_units"], reverse=True)
    
    return {
        "brand_id": brand_id,
        "total_students": len(students),
        "students_with_incomplete": len(incomplete_students),
        "units": [{"unit_id": u["unit_id"], "name": u["name"], "program_id": u["program_id"]} for u in units],
        "students": incomplete_students
    }

@router.get("/brand-head/reports/export-excel")
async def export_incomplete_students_excel(user: dict = Depends(get_current_user)):
    """Export incomplete students report as Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Get brand info
    brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
    brand_name = brand.get("name", "Unknown") if brand else "Unknown"
    
    # Get all units assigned to this brand
    units = await db.program_units.find(
        {"brand_id": brand_id},
        {"_id": 0}
    ).to_list(100)
    
    unit_ids = [u["unit_id"] for u in units]
    
    # Get all students
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    
    # Get all progress records
    progress_records = await db.student_progress.find(
        {"unit_id": {"$in": unit_ids}},
        {"_id": 0}
    ).to_list(5000)
    
    # Group progress by student
    student_progress = {}
    for record in progress_records:
        sid = record["student_id"]
        if sid not in student_progress:
            student_progress[sid] = {}
        student_progress[sid][record["unit_id"]] = record
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Incomplete Students"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    status_complete = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    status_progress = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
    status_not_started = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = f"{brand_name} - Incomplete Students Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Student Name", "Mobile", "Email", "City", "Progress %", "Completed", "Incomplete"]
    for unit in units:
        headers.append(unit["name"][:20])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 5
    for student in students:
        sid = student["student_id"]
        prog = student_progress.get(sid, {})
        
        completed = sum(1 for u in units if prog.get(u["unit_id"], {}).get("status") == "completed")
        incomplete = len(units) - completed
        
        if incomplete == 0:
            continue  # Skip students who completed all units
        
        progress_pct = round((completed / len(units)) * 100) if units else 0
        
        ws.cell(row=row, column=1, value=student.get("full_name", "")).border = thin_border
        ws.cell(row=row, column=2, value=student.get("mobile", "")).border = thin_border
        ws.cell(row=row, column=3, value=student.get("email", "")).border = thin_border
        ws.cell(row=row, column=4, value=student.get("city", "")).border = thin_border
        ws.cell(row=row, column=5, value=f"{progress_pct}%").border = thin_border
        ws.cell(row=row, column=6, value=completed).border = thin_border
        ws.cell(row=row, column=7, value=incomplete).border = thin_border
        
        # Unit status columns
        for col, unit in enumerate(units, 8):
            status = prog.get(unit["unit_id"], {}).get("status", "not_started")
            cell = ws.cell(row=row, column=col, value=status.replace("_", " ").title())
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if status == "completed":
                cell.fill = status_complete
                cell.font = Font(color="FFFFFF")
            elif status == "in_progress":
                cell.fill = status_progress
                cell.font = Font(color="FFFFFF")
            else:
                cell.fill = status_not_started
                cell.font = Font(color="FFFFFF")
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    for col in range(8, 8 + len(units)):
        ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{brand_name.replace(' ', '_')}_Incomplete_Students_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Re-class Management
class ReClassCreate(BaseModel):
    unit_id: str
    scheduled_date: str
    scheduled_time: str
    location: Optional[str] = ""
    notes: Optional[str] = ""
    student_ids: List[str] = []

@router.post("/brand-head/reclass")
async def create_reclass(data: ReClassCreate, user: dict = Depends(get_current_user)):
    """Create a re-class session for students with incomplete units"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Verify the unit belongs to this brand
    unit = await db.program_units.find_one({"unit_id": data.unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    if unit.get("brand_id") != brand_id:
        raise HTTPException(status_code=403, detail="This unit does not belong to your brand")
    
    reclass_id = f"reclass_{uuid.uuid4().hex[:12]}"
    reclass_doc = {
        "reclass_id": reclass_id,
        "unit_id": data.unit_id,
        "unit_name": unit["name"],
        "brand_id": brand_id,
        "scheduled_date": data.scheduled_date,
        "scheduled_time": data.scheduled_time,
        "location": data.location,
        "notes": data.notes,
        "student_ids": data.student_ids,
        "status": "scheduled",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.reclasses.insert_one(reclass_doc)
    
    return {k: v for k, v in reclass_doc.items() if k != "_id"}

@router.get("/brand-head/reclasses")
async def get_brand_reclasses(user: dict = Depends(get_current_user)):
    """Get all re-classes for the brand head's brand"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    reclasses = await db.reclasses.find(
        {"brand_id": brand_id},
        {"_id": 0}
    ).sort("scheduled_date", -1).to_list(100)
    
    return reclasses

@router.put("/brand-head/reclass/{reclass_id}")
async def update_reclass(reclass_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update a re-class (status, add/remove students)"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Verify the reclass belongs to this brand
    reclass = await db.reclasses.find_one({"reclass_id": reclass_id})
    if not reclass:
        raise HTTPException(status_code=404, detail="Re-class not found")
    if reclass.get("brand_id") != brand_id:
        raise HTTPException(status_code=403, detail="This re-class does not belong to your brand")
    
    data.pop("reclass_id", None)
    data.pop("brand_id", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.reclasses.update_one(
        {"reclass_id": reclass_id},
        {"$set": data}
    )
    
    return {"message": "Re-class updated"}

@router.delete("/brand-head/reclass/{reclass_id}")
async def delete_reclass(reclass_id: str, user: dict = Depends(get_current_user)):
    """Delete a re-class"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    reclass = await db.reclasses.find_one({"reclass_id": reclass_id})
    if not reclass:
        raise HTTPException(status_code=404, detail="Re-class not found")
    if reclass.get("brand_id") != brand_id:
        raise HTTPException(status_code=403, detail="This re-class does not belong to your brand")
    
    await db.reclasses.delete_one({"reclass_id": reclass_id})
    
    return {"message": "Re-class deleted"}
