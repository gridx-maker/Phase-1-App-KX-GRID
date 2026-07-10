# ============================================================
# KXGRID — Router: ADMIN
# ============================================================

from fastapi import APIRouter, Depends, HTTPException, Request, Response, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import uuid
import logging
import os
import base64
import io
import csv
import random
import string
import qrcode
from openpyxl import Workbook
import resend
import bcrypt
from io import BytesIO
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Core configurations
from core.database import db
from core.config import twilio_client, SENDER_EMAIL, resend as resend_client
from core.security import get_current_user, hash_password, verify_password, create_token, otp_store

# Shared models
from models import *

# Email and reporting helpers
from core.email import (
    send_welcome_email,
    send_email_async,
    send_report_email,
    generate_assessment_report_data,
    create_university_report_excel
)

logger = logging.getLogger(__name__)

router = APIRouter()

@router.get("/admin/nfc-users/template")
async def download_nfc_template(user: dict = Depends(get_current_user)):
    """Download Excel template for bulk NFC user upload"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "NFC Users"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["NFC_ID", "Student_Name", "Mobile", "Email", "Program", "Role"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Sample data row
    sample_data = ["NFC001", "John Doe", "+919876543210", "john@example.com", "Motorsport Certification", "student"]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
    
    # Instructions
    ws.cell(row=4, column=1, value="Instructions:").font = Font(bold=True)
    ws.cell(row=5, column=1, value="1. NFC_ID: Unique NFC card identifier (required)")
    ws.cell(row=6, column=1, value="2. Student_Name: Full name of the user (required)")
    ws.cell(row=7, column=1, value="3. Mobile: Mobile number with country code (required for password reset)")
    ws.cell(row=8, column=1, value="4. Email: Email address (optional)")
    ws.cell(row=9, column=1, value="5. Program: Program name (optional)")
    ws.cell(row=10, column=1, value="6. Role: student, trainer, brand_head, admin (default: student)")
    ws.cell(row=11, column=1, value="")
    ws.cell(row=12, column=1, value="Default password for all users: NFC1234").font = Font(bold=True, color="FF0000")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=NFC_Users_Template.xlsx"}
    )

@router.post("/admin/nfc-users/upload")
async def upload_nfc_users(file: UploadFile, user: dict = Depends(get_current_user)):
    """Bulk upload NFC users from Excel file"""
    from openpyxl import load_workbook
    from io import BytesIO
    
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")
    
    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    
    success_count = 0
    error_rows = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        nfc_id = str(row[0]).upper().strip() if row[0] else None
        name = str(row[1]).strip() if row[1] else None
        mobile = str(row[2]).strip() if row[2] else None
        email = str(row[3]).strip() if row[3] else None
        program = str(row[4]).strip() if row[4] else None
        role = str(row[5]).strip().lower() if row[5] else "student"
        
        if not nfc_id or not name:
            error_rows.append({"row": row_idx, "error": "NFC_ID and Student_Name are required"})
            continue
        
        # Validate role
        if role not in ["student", "trainer", "brand_head", "admin"]:
            role = "student"
        
        # Check if NFC ID already exists
        existing = await db.nfc_users.find_one({"nfc_id": nfc_id})
        
        nfc_user_doc = {
            "nfc_id": nfc_id,
            "name": name,
            "mobile": mobile,
            "email": email,
            "program": program,
            "role": role,
            "password": DEFAULT_NFC_PASSWORD,
            "password_changed": False,
            "is_active": True,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "uploaded_by": user["user_id"]
        }
        
        if existing:
            # Update existing
            await db.nfc_users.update_one({"nfc_id": nfc_id}, {"$set": nfc_user_doc})
        else:
            # Insert new
            await db.nfc_users.insert_one(nfc_user_doc)
        
        success_count += 1
    
    return {
        "message": f"Uploaded {success_count} NFC users successfully",
        "success_count": success_count,
        "error_count": len(error_rows),
        "errors": error_rows[:10]  # Return first 10 errors
    }

@router.get("/admin/nfc-users")
async def get_all_nfc_users(
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get all NFC users with optional search by NFC ID or name"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    query = {}
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        query = {"$or": [
            {"nfc_id": search_regex},
            {"name": search_regex},
            {"mobile": search_regex},
            {"email": search_regex}
        ]}
    
    nfc_users = await db.nfc_users.find(query, {"_id": 0}).sort("uploaded_at", -1).to_list(500)
    return nfc_users

@router.put("/admin/nfc-users/{nfc_id}/reset-password")
async def admin_reset_nfc_password(nfc_id: str, data: AdminPasswordReset = None, user: dict = Depends(get_current_user)):
    """Admin directly resets a user's password"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    nfc_id = nfc_id.upper().strip()
    
    # Find NFC user
    nfc_user = await db.nfc_users.find_one({"nfc_id": nfc_id})
    if not nfc_user:
        raise HTTPException(status_code=404, detail="NFC user not found")
    
    # Determine new password
    new_password = data.new_password if data and data.new_password else DEFAULT_NFC_PASSWORD
    
    # Hash if not default
    if new_password != DEFAULT_NFC_PASSWORD:
        hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    else:
        hashed_password = DEFAULT_NFC_PASSWORD
    
    # Update in nfc_users
    await db.nfc_users.update_one(
        {"nfc_id": nfc_id},
        {"$set": {
            "password": hashed_password,
            "password_changed": new_password != DEFAULT_NFC_PASSWORD,
            "password_reset_at": datetime.now(timezone.utc).isoformat(),
            "password_reset_by": user["user_id"]
        }}
    )
    
    # Also update in users if linked
    await db.users.update_one(
        {"nfc_card_id": nfc_id},
        {"$set": {
            "password": hashed_password,
            "password_changed": new_password != DEFAULT_NFC_PASSWORD
        }}
    )
    
    return {"message": f"Password reset to {'default (NFC1234)' if new_password == DEFAULT_NFC_PASSWORD else 'custom password'}"}

@router.put("/admin/nfc-users/{nfc_id}/toggle-status")
async def toggle_nfc_user_status(nfc_id: str, user: dict = Depends(get_current_user)):
    """Activate or deactivate an NFC user"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    nfc_id = nfc_id.upper().strip()
    
    nfc_user = await db.nfc_users.find_one({"nfc_id": nfc_id})
    if not nfc_user:
        raise HTTPException(status_code=404, detail="NFC user not found")
    
    new_status = not nfc_user.get("is_active", True)
    
    await db.nfc_users.update_one(
        {"nfc_id": nfc_id},
        {"$set": {"is_active": new_status}}
    )
    
    return {"message": f"NFC user {'activated' if new_status else 'deactivated'}", "is_active": new_status}

@router.post("/admin/assessment-categories")
async def create_assessment_category(data: AssessmentCategoryCreate, user: dict = Depends(get_current_user)):
    """Create a custom assessment category"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    category_id = f"cat_{uuid.uuid4().hex[:12]}"
    category_doc = {
        "category_id": category_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.assessment_categories.insert_one(category_doc)
    return {k: v for k, v in category_doc.items() if k != "_id"}

@router.get("/admin/assessment-categories")
async def get_assessment_categories(user: dict = Depends(get_current_user)):
    """Get all assessment categories"""
    if user.get("role") not in ["trainer", "admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    categories = await db.assessment_categories.find({}, {"_id": 0}).to_list(100)
    
    # If no categories exist, seed default ones
    if not categories:
        default_categories = [
            {"name": "Skill Control", "description": "Ability to control vehicle/equipment", "scale_min": 1, "scale_max": 5, "is_active": True, "brand_ids": []},
            {"name": "Discipline", "description": "Following rules and protocols", "scale_min": 1, "scale_max": 5, "is_active": True, "brand_ids": []},
            {"name": "Safety Awareness", "description": "Understanding and practicing safety measures", "scale_min": 1, "scale_max": 5, "is_active": True, "brand_ids": []},
            {"name": "Execution", "description": "Quality of task execution", "scale_min": 1, "scale_max": 5, "is_active": True, "brand_ids": []},
            {"name": "Teamwork", "description": "Collaboration with others", "scale_min": 1, "scale_max": 5, "is_active": True, "brand_ids": []}
        ]
        
        for cat in default_categories:
            cat_id = f"cat_{uuid.uuid4().hex[:12]}"
            cat_doc = {
                "category_id": cat_id,
                **cat,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": "system"
            }
            await db.assessment_categories.insert_one(cat_doc)
            categories.append({k: v for k, v in cat_doc.items() if k != "_id"})
    
    # Enrich with brand names
    all_brands = await db.brands.find({}, {"_id": 0, "brand_id": 1, "name": 1}).to_list(100)
    brand_map = {b["brand_id"]: b["name"] for b in all_brands}
    
    for cat in categories:
        cat["brand_names"] = [brand_map.get(bid, "Unknown") for bid in cat.get("brand_ids", [])]
    
    return categories

@router.put("/admin/assessment-categories/{category_id}")
async def update_assessment_category(category_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update an assessment category"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data.pop("category_id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.assessment_categories.update_one(
        {"category_id": category_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    return {"message": "Category updated"}

@router.delete("/admin/assessment-categories/{category_id}")
async def delete_assessment_category(category_id: str, user: dict = Depends(get_current_user)):
    """Delete an assessment category"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.assessment_categories.delete_one({"category_id": category_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    return {"message": "Category deleted"}

@router.get("/admin/stats")
async def get_admin_stats(user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    total_students = await db.students.count_documents({})
    active_students = await db.students.count_documents({"status": "active"})
    total_programs = await db.programs.count_documents({})
    total_batches = await db.batches.count_documents({})
    total_certificates = await db.certificates.count_documents({})
    
    # Medical flags
    medical_flags = await db.students.count_documents({"medical_conditions": {"$ne": []}})
    
    # Attendance stats
    attendance_pipeline = [
        {"$group": {"_id": None, "total": {"$sum": 1}}}
    ]
    attendance_result = await db.attendance.aggregate(attendance_pipeline).to_list(1)
    total_attendance = attendance_result[0]["total"] if attendance_result else 0
    
    # Top performers
    top_students = await db.students.find(
        {"status": "active"},
        {"_id": 0, "student_id": 1, "full_name": 1, "average_rating": 1}
    ).sort("average_rating", -1).limit(5).to_list(5)
    
    return {
        "total_students": total_students,
        "active_students": active_students,
        "total_programs": total_programs,
        "total_batches": total_batches,
        "total_certificates": total_certificates,
        "medical_flags": medical_flags,
        "total_attendance_records": total_attendance,
        "top_performers": top_students,
        "placement_ready": await db.students.count_documents({"average_rating": {"$gte": 4.0}})
    }

@router.put("/admin/students/{student_id}")
async def update_student(student_id: str, data: dict, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Remove protected fields
    data.pop("student_id", None)
    data.pop("user_id", None)
    data.pop("nfc_card_id", None)
    
    result = await db.students.update_one(
        {"student_id": student_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
    
    return {"message": "Student updated"}

@router.delete("/admin/students/{student_id}")
async def delete_student(student_id: str, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.students.delete_one({"student_id": student_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
    
    return {"message": "Student removed"}

@router.post("/admin/message")
async def create_admin_message(data: AdminMessage, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    message_id = f"msg_{uuid.uuid4().hex[:12]}"
    message_doc = {
        "message_id": message_id,
        **data.model_dump(),
        "created_by": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "active": True
    }
    
    await db.admin_messages.insert_one(message_doc)
    return {k: v for k, v in message_doc.items() if k != "_id"}

@router.get("/admin/messages")
async def get_admin_messages():
    messages = await db.admin_messages.find(
        {"active": True}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(10)
    return messages

@router.get("/admin/reports/batch/{batch_id}")
async def get_batch_report(batch_id: str, user: dict = Depends(get_current_user)):
    """Generate assessment report for a specific batch (for university)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    batch = await db.batches.find_one({"batch_id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    program = await db.programs.find_one({"program_id": batch.get("program_id")}, {"_id": 0})
    program_name = program.get("name", "Unknown") if program else "Unknown"
    
    report_data = await generate_assessment_report_data(batch_id=batch_id, report_type="batch")
    
    wb = create_university_report_excel(report_data, f"Batch: {batch_id} | Program: {program_name}")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"university_report_batch_{batch_id}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/admin/reports/weekly")
async def get_weekly_report(user: dict = Depends(get_current_user)):
    """Generate weekly assessment report for all active students (for university)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Get all students' data
    report_data = await generate_assessment_report_data(report_type="weekly")
    
    # Get week range
    today = datetime.now(timezone.utc)
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    
    wb = create_university_report_excel(
        report_data, 
        f"Weekly Report ({week_start.strftime('%Y-%m-%d')} to {week_end.strftime('%Y-%m-%d')})"
    )
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"university_weekly_report_{today.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/admin/reports/monthly")
async def get_monthly_report(user: dict = Depends(get_current_user)):
    """Generate monthly assessment report for all active students (for university)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    report_data = await generate_assessment_report_data(report_type="monthly")
    
    today = datetime.now(timezone.utc)
    month_name = today.strftime('%B %Y')
    
    wb = create_university_report_excel(report_data, f"Monthly Report - {month_name}")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"university_monthly_report_{today.strftime('%Y%m')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

class BatchCompletion(BaseModel):
    batch_id: str
    completion_date: Optional[str] = None

@router.post("/admin/reports/completion")
async def generate_completion_report(data: BatchCompletion, user: dict = Depends(get_current_user)):
    """Generate program completion report for a batch (final report for university)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    batch = await db.batches.find_one({"batch_id": data.batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    program = await db.programs.find_one({"program_id": batch.get("program_id")}, {"_id": 0})
    program_name = program.get("name", "Unknown") if program else "Unknown"
    
    report_data = await generate_assessment_report_data(batch_id=data.batch_id, report_type="completion")
    
    completion_date = data.completion_date or datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    wb = create_university_report_excel(
        report_data, 
        f"PROGRAM COMPLETION | Batch: {data.batch_id} | Program: {program_name} | Completed: {completion_date}"
    )
    
    # Mark batch as completed
    await db.batches.update_one(
        {"batch_id": data.batch_id},
        {"$set": {
            "status": "completed",
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "completion_report_generated": True
        }}
    )
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"university_completion_report_{data.batch_id}_{completion_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/admin/media/gallery")
async def get_all_media(user: dict = Depends(get_current_user)):
    """Get all media items (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    items = await db.media_gallery.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    return items

@router.post("/admin/media/gallery")
async def add_media_item(data: MediaItem, user: dict = Depends(get_current_user)):
    """Add media item (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Validate that either url or base64 is provided
    if not data.url and not data.media_base64:
        raise HTTPException(status_code=400, detail="Either URL or uploaded file is required")
    
    media_id = f"media_{uuid.uuid4().hex[:12]}"
    media_doc = {
        "media_id": media_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    await db.media_gallery.insert_one(media_doc)
    return {k: v for k, v in media_doc.items() if k != "_id"}

@router.put("/admin/media/gallery/{media_id}")
async def update_media_item(media_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update media item (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data.pop("media_id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.media_gallery.update_one(
        {"media_id": media_id},
        {"$set": data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Media item not found")
    return {"message": "Media updated"}

@router.delete("/admin/media/gallery/{media_id}")
async def delete_media_item(media_id: str, user: dict = Depends(get_current_user)):
    """Delete media item (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.media_gallery.delete_one({"media_id": media_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Media item not found")
    return {"message": "Media deleted"}

@router.post("/admin/reports/email")
async def email_report(data: ReportEmailRequest, user: dict = Depends(get_current_user)):
    """Send report to configured university email"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Get university email from settings
    settings = await db.site_settings.find_one({"setting_id": "main"}, {"_id": 0})
    if not settings or not settings.get("university_email"):
        raise HTTPException(status_code=400, detail="University email not configured. Please set it in CMS Settings.")
    
    recipient_email = settings["university_email"]
    
    result = await send_report_email(data.report_type, recipient_email, data.batch_id)
    return {"message": f"Report sent to {recipient_email}", **result}

@router.post("/admin/reports/email-custom")
async def email_report_custom(data: dict, user: dict = Depends(get_current_user)):
    """Send report to a custom email address"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    recipient_email = data.get("email")
    report_type = data.get("report_type", "weekly")
    batch_id = data.get("batch_id")
    
    if not recipient_email:
        raise HTTPException(status_code=400, detail="Email address required")
    
    result = await send_report_email(report_type, recipient_email, batch_id)
    return {"message": f"Report sent to {recipient_email}", **result}

@router.post("/admin/test-email")
async def test_email(user: dict = Depends(get_current_user)):
    """Send a test email to verify configuration"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    settings = await db.site_settings.find_one({"setting_id": "main"}, {"_id": 0})
    if not settings or not settings.get("university_email"):
        raise HTTPException(status_code=400, detail="University email not configured")
    
    html_content = """
    <html>
    <body style="font-family: Arial, sans-serif; background-color: #0a0a0a; color: #ffffff; padding: 40px;">
        <div style="max-width: 600px; margin: 0 auto; background-color: #1a1a1a; border-radius: 12px; padding: 40px;">
            <h1 style="color: #00f0ff;">KotlerX Email Test</h1>
            <p style="color: #ffffff;">This is a test email from KotlerX platform.</p>
            <p style="color: #a0a0a0;">If you received this email, the email configuration is working correctly.</p>
        </div>
    </body>
    </html>
    """
    
    try:
        await send_email_async(settings["university_email"], "KotlerX - Test Email", html_content)
        return {"message": f"Test email sent to {settings['university_email']}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email failed: {str(e)}")

@router.post("/admin/brand-heads/assign")
async def assign_brand_head(data: BrandHeadAssign, user: dict = Depends(get_current_user)):
    """Assign a user as Brand Head for a specific brand (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Verify brand exists
    brand = await db.brands.find_one({"brand_id": data.brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    # Verify user exists
    target_user = await db.users.find_one({"user_id": data.user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update user role and assigned brand
    await db.users.update_one(
        {"user_id": data.user_id},
        {"$set": {
            "role": "brand_head",
            "assigned_brand_id": data.brand_id,
            "assigned_brand_name": brand.get("name"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update brand with brand_head_id
    await db.brands.update_one(
        {"brand_id": data.brand_id},
        {"$set": {
            "brand_head_id": data.user_id,
            "brand_head_name": target_user.get("name"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"User assigned as Brand Head for {brand.get('name')}"}

@router.delete("/admin/brand-heads/{user_id}")
async def remove_brand_head(user_id: str, user: dict = Depends(get_current_user)):
    """Remove Brand Head role from a user (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    target_user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    brand_id = target_user.get("assigned_brand_id")
    
    # Remove brand head from user
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role": "student"}, "$unset": {"assigned_brand_id": "", "assigned_brand_name": ""}}
    )
    
    # Remove brand head from brand
    if brand_id:
        await db.brands.update_one(
            {"brand_id": brand_id},
            {"$unset": {"brand_head_id": "", "brand_head_name": ""}}
        )
    
    return {"message": "Brand Head role removed"}

@router.post("/admin/crew/assign-brand")
async def assign_crew_to_brand(data: dict, user: dict = Depends(get_current_user)):
    """Assign a crew/trainer to a specific brand (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    user_id = data.get("user_id")
    brand_id = data.get("brand_id")
    
    if not user_id:
        raise HTTPException(status_code=400, detail="user_id required")
    
    # Verify user exists and is a trainer
    target_user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if brand_id:
        # Verify brand exists
        brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
        if not brand:
            raise HTTPException(status_code=404, detail="Brand not found")
        
        # Assign to brand
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {
                "assigned_brand_id": brand_id,
                "assigned_brand_name": brand.get("name"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        return {"message": f"Crew assigned to {brand.get('name')}"}
    else:
        # Remove brand assignment
        await db.users.update_one(
            {"user_id": user_id},
            {"$unset": {"assigned_brand_id": "", "assigned_brand_name": ""}}
        )
        return {"message": "Brand assignment removed"}

class CreateCrewRequest(BaseModel):
    name: str
    email: str
    password: str
    brand_id: Optional[str] = None

@router.post("/admin/crew/create")
async def create_crew_member(data: CreateCrewRequest, user: dict = Depends(get_current_user)):
    """Create a new crew/trainer member (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    password_hash = bcrypt.hashpw(data.password.encode(), bcrypt.gensalt()).decode()
    
    new_crew = {
        "user_id": user_id,
        "email": data.email.lower(),
        "name": data.name,
        "password_hash": password_hash,
        "role": "trainer",
        "assigned_brand_id": data.brand_id if data.brand_id else None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("user_id")
    }
    
    await db.users.insert_one(new_crew)
    
    return {"message": "Crew member created", "user_id": user_id}

@router.get("/admin/crew")
async def get_all_crew(user: dict = Depends(get_current_user)):
    """Get all crew/trainers (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    crew = await db.users.find(
        {"role": "trainer"},
        {"_id": 0, "password": 0}
    ).to_list(100)
    
    return crew

@router.get("/admin/brand-heads")
async def get_all_brand_heads(user: dict = Depends(get_current_user)):
    """Get all brand heads (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    brand_heads = await db.users.find(
        {"role": "brand_head"},
        {"_id": 0, "password": 0}
    ).to_list(100)
    
    return brand_heads

@router.post("/admin/users/brand-manager")
async def create_brand_manager_account(data: CreateUserAccount, user: dict = Depends(get_current_user)):
    """Admin creates a Brand Manager (Brand Head) account with credentials"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    if not data.brand_id:
        raise HTTPException(status_code=400, detail="Brand ID is required for Brand Manager")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Verify brand exists
    brand = await db.brands.find_one({"brand_id": data.brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    # Create user account
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    hashed_password = bcrypt.hashpw(data.password.encode('utf-8'), bcrypt.gensalt())
    
    user_doc = {
        "user_id": user_id,
        "email": data.email,
        "password": hashed_password.decode('utf-8'),
        "name": data.name,
        "role": "brand_head",
        "assigned_brand_id": data.brand_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.users.insert_one(user_doc)
    
    # Update brand with brand_head info
    await db.brands.update_one(
        {"brand_id": data.brand_id},
        {"$set": {"brand_head_id": user_id, "brand_head_name": data.name}}
    )
    
    # Send welcome email with credentials (non-blocking, doesn't fail if email not configured)
    email_sent = False
    try:
        result = await send_welcome_email(data.email, data.name, "brand_head", data.password, brand.get("name"))
        email_sent = result is not None
    except Exception as e:
        logger.warning(f"Could not send welcome email: {e}")
    
    return {
        "message": "Brand Manager created successfully",
        "user_id": user_id,
        "email": data.email,
        "name": data.name,
        "role": "brand_head",
        "assigned_brand_id": data.brand_id,
        "brand_name": brand.get("name"),
        "email_sent": email_sent
    }

@router.get("/admin/callback-requests")
async def get_callback_requests(user: dict = Depends(get_current_user)):
    """Get all callback requests (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    requests = await db.callback_requests.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return requests

@router.put("/admin/callback-requests/{request_id}")
async def update_callback_request(request_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update callback request status"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = user["user_id"]
    
    await db.callback_requests.update_one(
        {"request_id": request_id},
        {"$set": data}
    )
    
    return {"message": "Callback request updated"}
