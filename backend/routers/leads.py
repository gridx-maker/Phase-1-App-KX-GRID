# ============================================================
# KXGRID — Router: LEADS
# ============================================================

from fastapi import APIRouter, Depends, HTTPException, Request, Response, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import uuid
import logging
import os
import base64
import io
import csv
import random
import string
import qrcode
from openpyxl import Workbook
import resend

# Core configurations
from core.database import db
from core.config import twilio_client, SENDER_EMAIL, resend as resend_client
from core.security import get_current_user, hash_password, verify_password, create_token, otp_store

# Shared models
from models import *

logger = logging.getLogger(__name__)

router = APIRouter()

@router.post("/leads")
async def create_lead(data: LeadRegistration):
    """Public endpoint for program interest registration"""
    lead_id = f"lead_{uuid.uuid4().hex[:12]}"
    lead_doc = {
        "lead_id": lead_id,
        **data.model_dump(),
        "status": "new",  # new, contacted, converted, dropped
        "created_at": datetime.now(timezone.utc).isoformat(),
        "notes": ""
    }
    
    await db.leads.insert_one(lead_doc)
    return {"message": "Lead registered", "lead_id": lead_id}

@router.get("/admin/leads")
async def get_leads(user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    leads = await db.leads.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return leads

@router.put("/admin/leads/{lead_id}")
async def update_lead(lead_id: str, data: dict, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data.pop("lead_id", None)
    data.pop("created_at", None)
    
    result = await db.leads.update_one(
        {"lead_id": lead_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    return {"message": "Lead updated"}

@router.get("/admin/leads/stats")
async def get_lead_stats(user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    total = await db.leads.count_documents({})
    new_leads = await db.leads.count_documents({"status": "new"})
    contacted = await db.leads.count_documents({"status": "contacted"})
    converted = await db.leads.count_documents({"status": "converted"})
    
    return {
        "total": total,
        "new": new_leads,
        "contacted": contacted,
        "converted": converted
    }

@router.get("/admin/leads/export")
async def export_leads_excel(user: dict = Depends(get_current_user)):
    """Export all leads to Excel file"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    leads = await db.leads.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Headers
    headers = ["Lead ID", "Name", "Location", "Mobile", "Program Interest", "Fee Type", "Status", "Created At", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
    
    # Data rows
    for row, lead in enumerate(leads, 2):
        ws.cell(row=row, column=1, value=lead.get("lead_id", ""))
        ws.cell(row=row, column=2, value=lead.get("name", ""))
        ws.cell(row=row, column=3, value=lead.get("location", ""))
        ws.cell(row=row, column=4, value=lead.get("mobile", ""))
        ws.cell(row=row, column=5, value=lead.get("program_interest", ""))
        ws.cell(row=row, column=6, value=lead.get("fee_type", ""))
        ws.cell(row=row, column=7, value=lead.get("status", ""))
        ws.cell(row=row, column=8, value=lead.get("created_at", ""))
        ws.cell(row=row, column=9, value=lead.get("notes", ""))
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"kotlerx_leads_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/callback-request")
async def create_callback_request(data: CallbackRequestCreate):
    """Create a callback request (public - converts to lead)"""
    request_id = f"callback_{uuid.uuid4().hex[:12]}"
    
    # Create callback request
    callback_doc = {
        "request_id": request_id,
        **data.model_dump(),
        "status": "new",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.callback_requests.insert_one(callback_doc)
    
    # Also create a lead entry
    lead_id = f"lead_{uuid.uuid4().hex[:12]}"
    lead_doc = {
        "lead_id": lead_id,
        "name": data.name,
        "mobile": data.phone,
        "email": data.email,
        "location": "Callback Request",
        "program_interest": data.message or "General Inquiry",
        "fee_type": "unknown",
        "status": "new",
        "source": "callback_request",
        "callback_request_id": request_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.leads.insert_one(lead_doc)
    
    return {"message": "Callback request received. Our team will contact you soon.", "request_id": request_id}
