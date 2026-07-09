from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File
from fastapi.security import HTTPBearer
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from jose import jwt, JWTError
import bcrypt
import random
import string
import httpx
import base64
import io
import qrcode
from openpyxl import Workbook
import resend
import asyncio
from twilio.rest import Client as TwilioClient

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Resend config
resend.api_key = os.environ.get('RESEND_API_KEY', '')

# Twilio config
TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID', '')
TWILIO_AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN', '')
TWILIO_PHONE_NUMBER = os.environ.get('TWILIO_PHONE_NUMBER', '')

# Initialize Twilio client only if credentials are provided
twilio_client = None
if TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN:
    try:
        twilio_client = TwilioClient(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
    except Exception as e:
        logging.warning(f"Failed to initialize Twilio client: {e}")
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')

# MongoDB connection
# MongoDB connection - use environment variable with fallback for local development
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'test_database')]

# JWT Config - Required environment variables with sensible defaults
JWT_SECRET = os.environ.get('JWT_SECRET_KEY', 'default_secret_key_change_in_production')
JWT_ALGORITHM = os.environ.get('JWT_ALGORITHM', 'HS256')
JWT_EXPIRATION = int(os.environ.get('JWT_EXPIRATION_HOURS', '168'))

# Create the main app
app = FastAPI(title="KotlerX API", version="1.0.0")

# Create router with /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer(auto_error=False)

# ======================== HEALTH CHECK (Required for Kubernetes) ========================
@app.get("/health")
def health_check():
    """Health check endpoint for Kubernetes liveness/readiness probes"""
    return {"status": "healthy", "service": "kxgrid-api"}

@app.get("/")
def app_root():
    """Root endpoint"""
    return {"message": "KotlerX API", "status": "running", "version": "1.0.0"}

# ======================== ROOT ENDPOINT ========================
@api_router.get("/")
async def root():
    return {"message": "KotlerX API v1.0", "status": "running"}

@api_router.get("/health")
def api_health_check():
    """Health check endpoint accessible via /api/health for Kubernetes"""
    return {"status": "healthy", "service": "kxgrid-api"}

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ======================== MODELS ========================

class UserBase(BaseModel):
    email: EmailStr
    name: str
    picture: Optional[str] = None

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "student"  # student, trainer, admin

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class OTPRequest(BaseModel):
    phone: str

class OTPVerify(BaseModel):
    phone: str
    otp: str

class LeadRegistration(BaseModel):
    name: str
    location: str
    mobile: str
    program_interest: str
    fee_type: str  # cash, loan

class StudentRegistration(BaseModel):
    user_id: str
    photo_url: Optional[str] = None
    full_name: str
    mobile: str
    age: int
    blood_group: str
    address: str
    city: str
    state: str
    emergency_contact: str
    highest_degree: str
    occupation_type: str  # student, working, business, govt, freelancer
    occupation_detail: Optional[str] = None  # college/office/company/dept name
    medical_conditions: List[str] = []
    other_medical: Optional[str] = None
    blood_donation_willing: bool = False
    nfc_card_id: Optional[str] = None

class ProgramCreate(BaseModel):
    name: str
    program_type: str  # certification, diploma, pg_diploma
    description: str
    duration_weeks: int
    batch_size: int = 20
    modules: List[Dict[str, Any]] = []
    registration_open: bool = True
    next_batch_date: Optional[str] = None
    brand_id: Optional[str] = None  # Primary brand (legacy)
    brand_ids: List[str] = []  # Multiple brands involved
    highlights: List[str] = []  # Program highlights visible in public view

class ProgramUpdate(BaseModel):
    name: Optional[str] = None
    program_type: Optional[str] = None
    description: Optional[str] = None
    duration_weeks: Optional[int] = None
    batch_size: Optional[int] = None
    registration_open: Optional[bool] = None
    next_batch_date: Optional[str] = None
    brand_id: Optional[str] = None
    brand_ids: Optional[List[str]] = None
    highlights: Optional[List[str]] = None  # Program highlights

# Program Unit Models
class UnitCreate(BaseModel):
    program_id: str
    name: str
    description: Optional[str] = ""
    brand_id: str  # Which brand handles this unit
    duration_weeks: int = 1
    order: int = 1
    theory_hours: int = 0
    practical_hours: int = 0
    assessments_required: List[Dict[str, Any]] = []  # [{name, type, passing_score}]

class UnitUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    brand_id: Optional[str] = None
    duration_weeks: Optional[int] = None
    order: Optional[int] = None
    theory_hours: Optional[int] = None
    practical_hours: Optional[int] = None
    assessments_required: Optional[List[Dict[str, Any]]] = None

# Student Unit Progress Models
class UnitProgressUpdate(BaseModel):
    status: str  # not_started, in_progress, completed
    completion_date: Optional[str] = None
    score: Optional[float] = None
    notes: Optional[str] = None

class BatchCreate(BaseModel):
    program_id: str
    trainer_id: str
    start_date: str
    schedule: Dict[str, Any] = {}

class AttendanceRecord(BaseModel):
    student_id: str
    batch_id: str
    nfc_card_id: str
    timestamp: Optional[str] = None
    gps_location: Optional[Dict[str, float]] = None
    session_type: str = "normal"  # normal, assessment
    offline_sync: bool = False

class AssessmentCreate(BaseModel):
    student_id: str
    batch_id: str
    trainer_id: str
    session_id: str
    skill_control: int = Field(ge=1, le=5)
    discipline: int = Field(ge=1, le=5)
    safety_awareness: int = Field(ge=1, le=5)
    execution: int = Field(ge=1, le=5)
    teamwork: int = Field(ge=1, le=5)
    notes: Optional[str] = None

class UpgradeApplication(BaseModel):
    student_id: str
    current_program_id: str
    target_program_id: str

# ======================== NFC ATTENDANCE SESSION MODELS ========================

class AttendanceSessionCreate(BaseModel):
    unit_id: str
    session_name: Optional[str] = None
    notes: Optional[str] = None

class NFCAttendanceRequest(BaseModel):
    nfc_card_id: str
    session_id: str

class AssessmentCategoryCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    assessment_type: str = "rating"  # rating, checkbox, text
    scale_min: int = 1
    scale_max: int = 5
    weight: float = 1.0  # Weight for calculating overall score
    is_required: bool = True  # Required for completing assessment
    is_active: bool = True
    brand_ids: List[str] = []  # Assign to multiple brands
    display_order: int = 0  # Order in assessment form

class NFCAssessmentCreate(BaseModel):
    student_id: str
    session_id: str
    unit_id: str
    ratings: Dict[str, int]  # {category_id: rating}
    notes: Optional[str] = None
    crew_nfc_confirmation: str  # Crew's NFC ID for confirmation

# ======================== PARTNERS & SPONSORS MODELS ========================

class PartnerCreate(BaseModel):
    name: str
    description: Optional[str] = None
    logo_url: Optional[str] = None
    logo_base64: Optional[str] = None
    website_url: Optional[str] = None
    partner_type: str = "partner"  # partner, sponsor, association
    is_visible: bool = True
    is_featured: bool = False  # Featured partners show as big logos
    order: int = 0

class ProgrammeDirectorUpdate(BaseModel):
    name: str
    designation: str
    message: str
    photo_url: Optional[str] = None
    photo_base64: Optional[str] = None

class ContactInfoUpdate(BaseModel):
    email: str
    phone: str
    whatsapp_number: str
    location_address: str
    location_maps_url: Optional[str] = None
    heading_text: str = "Questions? Please get in touch"
    subheading_text: str = "Our admission team will be happy to discuss your options"

class CallbackRequestCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    message: Optional[str] = None
    preferred_time: Optional[str] = None

class AdminCreate(BaseModel):
    email: EmailStr
    password: str
    name: str

# ======================== AUTH HELPERS ========================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, role: str) -> str:
    expire = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION)
    payload = {"user_id": user_id, "role": role, "exp": expire}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request):
    # Check cookie first
    session_token = request.cookies.get("session_token")
    if session_token:
        session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
        if session:
            expires_at = session.get("expires_at")
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at > datetime.now(timezone.utc):
                user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
                if user:
                    return user
    
    # Check Authorization header
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user = await db.users.find_one({"user_id": payload["user_id"]}, {"_id": 0})
            if user:
                return user
        except JWTError:
            pass
    
    raise HTTPException(status_code=401, detail="Not authenticated")

# ======================== OTP STORAGE ========================
otp_store: Dict[str, str] = {}

# ======================== AUTH ENDPOINTS ========================

@api_router.post("/auth/register")
async def register(data: UserCreate):
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    user_doc = {
        "user_id": user_id,
        "email": data.email,
        "name": data.name,
        "password_hash": hash_password(data.password),
        "role": data.role,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "picture": None
    }
    await db.users.insert_one(user_doc)
    
    token = create_token(user_id, data.role)
    return {"token": token, "user_id": user_id, "role": data.role, "name": data.name, "email": data.email}

@api_router.post("/auth/login")
async def login(data: UserLogin):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user or not verify_password(data.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["user_id"], user.get("role", "student"))
    return {"token": token, "user_id": user["user_id"], "role": user.get("role"), "name": user["name"], "email": user["email"]}

@api_router.post("/auth/session")
async def create_session(request: Request, response: Response):
    session_id = request.headers.get("X-Session-ID")
    if not session_id:
        raise HTTPException(status_code=400, detail="Session ID required")
    
    async with httpx.AsyncClient() as client:
        resp = await client.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": session_id}
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid session")
        
        oauth_data = resp.json()
    
    # Check if user exists
    user = await db.users.find_one({"email": oauth_data["email"]}, {"_id": 0})
    if not user:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user = {
            "user_id": user_id,
            "email": oauth_data["email"],
            "name": oauth_data["name"],
            "picture": oauth_data.get("picture"),
            "role": "student",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user)
    else:
        user_id = user["user_id"]
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {"name": oauth_data["name"], "picture": oauth_data.get("picture")}}
        )
    
    # Create session
    session_token = oauth_data.get("session_token", f"sess_{uuid.uuid4().hex}")
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7*24*60*60
    )
    
    user_response = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    return user_response

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    return {k: v for k, v in user.items() if k != "password_hash"}

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    response.delete_cookie("session_token", path="/")
    return {"message": "Logged out"}

# ======================== OTP ENDPOINTS ========================

@api_router.post("/otp/send")
async def send_otp(data: OTPRequest):
    use_mock = os.environ.get("USE_MOCK_OTP", "false").lower() == "true"
    
    if use_mock:
        otp = "123456"
        logger.info(f"MOCK OTP for {data.phone}: {otp}")
    else:
        otp = ''.join(random.choices(string.digits, k=6))
        
        # Send OTP via Twilio
        if twilio_client and TWILIO_PHONE_NUMBER:
            try:
                message = twilio_client.messages.create(
                    body=f"Your KXGRID verification code is: {otp}. Valid for 10 minutes.",
                    from_=TWILIO_PHONE_NUMBER,
                    to=data.phone
                )
                logger.info(f"OTP sent to {data.phone} via Twilio. SID: {message.sid}")
            except Exception as e:
                logger.error(f"Failed to send OTP via Twilio: {e}")
                raise HTTPException(status_code=500, detail="Failed to send OTP. Please try again.")
        else:
            logger.warning(f"Twilio not configured. OTP for {data.phone}: {otp}")
    
    otp_store[data.phone] = otp
    
    return {"status": "sent", "message": "OTP sent successfully"}

@api_router.post("/otp/verify")
async def verify_otp(data: OTPVerify):
    stored_otp = otp_store.get(data.phone)
    if not stored_otp or stored_otp != data.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP")
    
    del otp_store[data.phone]
    return {"valid": True, "message": "OTP verified"}

# ======================== NFC & MOBILE LOGIN ========================

class NFCLogin(BaseModel):
    nfc_card_id: str

class MobileLogin(BaseModel):
    mobile: str

@api_router.post("/auth/nfc-login")
async def nfc_login(data: NFCLogin):
    """Login using NFC Card ID - works for students, crew, and admin"""
    nfc_id = data.nfc_card_id.upper()
    
    # Check in students
    student = await db.students.find_one({"nfc_card_id": nfc_id, "status": "active"}, {"_id": 0})
    if student:
        user = await db.users.find_one({"user_id": student["user_id"]}, {"_id": 0})
        if user:
            token = create_token(user["user_id"], user.get("role", "student"))
            return {"token": token, "user_id": user["user_id"], "role": user.get("role"), "name": student.get("full_name", user.get("name")), "nfc_card_id": nfc_id}
    
    # Check in crew/admin NFC records
    nfc_record = await db.nfc_cards.find_one({"nfc_card_id": nfc_id, "status": "active"}, {"_id": 0})
    if nfc_record:
        user = await db.users.find_one({"user_id": nfc_record["user_id"]}, {"_id": 0})
        if user:
            token = create_token(user["user_id"], user.get("role", "student"))
            return {"token": token, "user_id": user["user_id"], "role": user.get("role"), "name": user.get("name"), "nfc_card_id": nfc_id}
    
    raise HTTPException(status_code=404, detail="NFC Card not found or inactive")

# ======================== NFC LOGIN WITH PASSWORD ========================

class NFCPasswordLogin(BaseModel):
    nfc_card_id: str
    password: str

class PasswordResetRequest(BaseModel):
    mobile: str

class PasswordResetVerify(BaseModel):
    mobile: str
    otp: str
    new_password: str

class AdminPasswordReset(BaseModel):
    user_id: str
    new_password: str

DEFAULT_NFC_PASSWORD = "NFC1234"

@api_router.post("/auth/nfc-password-login")
async def nfc_password_login(data: NFCPasswordLogin):
    """Login using NFC ID + Password"""
    nfc_id = data.nfc_card_id.upper().strip()
    
    # Check in nfc_users collection (bulk uploaded users)
    nfc_user = await db.nfc_users.find_one({"nfc_id": nfc_id, "is_active": True}, {"_id": 0})
    
    if not nfc_user:
        raise HTTPException(status_code=404, detail="NFC ID not found or not activated")
    
    # Verify password
    stored_password = nfc_user.get("password", DEFAULT_NFC_PASSWORD)
    
    # Check if password is hashed or plain
    if stored_password.startswith("$2"):
        # Hashed password
        if not bcrypt.checkpw(data.password.encode('utf-8'), stored_password.encode('utf-8')):
            raise HTTPException(status_code=401, detail="Invalid password")
    else:
        # Plain text password (default)
        if data.password != stored_password:
            raise HTTPException(status_code=401, detail="Invalid password")
    
    # Check if user account exists, if not create one
    user = await db.users.find_one({"nfc_card_id": nfc_id}, {"_id": 0, "password": 0})
    
    if not user:
        # Create user account from NFC user data
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user_doc = {
            "user_id": user_id,
            "name": nfc_user.get("name", "NFC User"),
            "email": nfc_user.get("email"),
            "mobile": nfc_user.get("mobile"),
            "role": nfc_user.get("role", "student"),
            "nfc_card_id": nfc_id,
            "password_changed": nfc_user.get("password_changed", False),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user_doc)
        user = {k: v for k, v in user_doc.items() if k not in ["_id", "password"]}
        
        # Also create student profile if role is student
        if user_doc["role"] == "student":
            student_doc = {
                "student_id": f"std_{uuid.uuid4().hex[:12]}",
                "user_id": user_id,
                "full_name": nfc_user.get("name"),
                "mobile": nfc_user.get("mobile"),
                "email": nfc_user.get("email"),
                "nfc_card_id": nfc_id,
                "status": "active",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.students.insert_one(student_doc)
    
    # Generate token
    token = create_token(user["user_id"], user.get("role", "student"))
    
    # Check if first login (password not changed)
    is_first_login = not nfc_user.get("password_changed", False)
    
    return {
        "token": token,
        "user_id": user["user_id"],
        "role": user.get("role", "student"),
        "name": user.get("name"),
        "nfc_card_id": nfc_id,
        "is_first_login": is_first_login,
        "show_password_reset_reminder": is_first_login
    }

@api_router.post("/auth/request-password-reset")
async def request_password_reset(data: PasswordResetRequest):
    """Request OTP for password reset"""
    mobile = data.mobile.strip()
    
    # Find user by mobile
    user = await db.users.find_one({"mobile": mobile}, {"_id": 0, "password": 0})
    if not user:
        # Check in students
        student = await db.students.find_one({"mobile": mobile}, {"_id": 0})
        if student:
            user = await db.users.find_one({"user_id": student["user_id"]}, {"_id": 0, "password": 0})
    
    if not user:
        # Check in nfc_users
        nfc_user = await db.nfc_users.find_one({"mobile": mobile}, {"_id": 0})
        if not nfc_user:
            raise HTTPException(status_code=404, detail="Mobile number not registered")
    
    # Generate OTP (mocked for now - will integrate Twilio later)
    otp = "123456"  # TODO: Replace with Twilio
    
    # Store OTP with expiry
    await db.otp_requests.update_one(
        {"mobile": mobile},
        {
            "$set": {
                "mobile": mobile,
                "otp": otp,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat(),
                "verified": False
            }
        },
        upsert=True
    )
    
    # TODO: Send SMS via Twilio
    logger.info(f"OTP for {mobile}: {otp} (MOCKED - will send via Twilio)")
    
    return {"message": "OTP sent to your mobile number", "mobile": mobile[-4:].rjust(len(mobile), '*')}

@api_router.post("/auth/verify-reset-password")
async def verify_and_reset_password(data: PasswordResetVerify):
    """Verify OTP and reset password"""
    mobile = data.mobile.strip()
    
    # Verify OTP
    otp_record = await db.otp_requests.find_one({"mobile": mobile}, {"_id": 0})
    if not otp_record:
        raise HTTPException(status_code=400, detail="No OTP request found. Please request a new OTP.")
    
    # Check expiry
    expires_at = datetime.fromisoformat(otp_record["expires_at"].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="OTP has expired. Please request a new one.")
    
    # Verify OTP
    if otp_record["otp"] != data.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP")
    
    # Hash new password
    hashed_password = bcrypt.hashpw(data.new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    # Update password in users collection
    user = await db.users.find_one({"mobile": mobile})
    if user:
        await db.users.update_one(
            {"mobile": mobile},
            {"$set": {"password": hashed_password, "password_changed": True}}
        )
    
    # Also update in nfc_users if exists
    nfc_user = await db.nfc_users.find_one({"mobile": mobile})
    if nfc_user:
        await db.nfc_users.update_one(
            {"mobile": mobile},
            {"$set": {"password": hashed_password, "password_changed": True}}
        )
    
    # Check students
    student = await db.students.find_one({"mobile": mobile})
    if student:
        await db.users.update_one(
            {"user_id": student["user_id"]},
            {"$set": {"password": hashed_password, "password_changed": True}}
        )
    
    # Mark OTP as used
    await db.otp_requests.delete_one({"mobile": mobile})
    
    return {"message": "Password reset successfully"}

@api_router.put("/auth/change-password")
async def change_password(new_password: str, user: dict = Depends(get_current_user)):
    """Change password for logged-in user"""
    hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    await db.users.update_one(
        {"user_id": user["user_id"]},
        {"$set": {"password": hashed_password, "password_changed": True}}
    )
    
    # Also update nfc_users if exists
    nfc_id = user.get("nfc_card_id")
    if nfc_id:
        await db.nfc_users.update_one(
            {"nfc_id": nfc_id},
            {"$set": {"password": hashed_password, "password_changed": True}}
        )
    
    return {"message": "Password changed successfully"}

# ======================== ADMIN NFC USER MANAGEMENT ========================

@api_router.get("/admin/nfc-users/template")
async def download_nfc_template(user: dict = Depends(get_current_user)):
    """Download Excel template for bulk NFC user upload"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "NFC Users"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["NFC_ID", "Student_Name", "Mobile", "Email", "Program", "Role"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Sample data row
    sample_data = ["NFC001", "John Doe", "+919876543210", "john@example.com", "Motorsport Certification", "student"]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
    
    # Instructions
    ws.cell(row=4, column=1, value="Instructions:").font = Font(bold=True)
    ws.cell(row=5, column=1, value="1. NFC_ID: Unique NFC card identifier (required)")
    ws.cell(row=6, column=1, value="2. Student_Name: Full name of the user (required)")
    ws.cell(row=7, column=1, value="3. Mobile: Mobile number with country code (required for password reset)")
    ws.cell(row=8, column=1, value="4. Email: Email address (optional)")
    ws.cell(row=9, column=1, value="5. Program: Program name (optional)")
    ws.cell(row=10, column=1, value="6. Role: student, trainer, brand_head, admin (default: student)")
    ws.cell(row=11, column=1, value="")
    ws.cell(row=12, column=1, value="Default password for all users: NFC1234").font = Font(bold=True, color="FF0000")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=NFC_Users_Template.xlsx"}
    )

@api_router.post("/admin/nfc-users/upload")
async def upload_nfc_users(file: UploadFile, user: dict = Depends(get_current_user)):
    """Bulk upload NFC users from Excel file"""
    from openpyxl import load_workbook
    from io import BytesIO
    
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")
    
    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    
    success_count = 0
    error_rows = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        nfc_id = str(row[0]).upper().strip() if row[0] else None
        name = str(row[1]).strip() if row[1] else None
        mobile = str(row[2]).strip() if row[2] else None
        email = str(row[3]).strip() if row[3] else None
        program = str(row[4]).strip() if row[4] else None
        role = str(row[5]).strip().lower() if row[5] else "student"
        
        if not nfc_id or not name:
            error_rows.append({"row": row_idx, "error": "NFC_ID and Student_Name are required"})
            continue
        
        # Validate role
        if role not in ["student", "trainer", "brand_head", "admin"]:
            role = "student"
        
        # Check if NFC ID already exists
        existing = await db.nfc_users.find_one({"nfc_id": nfc_id})
        
        nfc_user_doc = {
            "nfc_id": nfc_id,
            "name": name,
            "mobile": mobile,
            "email": email,
            "program": program,
            "role": role,
            "password": DEFAULT_NFC_PASSWORD,
            "password_changed": False,
            "is_active": True,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "uploaded_by": user["user_id"]
        }
        
        if existing:
            # Update existing
            await db.nfc_users.update_one({"nfc_id": nfc_id}, {"$set": nfc_user_doc})
        else:
            # Insert new
            await db.nfc_users.insert_one(nfc_user_doc)
        
        success_count += 1
    
    return {
        "message": f"Uploaded {success_count} NFC users successfully",
        "success_count": success_count,
        "error_count": len(error_rows),
        "errors": error_rows[:10]  # Return first 10 errors
    }

@api_router.get("/admin/nfc-users")
async def get_all_nfc_users(
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get all NFC users with optional search by NFC ID or name"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    query = {}
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        query = {"$or": [
            {"nfc_id": search_regex},
            {"name": search_regex},
            {"mobile": search_regex},
            {"email": search_regex}
        ]}
    
    nfc_users = await db.nfc_users.find(query, {"_id": 0}).sort("uploaded_at", -1).to_list(500)
    return nfc_users

@api_router.put("/admin/nfc-users/{nfc_id}/reset-password")
async def admin_reset_nfc_password(nfc_id: str, data: AdminPasswordReset = None, user: dict = Depends(get_current_user)):
    """Admin directly resets a user's password"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    nfc_id = nfc_id.upper().strip()
    
    # Find NFC user
    nfc_user = await db.nfc_users.find_one({"nfc_id": nfc_id})
    if not nfc_user:
        raise HTTPException(status_code=404, detail="NFC user not found")
    
    # Determine new password
    new_password = data.new_password if data and data.new_password else DEFAULT_NFC_PASSWORD
    
    # Hash if not default
    if new_password != DEFAULT_NFC_PASSWORD:
        hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    else:
        hashed_password = DEFAULT_NFC_PASSWORD
    
    # Update in nfc_users
    await db.nfc_users.update_one(
        {"nfc_id": nfc_id},
        {"$set": {
            "password": hashed_password,
            "password_changed": new_password != DEFAULT_NFC_PASSWORD,
            "password_reset_at": datetime.now(timezone.utc).isoformat(),
            "password_reset_by": user["user_id"]
        }}
    )
    
    # Also update in users if linked
    await db.users.update_one(
        {"nfc_card_id": nfc_id},
        {"$set": {
            "password": hashed_password,
            "password_changed": new_password != DEFAULT_NFC_PASSWORD
        }}
    )
    
    return {"message": f"Password reset to {'default (NFC1234)' if new_password == DEFAULT_NFC_PASSWORD else 'custom password'}"}

@api_router.put("/admin/nfc-users/{nfc_id}/toggle-status")
async def toggle_nfc_user_status(nfc_id: str, user: dict = Depends(get_current_user)):
    """Activate or deactivate an NFC user"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    nfc_id = nfc_id.upper().strip()
    
    nfc_user = await db.nfc_users.find_one({"nfc_id": nfc_id})
    if not nfc_user:
        raise HTTPException(status_code=404, detail="NFC user not found")
    
    new_status = not nfc_user.get("is_active", True)
    
    await db.nfc_users.update_one(
        {"nfc_id": nfc_id},
        {"$set": {"is_active": new_status}}
    )
    
    return {"message": f"NFC user {'activated' if new_status else 'deactivated'}", "is_active": new_status}

@api_router.post("/auth/mobile-login")
async def mobile_login(data: MobileLogin):
    """Login using verified mobile number"""
    mobile = data.mobile
    
    # Check in students
    student = await db.students.find_one({"mobile": mobile, "status": "active"}, {"_id": 0})
    if student:
        user = await db.users.find_one({"user_id": student["user_id"]}, {"_id": 0})
        if user:
            token = create_token(user["user_id"], user.get("role", "student"))
            return {"token": token, "user_id": user["user_id"], "role": user.get("role"), "name": student.get("full_name", user.get("name"))}
    
    # Check in users directly (for crew/admin who may have mobile linked)
    user = await db.users.find_one({"mobile": mobile}, {"_id": 0})
    if user:
        token = create_token(user["user_id"], user.get("role", "student"))
        return {"token": token, "user_id": user["user_id"], "role": user.get("role"), "name": user.get("name")}
    
    raise HTTPException(status_code=404, detail="Mobile number not registered")

# ======================== NFC MANAGEMENT (Admin) ========================

class NFCIssue(BaseModel):
    user_id: str
    nfc_card_id: str
    user_type: str  # student, trainer, admin

class NFCReplace(BaseModel):
    old_nfc_id: str
    new_nfc_id: str

@api_router.post("/admin/nfc/issue")
async def issue_nfc(data: NFCIssue, user: dict = Depends(get_current_user)):
    """Issue new NFC card to a user"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Check if NFC already exists
    existing = await db.nfc_cards.find_one({"nfc_card_id": data.nfc_card_id.upper()})
    if existing:
        raise HTTPException(status_code=400, detail="NFC Card ID already in use")
    
    nfc_doc = {
        "nfc_card_id": data.nfc_card_id.upper(),
        "user_id": data.user_id,
        "user_type": data.user_type,
        "status": "active",
        "issued_at": datetime.now(timezone.utc).isoformat(),
        "issued_by": user["user_id"]
    }
    
    await db.nfc_cards.insert_one(nfc_doc)
    
    # If student, update their NFC
    if data.user_type == "student":
        await db.students.update_one(
            {"user_id": data.user_id},
            {"$set": {"nfc_card_id": data.nfc_card_id.upper()}}
        )
    
    # Update user record with mobile if needed
    await db.users.update_one(
        {"user_id": data.user_id},
        {"$set": {"nfc_card_id": data.nfc_card_id.upper()}}
    )
    
    return {"message": "NFC Card issued", "nfc_card_id": data.nfc_card_id.upper()}

@api_router.post("/admin/nfc/replace")
async def replace_nfc(data: NFCReplace, user: dict = Depends(get_current_user)):
    """Replace old NFC with new one - transfers all data"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    old_nfc = data.old_nfc_id.upper()
    new_nfc = data.new_nfc_id.upper()
    
    # Check if new NFC already exists
    existing = await db.nfc_cards.find_one({"nfc_card_id": new_nfc})
    if existing:
        raise HTTPException(status_code=400, detail="New NFC Card ID already in use")
    
    # Find old NFC record
    old_record = await db.nfc_cards.find_one({"nfc_card_id": old_nfc})
    student_record = await db.students.find_one({"nfc_card_id": old_nfc})
    
    if not old_record and not student_record:
        raise HTTPException(status_code=404, detail="Old NFC Card not found")
    
    user_id = old_record["user_id"] if old_record else student_record["user_id"]
    user_type = old_record["user_type"] if old_record else "student"
    
    # Create new NFC record
    new_nfc_doc = {
        "nfc_card_id": new_nfc,
        "user_id": user_id,
        "user_type": user_type,
        "status": "active",
        "issued_at": datetime.now(timezone.utc).isoformat(),
        "issued_by": user["user_id"],
        "replaced_from": old_nfc
    }
    await db.nfc_cards.insert_one(new_nfc_doc)
    
    # Mark old NFC as pending deletion (10 days)
    deletion_date = datetime.now(timezone.utc) + timedelta(days=10)
    
    if old_record:
        await db.nfc_cards.update_one(
            {"nfc_card_id": old_nfc},
            {"$set": {
                "status": "pending_deletion",
                "replaced_by": new_nfc,
                "deletion_scheduled_at": deletion_date.isoformat(),
                "replaced_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # Update student record
    if student_record:
        await db.students.update_one(
            {"nfc_card_id": old_nfc},
            {"$set": {"nfc_card_id": new_nfc}}
        )
    
    # Update user record
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"nfc_card_id": new_nfc}}
    )
    
    # Create pending deletion record
    await db.pending_nfc_deletions.insert_one({
        "old_nfc_id": old_nfc,
        "new_nfc_id": new_nfc,
        "user_id": user_id,
        "replaced_at": datetime.now(timezone.utc).isoformat(),
        "deletion_due": deletion_date.isoformat(),
        "status": "pending",
        "verified": False
    })
    
    return {
        "message": "NFC replaced successfully",
        "old_nfc": old_nfc,
        "new_nfc": new_nfc,
        "deletion_scheduled": deletion_date.isoformat()
    }

@api_router.get("/admin/nfc/all")
async def get_all_nfc_cards(user: dict = Depends(get_current_user)):
    """Get all NFC cards with user info"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Get all NFC cards from nfc_cards collection
    nfc_cards = await db.nfc_cards.find({"status": "active"}, {"_id": 0}).to_list(500)
    
    # Get all students with NFC cards
    students_with_nfc = await db.students.find(
        {"nfc_card_id": {"$exists": True, "$ne": None}},
        {"_id": 0, "user_id": 1, "full_name": 1, "nfc_card_id": 1, "mobile": 1, "email": 1}
    ).to_list(500)
    
    # Get all users for reference
    user_ids = set([n.get("user_id") for n in nfc_cards])
    user_ids.update([s.get("user_id") for s in students_with_nfc])
    
    users = await db.users.find(
        {"user_id": {"$in": list(user_ids)}},
        {"_id": 0, "user_id": 1, "name": 1, "email": 1, "role": 1, "nfc_card_id": 1}
    ).to_list(500)
    users_map = {u["user_id"]: u for u in users}
    
    # Build combined list
    result = []
    seen_nfc = set()
    
    # Add from students (most common)
    for s in students_with_nfc:
        nfc_id = s.get("nfc_card_id")
        if nfc_id and nfc_id not in seen_nfc:
            seen_nfc.add(nfc_id)
            user_info = users_map.get(s.get("user_id"), {})
            result.append({
                "nfc_card_id": nfc_id,
                "user_id": s.get("user_id"),
                "name": s.get("full_name") or user_info.get("name"),
                "email": s.get("email") or user_info.get("email"),
                "mobile": s.get("mobile"),
                "user_type": "student",
                "status": "active"
            })
    
    # Add from nfc_cards collection (crew/admin)
    for n in nfc_cards:
        nfc_id = n.get("nfc_card_id")
        if nfc_id and nfc_id not in seen_nfc:
            seen_nfc.add(nfc_id)
            user_info = users_map.get(n.get("user_id"), {})
            result.append({
                "nfc_card_id": nfc_id,
                "user_id": n.get("user_id"),
                "name": user_info.get("name"),
                "email": user_info.get("email"),
                "user_type": n.get("user_type", "unknown"),
                "status": n.get("status", "active"),
                "issued_at": n.get("issued_at")
            })
    
    return result

@api_router.delete("/admin/nfc/{nfc_card_id}")
async def revoke_nfc_card(nfc_card_id: str, user: dict = Depends(get_current_user)):
    """Revoke/deactivate an NFC card"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    nfc_id = nfc_card_id.upper()
    
    # Update nfc_cards collection
    await db.nfc_cards.update_one(
        {"nfc_card_id": nfc_id},
        {"$set": {"status": "revoked", "revoked_at": datetime.now(timezone.utc).isoformat(), "revoked_by": user["user_id"]}}
    )
    
    # Clear from student record
    await db.students.update_one(
        {"nfc_card_id": nfc_id},
        {"$unset": {"nfc_card_id": ""}}
    )
    
    # Clear from user record
    await db.users.update_one(
        {"nfc_card_id": nfc_id},
        {"$unset": {"nfc_card_id": ""}}
    )
    
    return {"message": "NFC Card revoked"}

@api_router.get("/admin/nfc/pending-deletions")
async def get_pending_deletions(user: dict = Depends(get_current_user)):
    """Get all NFCs pending deletion"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    pending = await db.pending_nfc_deletions.find(
        {"status": "pending"},
        {"_id": 0}
    ).to_list(100)
    
    # Check which ones are due
    now = datetime.now(timezone.utc)
    for p in pending:
        due_date = datetime.fromisoformat(p["deletion_due"].replace("Z", "+00:00"))
        p["is_due"] = now >= due_date
        p["days_remaining"] = max(0, (due_date - now).days)
    
    return pending

@api_router.post("/admin/nfc/confirm-deletion/{old_nfc_id}")
async def confirm_nfc_deletion(old_nfc_id: str, user: dict = Depends(get_current_user)):
    """Admin confirms and permanently deletes old NFC"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    old_nfc = old_nfc_id.upper()
    
    # Update pending deletion record
    result = await db.pending_nfc_deletions.update_one(
        {"old_nfc_id": old_nfc, "status": "pending"},
        {"$set": {
            "status": "deleted",
            "verified": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": user["user_id"]
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pending deletion not found")
    
    # Delete old NFC record
    await db.nfc_cards.delete_one({"nfc_card_id": old_nfc})
    
    return {"message": f"NFC {old_nfc} permanently deleted"}

# ======================== STUDENT ENDPOINTS ========================

@api_router.post("/students/register")
async def register_student(data: StudentRegistration, user: dict = Depends(get_current_user)):
    student_id = f"std_{uuid.uuid4().hex[:12]}"
    nfc_card_id = data.nfc_card_id or f"NFC_{uuid.uuid4().hex[:8].upper()}"
    
    student_doc = {
        "student_id": student_id,
        "user_id": data.user_id,
        **data.model_dump(exclude={"user_id", "nfc_card_id"}),
        "nfc_card_id": nfc_card_id,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "badges": [],
        "total_attendance": 0,
        "average_rating": 0.0
    }
    
    await db.students.insert_one(student_doc)
    await db.users.update_one({"user_id": data.user_id}, {"$set": {"student_id": student_id, "registration_complete": True}})
    
    return {k: v for k, v in student_doc.items() if k != "_id"}

@api_router.get("/students/profile")
async def get_student_profile(user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not student:
        return {"registered": False}
    return {"registered": True, **student}

@api_router.put("/students/{student_id}/nfc")
async def assign_nfc_card(student_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Assign or update NFC card for a student (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    nfc_card_id = data.get("nfc_card_id", "").upper()
    if not nfc_card_id:
        raise HTTPException(status_code=400, detail="NFC Card ID required")
    
    # Check if NFC already assigned to another student
    existing = await db.students.find_one({"nfc_card_id": nfc_card_id, "student_id": {"$ne": student_id}})
    if existing:
        raise HTTPException(status_code=400, detail=f"NFC card already assigned to {existing.get('full_name')}")
    
    result = await db.students.update_one(
        {"student_id": student_id},
        {"$set": {"nfc_card_id": nfc_card_id}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
    
    return {"message": f"NFC card {nfc_card_id} assigned", "nfc_card_id": nfc_card_id}

@api_router.get("/students/nfc/{nfc_id}")
async def get_student_by_nfc(nfc_id: str):
    """Public endpoint for NFC ID card view"""
    student = await db.students.find_one({"nfc_card_id": nfc_id.upper()}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student

@api_router.get("/students/{student_id}")
async def get_student(student_id: str, user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student

@api_router.get("/students")
async def list_students(user: dict = Depends(get_current_user)):
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    return students

# ======================== STUDENT UNIT PROGRESS ========================

@api_router.get("/student/progress")
async def get_student_progress(user: dict = Depends(get_current_user)):
    """Get current student's unit progress across all programs"""
    student = await db.students.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student profile not found")
    
    student_id = student["student_id"]
    
    # Get all programs
    programs = await db.programs.find({"status": "active"}, {"_id": 0}).to_list(100)
    
    # Get all progress records for this student
    progress_records = await db.student_progress.find(
        {"student_id": student_id},
        {"_id": 0}
    ).to_list(500)
    
    # Create a lookup dict for quick access
    progress_lookup = {p["unit_id"]: p for p in progress_records}
    
    result = []
    for program in programs:
        # Get units for this program
        units = await db.program_units.find(
            {"program_id": program["program_id"]},
            {"_id": 0}
        ).sort("order", 1).to_list(50)
        
        if not units:
            continue
        
        # Enrich units with progress data
        units_with_progress = []
        completed_count = 0
        
        # Batch fetch brands to avoid N+1 queries
        brand_ids = list(set([u.get("brand_id") for u in units if u.get("brand_id")]))
        brands_list = await db.brands.find({"brand_id": {"$in": brand_ids}}, {"_id": 0, "brand_id": 1, "name": 1, "color": 1}).to_list(100) if brand_ids else []
        brands_map = {b["brand_id"]: b for b in brands_list}
        
        for unit in units:
            unit_progress = progress_lookup.get(unit["unit_id"], {
                "status": "not_started",
                "completion_date": None,
                "score": None
            })
            unit["progress"] = unit_progress
            if unit_progress.get("status") == "completed":
                completed_count += 1
            
            brand = brands_map.get(unit.get("brand_id"), {})
            unit["brand_name"] = brand.get("name", "Unknown")
            unit["brand_color"] = brand.get("color", "#00f0ff")
            
            units_with_progress.append(unit)
        
        total_units = len(units)
        progress_percent = round((completed_count / total_units) * 100) if total_units > 0 else 0
        
        result.append({
            "program_id": program["program_id"],
            "program_name": program["name"],
            "program_type": program.get("program_type", "certification"),
            "total_units": total_units,
            "completed_units": completed_count,
            "progress_percent": progress_percent,
            "units": units_with_progress
        })
    
    return result

@api_router.get("/student/progress/{program_id}")
async def get_student_program_progress(program_id: str, user: dict = Depends(get_current_user)):
    """Get student's progress for a specific program"""
    student = await db.students.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student profile not found")
    
    student_id = student["student_id"]
    
    # Get program
    program = await db.programs.find_one({"program_id": program_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    # Get units
    units = await db.program_units.find(
        {"program_id": program_id},
        {"_id": 0}
    ).sort("order", 1).to_list(50)
    
    # Get progress records
    progress_records = await db.student_progress.find(
        {"student_id": student_id, "program_id": program_id},
        {"_id": 0}
    ).to_list(50)
    progress_lookup = {p["unit_id"]: p for p in progress_records}
    
    units_with_progress = []
    completed_count = 0
    
    # Batch fetch brands
    brand_ids = list(set([u.get("brand_id") for u in units if u.get("brand_id")]))
    brands_list = await db.brands.find({"brand_id": {"$in": brand_ids}}, {"_id": 0, "brand_id": 1, "name": 1, "color": 1}).to_list(100) if brand_ids else []
    brands_map = {b["brand_id"]: b for b in brands_list}
    
    for unit in units:
        unit_progress = progress_lookup.get(unit["unit_id"], {
            "status": "not_started",
            "completion_date": None,
            "score": None
        })
        unit["progress"] = unit_progress
        if unit_progress.get("status") == "completed":
            completed_count += 1
        
        brand = brands_map.get(unit.get("brand_id"), {})
        unit["brand_name"] = brand.get("name", "Unknown")
        unit["brand_color"] = brand.get("color", "#00f0ff")
        
        units_with_progress.append(unit)
    
    return {
        "program": program,
        "total_units": len(units),
        "completed_units": completed_count,
        "progress_percent": round((completed_count / len(units)) * 100) if units else 0,
        "units": units_with_progress
    }

@api_router.put("/student/progress/{unit_id}")
async def update_student_unit_progress(
    unit_id: str, 
    data: UnitProgressUpdate, 
    user: dict = Depends(get_current_user)
):
    """Update unit progress (admin/crew can update for students, students can mark in_progress)"""
    # Get the unit to find program_id
    unit = await db.program_units.find_one({"unit_id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    
    # Determine whose progress to update
    if user.get("role") in ["admin", "trainer"]:
        # Admin/trainer updating a student's progress - need student_id in query params
        # For now, this endpoint is for self-update; admin uses different endpoint
        student = await db.students.find_one({"user_id": user["user_id"]}, {"_id": 0})
    else:
        student = await db.students.find_one({"user_id": user["user_id"]}, {"_id": 0})
    
    if not student:
        raise HTTPException(status_code=404, detail="Student profile not found")
    
    student_id = student["student_id"]
    
    # Students can only mark as in_progress, not completed (that requires admin/crew)
    if user.get("role") == "student" and data.status == "completed":
        raise HTTPException(status_code=403, detail="Only trainers/admin can mark units as completed")
    
    progress_data = {
        "student_id": student_id,
        "unit_id": unit_id,
        "program_id": unit["program_id"],
        "status": data.status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": user["user_id"]
    }
    
    if data.completion_date:
        progress_data["completion_date"] = data.completion_date
    if data.score is not None:
        progress_data["score"] = data.score
    if data.notes:
        progress_data["notes"] = data.notes
    
    await db.student_progress.update_one(
        {"student_id": student_id, "unit_id": unit_id},
        {"$set": progress_data},
        upsert=True
    )
    
    return {"message": "Progress updated", "status": data.status}

@api_router.put("/admin/student/{student_id}/progress/{unit_id}")
async def admin_update_student_progress(
    student_id: str,
    unit_id: str,
    data: UnitProgressUpdate,
    user: dict = Depends(get_current_user)
):
    """Admin/Crew update a student's unit progress"""
    if user.get("role") not in ["admin", "trainer"]:
        raise HTTPException(status_code=403, detail="Admin or trainer only")
    
    # Verify student exists
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Get the unit
    unit = await db.program_units.find_one({"unit_id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    
    progress_data = {
        "student_id": student_id,
        "unit_id": unit_id,
        "program_id": unit["program_id"],
        "status": data.status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": user["user_id"]
    }
    
    if data.status == "completed" and not data.completion_date:
        progress_data["completion_date"] = datetime.now(timezone.utc).isoformat()
    elif data.completion_date:
        progress_data["completion_date"] = data.completion_date
    
    if data.score is not None:
        progress_data["score"] = data.score
    if data.notes:
        progress_data["notes"] = data.notes
    
    await db.student_progress.update_one(
        {"student_id": student_id, "unit_id": unit_id},
        {"$set": progress_data},
        upsert=True
    )
    
    return {"message": "Progress updated", "status": data.status}

@api_router.get("/admin/progress/program/{program_id}")
async def get_program_progress_overview(program_id: str, user: dict = Depends(get_current_user)):
    """Get progress overview for all students in a program (admin/crew)"""
    if user.get("role") not in ["admin", "trainer", "brand_head"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get all units for this program
    units = await db.program_units.find(
        {"program_id": program_id},
        {"_id": 0}
    ).sort("order", 1).to_list(50)
    
    if not units:
        return {"program_id": program_id, "units": [], "students": []}
    
    # Get all students
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    
    # Get all progress records for this program
    progress_records = await db.student_progress.find(
        {"program_id": program_id},
        {"_id": 0}
    ).to_list(5000)
    
    # Group by student
    student_progress = {}
    for record in progress_records:
        sid = record["student_id"]
        if sid not in student_progress:
            student_progress[sid] = {}
        student_progress[sid][record["unit_id"]] = record
    
    # Build result
    result_students = []
    for student in students:
        sid = student["student_id"]
        prog = student_progress.get(sid, {})
        
        completed = sum(1 for u in units if prog.get(u["unit_id"], {}).get("status") == "completed")
        in_progress = sum(1 for u in units if prog.get(u["unit_id"], {}).get("status") == "in_progress")
        
        result_students.append({
            "student_id": sid,
            "student_name": student.get("full_name", "Unknown"),
            "completed_units": completed,
            "in_progress_units": in_progress,
            "total_units": len(units),
            "progress_percent": round((completed / len(units)) * 100) if units else 0,
            "unit_status": {u["unit_id"]: prog.get(u["unit_id"], {"status": "not_started"}) for u in units}
        })
    
    return {
        "program_id": program_id,
        "units": units,
        "students": result_students
    }

# ======================== PROGRAM ENDPOINTS ========================

@api_router.post("/programs")
async def create_program(data: ProgramCreate, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    program_id = f"prog_{uuid.uuid4().hex[:12]}"
    program_doc = {
        "program_id": program_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "active",
        "total_enrolled": 0,
        "registration_open": data.registration_open,
        "next_batch_date": data.next_batch_date
    }
    await db.programs.insert_one(program_doc)
    return {k: v for k, v in program_doc.items() if k != "_id"}

@api_router.get("/programs")
async def list_programs():
    programs = await db.programs.find({}, {"_id": 0}).to_list(100)
    return programs

@api_router.get("/programs/{program_id}")
async def get_program(program_id: str):
    program = await db.programs.find_one({"program_id": program_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    return program

@api_router.put("/programs/{program_id}")
async def update_program(program_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update program details including registration status"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Remove protected fields
    data.pop("program_id", None)
    data.pop("created_at", None)
    
    result = await db.programs.update_one(
        {"program_id": program_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Program not found")
    
    return {"message": "Program updated"}

class RegistrationToggle(BaseModel):
    registration_open: bool
    next_batch_date: Optional[str] = None

@api_router.put("/programs/{program_id}/registration")
async def toggle_registration(program_id: str, data: RegistrationToggle, user: dict = Depends(get_current_user)):
    """Toggle program registration open/closed"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    update_data = {"registration_open": data.registration_open}
    if data.next_batch_date:
        update_data["next_batch_date"] = data.next_batch_date
    
    result = await db.programs.update_one(
        {"program_id": program_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Program not found")
    
    status = "opened" if data.registration_open else "closed"
    return {"message": f"Registration {status}", "next_batch_date": data.next_batch_date}

@api_router.delete("/programs/{program_id}")
async def delete_program(program_id: str, user: dict = Depends(get_current_user)):
    """Delete a program (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.programs.delete_one({"program_id": program_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Program not found")
    
    return {"message": "Program deleted"}

# ======================== PROGRAM UNITS ========================

@api_router.post("/programs/{program_id}/units")
async def create_unit(program_id: str, data: UnitCreate, user: dict = Depends(get_current_user)):
    """Create a new unit for a program (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Verify program exists
    program = await db.programs.find_one({"program_id": program_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    # Verify brand exists
    brand = await db.brands.find_one({"brand_id": data.brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    unit_id = f"unit_{uuid.uuid4().hex[:12]}"
    unit_doc = {
        "unit_id": unit_id,
        "program_id": program_id,
        **data.model_dump(),
        "brand_name": brand.get("name"),
        "brand_color": brand.get("color"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.program_units.insert_one(unit_doc)
    
    # Update program's brand_ids to include this brand if not already
    await db.programs.update_one(
        {"program_id": program_id},
        {"$addToSet": {"brand_ids": data.brand_id}}
    )
    
    return {k: v for k, v in unit_doc.items() if k != "_id"}

@api_router.get("/programs/{program_id}/units")
async def get_program_units(program_id: str):
    """Get all units for a program"""
    units = await db.program_units.find(
        {"program_id": program_id},
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    return units

@api_router.get("/units/{unit_id}")
async def get_unit(unit_id: str):
    """Get a specific unit"""
    unit = await db.program_units.find_one({"unit_id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    return unit

@api_router.put("/units/{unit_id}")
async def update_unit(unit_id: str, data: UnitUpdate, user: dict = Depends(get_current_user)):
    """Update a unit (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    # If brand_id is being updated, get brand info
    if "brand_id" in update_data:
        brand = await db.brands.find_one({"brand_id": update_data["brand_id"]}, {"_id": 0})
        if brand:
            update_data["brand_name"] = brand.get("name")
            update_data["brand_color"] = brand.get("color")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = user["user_id"]
    
    result = await db.program_units.update_one(
        {"unit_id": unit_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Unit not found")
    
    return {"message": "Unit updated"}

@api_router.delete("/units/{unit_id}")
async def delete_unit(unit_id: str, user: dict = Depends(get_current_user)):
    """Delete a unit (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.program_units.delete_one({"unit_id": unit_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Unit not found")
    
    return {"message": "Unit deleted"}

@api_router.post("/units/reorder")
async def reorder_units(data: dict, user: dict = Depends(get_current_user)):
    """Reorder units within a program (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    unit_orders = data.get("unit_orders", [])
    for item in unit_orders:
        await db.program_units.update_one(
            {"unit_id": item["unit_id"]},
            {"$set": {"order": item["order"]}}
        )
    
    return {"message": "Units reordered"}

# ======================== BATCH ENDPOINTS ========================

@api_router.post("/batches")
async def create_batch(data: BatchCreate, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "trainer"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    batch_id = f"batch_{uuid.uuid4().hex[:12]}"
    batch_doc = {
        "batch_id": batch_id,
        **data.model_dump(),
        "students": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "active"
    }
    await db.batches.insert_one(batch_doc)
    return {k: v for k, v in batch_doc.items() if k != "_id"}

@api_router.get("/batches")
async def list_batches(user: dict = Depends(get_current_user)):
    query = {}
    if user.get("role") == "trainer":
        query["trainer_id"] = user["user_id"]
    batches = await db.batches.find(query, {"_id": 0}).to_list(100)
    return batches

@api_router.post("/batches/{batch_id}/enroll/{student_id}")
async def enroll_student(batch_id: str, student_id: str, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "trainer"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.batches.update_one(
        {"batch_id": batch_id},
        {"$addToSet": {"students": student_id}}
    )
    await db.students.update_one(
        {"student_id": student_id},
        {"$set": {"batch_id": batch_id}}
    )
    return {"message": "Student enrolled"}

# ======================== ATTENDANCE ENDPOINTS (NFC SIMULATION) ========================

@api_router.post("/attendance/mark")
async def mark_attendance(data: AttendanceRecord, user: dict = Depends(get_current_user)):
    # Verify NFC card belongs to student
    student = await db.students.find_one({"nfc_card_id": data.nfc_card_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Invalid NFC card")
    
    attendance_id = f"att_{uuid.uuid4().hex[:12]}"
    timestamp = data.timestamp or datetime.now(timezone.utc).isoformat()
    
    attendance_doc = {
        "attendance_id": attendance_id,
        "student_id": student["student_id"],
        "batch_id": data.batch_id,
        "nfc_card_id": data.nfc_card_id,
        "timestamp": timestamp,
        "gps_location": data.gps_location,
        "session_type": data.session_type,
        "marked_by": user["user_id"],
        "offline_sync": data.offline_sync,
        "synced_at": datetime.now(timezone.utc).isoformat() if data.offline_sync else None
    }
    
    await db.attendance.insert_one(attendance_doc)
    await db.students.update_one(
        {"student_id": student["student_id"]},
        {"$inc": {"total_attendance": 1}}
    )
    
    return {k: v for k, v in attendance_doc.items() if k != "_id"}

@api_router.post("/attendance/sync")
async def sync_offline_attendance(records: List[AttendanceRecord], user: dict = Depends(get_current_user)):
    synced = []
    for record in records:
        record.offline_sync = True
        result = await mark_attendance(record, user)
        synced.append(result)
    return {"synced": len(synced), "records": synced}

@api_router.get("/attendance/student/{student_id}")
async def get_student_attendance(student_id: str, user: dict = Depends(get_current_user)):
    records = await db.attendance.find({"student_id": student_id}, {"_id": 0}).to_list(500)
    return records

@api_router.get("/attendance/batch/{batch_id}")
async def get_batch_attendance(batch_id: str, user: dict = Depends(get_current_user)):
    records = await db.attendance.find({"batch_id": batch_id}, {"_id": 0}).to_list(1000)
    return records

# ======================== NFC ATTENDANCE SESSION ENDPOINTS ========================

@api_router.post("/crew/attendance-session/start")
async def start_attendance_session(data: AttendanceSessionCreate, user: dict = Depends(get_current_user)):
    """Crew starts a new attendance session for a specific unit"""
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Crew/Admin only")
    
    # Verify unit exists
    unit = await db.program_units.find_one({"unit_id": data.unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    
    # Get brand info
    brand = await db.brands.find_one({"brand_id": unit.get("brand_id")}, {"_id": 0})
    
    session_id = f"sess_{uuid.uuid4().hex[:12]}"
    session_doc = {
        "session_id": session_id,
        "unit_id": data.unit_id,
        "unit_name": unit.get("name"),
        "program_id": unit.get("program_id"),
        "brand_id": unit.get("brand_id"),
        "brand_name": brand.get("name") if brand else "Unknown",
        "session_name": data.session_name or f"Session - {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}",
        "notes": data.notes,
        "crew_id": user["user_id"],
        "crew_name": user.get("name"),
        "status": "active",
        "started_at": datetime.now(timezone.utc).isoformat(),
        "ended_at": None,
        "attendance_count": 0,
        "attendees": []
    }
    
    await db.attendance_sessions.insert_one(session_doc)
    return {k: v for k, v in session_doc.items() if k != "_id"}

@api_router.post("/crew/attendance-session/{session_id}/nfc-tap")
async def record_nfc_attendance(session_id: str, data: NFCAttendanceRequest, user: dict = Depends(get_current_user)):
    """Record attendance when student taps NFC card"""
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Crew/Admin only")
    
    # Verify session exists and is active
    session = await db.attendance_sessions.find_one({"session_id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.get("status") != "active":
        raise HTTPException(status_code=400, detail="Session is not active")
    
    nfc_id = data.nfc_card_id.upper().strip()
    
    # Find student by NFC ID
    student = await db.students.find_one({"nfc_card_id": nfc_id, "status": "active"}, {"_id": 0})
    
    if not student:
        # Also check nfc_users collection
        nfc_user = await db.nfc_users.find_one({"nfc_id": nfc_id, "is_active": True}, {"_id": 0})
        if nfc_user:
            # Find or create student profile
            user_record = await db.users.find_one({"nfc_card_id": nfc_id}, {"_id": 0})
            if user_record:
                student = await db.students.find_one({"user_id": user_record["user_id"]}, {"_id": 0})
        
        if not student:
            raise HTTPException(status_code=404, detail=f"No student found with NFC ID: {nfc_id}")
    
    # Check if already marked attendance in this session
    existing = await db.session_attendance.find_one({
        "session_id": session_id,
        "student_id": student["student_id"]
    })
    if existing:
        return {
            "status": "already_marked",
            "message": f"{student.get('full_name', 'Student')} already marked attendance",
            "student": {
                "student_id": student["student_id"],
                "full_name": student.get("full_name"),
                "nfc_card_id": nfc_id
            },
            "marked_at": existing.get("marked_at")
        }
    
    # Record attendance
    attendance_id = f"satt_{uuid.uuid4().hex[:12]}"
    attendance_doc = {
        "attendance_id": attendance_id,
        "session_id": session_id,
        "unit_id": session.get("unit_id"),
        "program_id": session.get("program_id"),
        "student_id": student["student_id"],
        "student_name": student.get("full_name"),
        "nfc_card_id": nfc_id,
        "marked_by": user["user_id"],
        "marked_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.session_attendance.insert_one(attendance_doc)
    
    # Update session attendance count
    await db.attendance_sessions.update_one(
        {"session_id": session_id},
        {
            "$inc": {"attendance_count": 1},
            "$push": {"attendees": student["student_id"]}
        }
    )
    
    # Update student's total attendance
    await db.students.update_one(
        {"student_id": student["student_id"]},
        {"$inc": {"total_attendance": 1}}
    )
    
    return {
        "status": "success",
        "message": f"Attendance marked for {student.get('full_name', 'Student')}",
        "student": {
            "student_id": student["student_id"],
            "full_name": student.get("full_name"),
            "nfc_card_id": nfc_id,
            "photo_url": student.get("photo_url")
        },
        "attendance_id": attendance_id,
        "marked_at": attendance_doc["marked_at"]
    }

@api_router.put("/crew/attendance-session/{session_id}/end")
async def end_attendance_session(session_id: str, user: dict = Depends(get_current_user)):
    """End an active attendance session"""
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Crew/Admin only")
    
    result = await db.attendance_sessions.update_one(
        {"session_id": session_id, "status": "active"},
        {
            "$set": {
                "status": "completed",
                "ended_at": datetime.now(timezone.utc).isoformat(),
                "ended_by": user["user_id"]
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Active session not found")
    
    # Get final session data
    session = await db.attendance_sessions.find_one({"session_id": session_id}, {"_id": 0})
    return session

@api_router.get("/crew/attendance-sessions")
async def get_crew_sessions(user: dict = Depends(get_current_user)):
    """Get all attendance sessions for the logged-in crew member"""
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Crew/Admin only")
    
    query = {} if user.get("role") == "admin" else {"crew_id": user["user_id"]}
    sessions = await db.attendance_sessions.find(query, {"_id": 0}).sort("started_at", -1).to_list(100)
    return sessions

@api_router.get("/crew/attendance-session/{session_id}")
async def get_session_details(session_id: str, user: dict = Depends(get_current_user)):
    """Get details of a specific attendance session including all attendees"""
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Crew/Admin only")
    
    session = await db.attendance_sessions.find_one({"session_id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get all attendance records for this session
    attendance_records = await db.session_attendance.find(
        {"session_id": session_id},
        {"_id": 0}
    ).sort("marked_at", -1).to_list(500)
    
    session["attendance_records"] = attendance_records
    return session

@api_router.get("/crew/active-session")
async def get_active_session(user: dict = Depends(get_current_user)):
    """Get the current active session for the crew member"""
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Crew/Admin only")
    
    session = await db.attendance_sessions.find_one(
        {"crew_id": user["user_id"], "status": "active"},
        {"_id": 0}
    )
    return session

# ======================== ASSESSMENT CATEGORIES (Admin) ========================

@api_router.post("/admin/assessment-categories")
async def create_assessment_category(data: AssessmentCategoryCreate, user: dict = Depends(get_current_user)):
    """Create a custom assessment category"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    category_id = f"cat_{uuid.uuid4().hex[:12]}"
    category_doc = {
        "category_id": category_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.assessment_categories.insert_one(category_doc)
    return {k: v for k, v in category_doc.items() if k != "_id"}

@api_router.get("/admin/assessment-categories")
async def get_assessment_categories(user: dict = Depends(get_current_user)):
    """Get all assessment categories"""
    if user.get("role") not in ["trainer", "admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    categories = await db.assessment_categories.find({}, {"_id": 0}).to_list(100)
    
    # If no categories exist, seed default ones
    if not categories:
        default_categories = [
            {"name": "Skill Control", "description": "Ability to control vehicle/equipment", "scale_min": 1, "scale_max": 5, "is_active": True, "brand_ids": []},
            {"name": "Discipline", "description": "Following rules and protocols", "scale_min": 1, "scale_max": 5, "is_active": True, "brand_ids": []},
            {"name": "Safety Awareness", "description": "Understanding and practicing safety measures", "scale_min": 1, "scale_max": 5, "is_active": True, "brand_ids": []},
            {"name": "Execution", "description": "Quality of task execution", "scale_min": 1, "scale_max": 5, "is_active": True, "brand_ids": []},
            {"name": "Teamwork", "description": "Collaboration with others", "scale_min": 1, "scale_max": 5, "is_active": True, "brand_ids": []}
        ]
        
        for cat in default_categories:
            cat_id = f"cat_{uuid.uuid4().hex[:12]}"
            cat_doc = {
                "category_id": cat_id,
                **cat,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": "system"
            }
            await db.assessment_categories.insert_one(cat_doc)
            categories.append({k: v for k, v in cat_doc.items() if k != "_id"})
    
    # Enrich with brand names
    all_brands = await db.brands.find({}, {"_id": 0, "brand_id": 1, "name": 1}).to_list(100)
    brand_map = {b["brand_id"]: b["name"] for b in all_brands}
    
    for cat in categories:
        cat["brand_names"] = [brand_map.get(bid, "Unknown") for bid in cat.get("brand_ids", [])]
    
    return categories

@api_router.put("/admin/assessment-categories/{category_id}")
async def update_assessment_category(category_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update an assessment category"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data.pop("category_id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.assessment_categories.update_one(
        {"category_id": category_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    return {"message": "Category updated"}

@api_router.delete("/admin/assessment-categories/{category_id}")
async def delete_assessment_category(category_id: str, user: dict = Depends(get_current_user)):
    """Delete an assessment category"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.assessment_categories.delete_one({"category_id": category_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    return {"message": "Category deleted"}

# ======================== NFC ASSESSMENT ENDPOINTS ========================

@api_router.post("/crew/assessment/nfc-submit")
async def submit_nfc_assessment(data: NFCAssessmentCreate, user: dict = Depends(get_current_user)):
    """Submit assessment with NFC confirmation from crew"""
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Crew/Admin only")
    
    # Verify crew NFC card
    crew_nfc = data.crew_nfc_confirmation.upper().strip()
    
    # Check if NFC belongs to the crew member
    nfc_record = await db.nfc_cards.find_one({"nfc_card_id": crew_nfc, "user_id": user["user_id"], "status": "active"})
    crew_user = await db.users.find_one({"user_id": user["user_id"], "nfc_card_id": crew_nfc})
    nfc_user_record = await db.nfc_users.find_one({"nfc_id": crew_nfc, "is_active": True})
    
    # Allow if NFC matches current user OR if user is admin
    if not (nfc_record or crew_user or (nfc_user_record and user.get("role") == "admin")):
        # For flexibility, also check if NFC belongs to any trainer
        any_crew_nfc = await db.nfc_users.find_one({"nfc_id": crew_nfc, "role": {"$in": ["trainer", "admin"]}, "is_active": True})
        if not any_crew_nfc:
            raise HTTPException(status_code=400, detail="Invalid crew NFC confirmation. Please tap your crew NFC card.")
    
    # Verify student exists
    student = await db.students.find_one({"student_id": data.student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Verify session exists
    session = await db.attendance_sessions.find_one({"session_id": data.session_id}, {"_id": 0})
    
    # Get categories for calculating average
    categories = await db.assessment_categories.find({"is_active": True}, {"_id": 0}).to_list(100)
    category_map = {c["category_id"]: c for c in categories}
    
    # Calculate average rating
    total_rating = 0
    rating_count = 0
    rating_details = []
    
    for cat_id, rating in data.ratings.items():
        category = category_map.get(cat_id)
        if category:
            total_rating += rating
            rating_count += 1
            rating_details.append({
                "category_id": cat_id,
                "category_name": category.get("name"),
                "rating": rating,
                "scale_max": category.get("scale_max", 5)
            })
    
    avg_rating = round(total_rating / rating_count, 2) if rating_count > 0 else 0
    
    # Create assessment record
    assessment_id = f"nfcass_{uuid.uuid4().hex[:12]}"
    assessment_doc = {
        "assessment_id": assessment_id,
        "student_id": data.student_id,
        "student_name": student.get("full_name"),
        "session_id": data.session_id,
        "unit_id": data.unit_id,
        "ratings": data.ratings,
        "rating_details": rating_details,
        "average_rating": avg_rating,
        "notes": data.notes,
        "crew_id": user["user_id"],
        "crew_name": user.get("name"),
        "crew_nfc_confirmation": crew_nfc,
        "confirmed_at": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.nfc_assessments.insert_one(assessment_doc)
    
    # Update student's average rating
    all_assessments = await db.nfc_assessments.find({"student_id": data.student_id}, {"_id": 0}).to_list(500)
    if all_assessments:
        overall_avg = sum(a.get("average_rating", 0) for a in all_assessments) / len(all_assessments)
        await db.students.update_one(
            {"student_id": data.student_id},
            {"$set": {"average_rating": round(overall_avg, 2)}}
        )
    
    return {
        "status": "success",
        "message": f"Assessment submitted for {student.get('full_name')}",
        "assessment_id": assessment_id,
        "average_rating": avg_rating,
        "confirmed_by": user.get("name"),
        "confirmed_at": assessment_doc["confirmed_at"]
    }

@api_router.get("/crew/assessments/session/{session_id}")
async def get_session_assessments(session_id: str, user: dict = Depends(get_current_user)):
    """Get all assessments for a specific session"""
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Crew/Admin only")
    
    assessments = await db.nfc_assessments.find(
        {"session_id": session_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    
    return assessments

@api_router.get("/crew/units")
async def get_units_for_crew(user: dict = Depends(get_current_user)):
    """Get all units available for the crew member (based on their brand if brand-locked)"""
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Crew/Admin only")
    
    # Check if crew is brand-locked
    assigned_brand = user.get("assigned_brand_id")
    
    if assigned_brand and user.get("role") != "admin":
        # Get units only for assigned brand
        units = await db.program_units.find(
            {"brand_id": assigned_brand},
            {"_id": 0}
        ).sort("order", 1).to_list(100)
    else:
        # Admin can see all units
        units = await db.program_units.find({}, {"_id": 0}).sort("order", 1).to_list(500)
    
    # Enrich with program and brand info
    enriched_units = []
    for unit in units:
        program = await db.programs.find_one({"program_id": unit.get("program_id")}, {"_id": 0, "name": 1})
        brand = await db.brands.find_one({"brand_id": unit.get("brand_id")}, {"_id": 0, "name": 1, "color": 1})
        
        unit["program_name"] = program.get("name") if program else "Unknown"
        unit["brand_name"] = brand.get("name") if brand else "Unknown"
        unit["brand_color"] = brand.get("color", "#00f0ff") if brand else "#00f0ff"
        enriched_units.append(unit)
    
    return enriched_units

# ======================== ASSESSMENT ENDPOINTS ========================

@api_router.post("/assessments")
async def create_assessment(data: AssessmentCreate, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    assessment_id = f"assess_{uuid.uuid4().hex[:12]}"
    avg_rating = (data.skill_control + data.discipline + data.safety_awareness + data.execution + data.teamwork) / 5
    
    assessment_doc = {
        "assessment_id": assessment_id,
        **data.model_dump(),
        "average_rating": round(avg_rating, 2),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.assessments.insert_one(assessment_doc)
    
    # Update student average
    all_assessments = await db.assessments.find({"student_id": data.student_id}, {"_id": 0}).to_list(100)
    if all_assessments:
        overall_avg = sum(a["average_rating"] for a in all_assessments) / len(all_assessments)
        await db.students.update_one(
            {"student_id": data.student_id},
            {"$set": {"average_rating": round(overall_avg, 2)}}
        )
    
    return {k: v for k, v in assessment_doc.items() if k != "_id"}

@api_router.get("/assessments/student/{student_id}")
async def get_student_assessments(student_id: str, user: dict = Depends(get_current_user)):
    assessments = await db.assessments.find({"student_id": student_id}, {"_id": 0}).to_list(100)
    return assessments

# ======================== AI GAP ANALYSIS ========================

@api_router.get("/analysis/student/{student_id}")
async def get_gap_analysis(student_id: str, user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    assessments = await db.assessments.find({"student_id": student_id}, {"_id": 0}).to_list(100)
    
    if not assessments:
        return {
            "student_id": student_id,
            "insights": "No assessments available yet",
            "strong_areas": [],
            "improvement_areas": [],
            "recommendations": []
        }
    
    # Calculate averages per skill
    skills = {
        "skill_control": [],
        "discipline": [],
        "safety_awareness": [],
        "execution": [],
        "teamwork": []
    }
    
    for a in assessments:
        for skill in skills:
            skills[skill].append(a.get(skill, 3))
    
    skill_avgs = {k: sum(v)/len(v) if v else 0 for k, v in skills.items()}
    
    # Determine strong and weak areas
    strong = [k.replace("_", " ").title() for k, v in skill_avgs.items() if v >= 4]
    weak = [k.replace("_", " ").title() for k, v in skill_avgs.items() if v < 3]
    
    # Generate AI insights using Gemini
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        chat = LlmChat(
            api_key=os.environ.get("EMERGENT_LLM_KEY"),
            session_id=f"analysis_{student_id}_{uuid.uuid4().hex[:8]}",
            system_message="You are an expert motorsport and automotive education trainer. Provide concise, actionable insights for student improvement."
        ).with_model("gemini", "gemini-3-flash-preview")
        
        prompt = f"""Analyze this motorsport student's performance:
Student: {student.get('full_name', 'Unknown')}
Skill Averages (1-5 scale):
- Skill Control: {skill_avgs['skill_control']:.1f}
- Discipline: {skill_avgs['discipline']:.1f}
- Safety Awareness: {skill_avgs['safety_awareness']:.1f}
- Execution: {skill_avgs['execution']:.1f}
- Teamwork: {skill_avgs['teamwork']:.1f}

Medical conditions: {', '.join(student.get('medical_conditions', [])) or 'None'}

Provide in JSON format:
{{"insights": "2-3 sentence summary", "recommendations": ["3 specific actionable items"], "safety_flags": ["any concerns based on medical/performance"]}}"""
        
        response = await chat.send_message(UserMessage(text=prompt))
        
        import json
        try:
            # Try to parse JSON from response
            json_start = response.find("{")
            json_end = response.rfind("}") + 1
            if json_start >= 0 and json_end > json_start:
                ai_response = json.loads(response[json_start:json_end])
            else:
                ai_response = {"insights": response, "recommendations": [], "safety_flags": []}
        except:
            ai_response = {"insights": response, "recommendations": [], "safety_flags": []}
        
    except Exception as e:
        logger.error(f"AI analysis error: {e}")
        ai_response = {
            "insights": f"Based on assessments, overall performance is {'good' if student.get('average_rating', 0) >= 3.5 else 'needs improvement'}.",
            "recommendations": ["Focus on weak areas", "Practice consistently", "Follow safety protocols"],
            "safety_flags": []
        }
    
    return {
        "student_id": student_id,
        "student_name": student.get("full_name"),
        "skill_averages": skill_avgs,
        "strong_areas": strong,
        "improvement_areas": weak,
        "ai_insights": ai_response.get("insights", ""),
        "recommendations": ai_response.get("recommendations", []),
        "safety_flags": ai_response.get("safety_flags", []),
        "total_assessments": len(assessments),
        "average_rating": student.get("average_rating", 0)
    }

# ======================== LEADERBOARD & BADGES ========================

@api_router.get("/leaderboard")
async def get_leaderboard(batch_id: Optional[str] = None, program_id: Optional[str] = None):
    query = {"status": "active"}
    if batch_id:
        query["batch_id"] = batch_id
    
    students = await db.students.find(query, {"_id": 0}).to_list(1000)
    
    # Sort by average rating
    sorted_students = sorted(students, key=lambda x: (x.get("average_rating", 0), x.get("total_attendance", 0)), reverse=True)
    
    leaderboard = []
    for rank, student in enumerate(sorted_students[:50], 1):
        leaderboard.append({
            "rank": rank,
            "student_id": student["student_id"],
            "name": student.get("full_name", "Unknown"),
            "average_rating": student.get("average_rating", 0),
            "total_attendance": student.get("total_attendance", 0),
            "badges": student.get("badges", [])
        })
    
    return leaderboard

@api_router.post("/badges/award")
async def award_badge(student_id: str, badge_type: str, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    valid_badges = ["batch_leader", "programme_leader", "discipline_champion", "safety_star", "top_performer"]
    if badge_type not in valid_badges:
        raise HTTPException(status_code=400, detail="Invalid badge type")
    
    badge_doc = {
        "badge_type": badge_type,
        "awarded_at": datetime.now(timezone.utc).isoformat(),
        "awarded_by": user["user_id"]
    }
    
    await db.students.update_one(
        {"student_id": student_id},
        {"$push": {"badges": badge_doc}}
    )
    
    return {"message": f"Badge '{badge_type}' awarded", "badge": badge_doc}

# ======================== CERTIFICATES ========================

@api_router.post("/certificates/generate/{student_id}")
async def generate_certificate(student_id: str, program_id: str, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    program = await db.programs.find_one({"program_id": program_id}, {"_id": 0})
    
    if not student or not program:
        raise HTTPException(status_code=404, detail="Student or program not found")
    
    cert_id = f"cert_{uuid.uuid4().hex[:12]}"
    
    # Generate QR code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(f"https://kotlerx.verify/{cert_id}")
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = io.BytesIO()
    qr_img.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    cert_doc = {
        "certificate_id": cert_id,
        "student_id": student_id,
        "student_name": student.get("full_name"),
        "program_id": program_id,
        "program_name": program.get("name"),
        "program_type": program.get("program_type"),
        "issued_at": datetime.now(timezone.utc).isoformat(),
        "issued_by": user["user_id"],
        "qr_code": qr_base64,
        "status": "issued",
        "verified": False
    }
    
    await db.certificates.insert_one(cert_doc)
    return {k: v for k, v in cert_doc.items() if k != "_id"}

@api_router.get("/certificates/student/{student_id}")
async def get_student_certificates(student_id: str, user: dict = Depends(get_current_user)):
    certs = await db.certificates.find({"student_id": student_id}, {"_id": 0}).to_list(50)
    return certs

@api_router.get("/certificates/verify/{cert_id}")
async def verify_certificate(cert_id: str):
    cert = await db.certificates.find_one({"certificate_id": cert_id}, {"_id": 0})
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    return {"valid": True, "certificate": cert}

# ======================== UPGRADE & DISCOUNT ========================

@api_router.post("/upgrades/apply")
async def apply_upgrade(data: UpgradeApplication, user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"student_id": data.student_id}, {"_id": 0})
    current_program = await db.programs.find_one({"program_id": data.current_program_id}, {"_id": 0})
    target_program = await db.programs.find_one({"program_id": data.target_program_id}, {"_id": 0})
    
    if not all([student, current_program, target_program]):
        raise HTTPException(status_code=404, detail="Invalid data")
    
    # Calculate discount based on current progress
    discount_percent = 15  # Base discount for early upgrade
    
    upgrade_id = f"upg_{uuid.uuid4().hex[:12]}"
    upgrade_doc = {
        "upgrade_id": upgrade_id,
        "student_id": data.student_id,
        "current_program": current_program.get("name"),
        "target_program": target_program.get("name"),
        "discount_percent": discount_percent,
        "status": "pending",
        "applied_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.upgrades.insert_one(upgrade_doc)
    return {k: v for k, v in upgrade_doc.items() if k != "_id"}

@api_router.get("/upgrades/available/{student_id}")
async def get_available_upgrades(student_id: str, user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Get all programs for upgrade paths
    programs = await db.programs.find({}, {"_id": 0}).to_list(100)
    
    upgrade_paths = {
        "certification": ["diploma", "pg_diploma"],
        "diploma": ["pg_diploma"]
    }
    
    available = []
    current_type = "certification"  # Default
    
    for prog in programs:
        if prog.get("program_type") in upgrade_paths.get(current_type, []):
            available.append({
                "program_id": prog["program_id"],
                "name": prog["name"],
                "type": prog["program_type"],
                "discount_available": 15
            })
    
    return available

# ======================== ADMIN DASHBOARD ========================

@api_router.get("/admin/stats")
async def get_admin_stats(user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    total_students = await db.students.count_documents({})
    active_students = await db.students.count_documents({"status": "active"})
    total_programs = await db.programs.count_documents({})
    total_batches = await db.batches.count_documents({})
    total_certificates = await db.certificates.count_documents({})
    
    # Medical flags
    medical_flags = await db.students.count_documents({"medical_conditions": {"$ne": []}})
    
    # Attendance stats
    attendance_pipeline = [
        {"$group": {"_id": None, "total": {"$sum": 1}}}
    ]
    attendance_result = await db.attendance.aggregate(attendance_pipeline).to_list(1)
    total_attendance = attendance_result[0]["total"] if attendance_result else 0
    
    # Top performers
    top_students = await db.students.find(
        {"status": "active"},
        {"_id": 0, "student_id": 1, "full_name": 1, "average_rating": 1}
    ).sort("average_rating", -1).limit(5).to_list(5)
    
    return {
        "total_students": total_students,
        "active_students": active_students,
        "total_programs": total_programs,
        "total_batches": total_batches,
        "total_certificates": total_certificates,
        "medical_flags": medical_flags,
        "total_attendance_records": total_attendance,
        "top_performers": top_students,
        "placement_ready": await db.students.count_documents({"average_rating": {"$gte": 4.0}})
    }

# ======================== ADMIN STUDENT MANAGEMENT ========================

@api_router.put("/admin/students/{student_id}")
async def update_student(student_id: str, data: dict, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Remove protected fields
    data.pop("student_id", None)
    data.pop("user_id", None)
    data.pop("nfc_card_id", None)
    
    result = await db.students.update_one(
        {"student_id": student_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
    
    return {"message": "Student updated"}

@api_router.delete("/admin/students/{student_id}")
async def delete_student(student_id: str, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.students.delete_one({"student_id": student_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
    
    return {"message": "Student removed"}

# ======================== ADMIN MESSAGES ========================

class AdminMessage(BaseModel):
    title: str
    content: str
    type: str = "info"  # info, announcement, alert

@api_router.post("/admin/message")
async def create_admin_message(data: AdminMessage, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    message_id = f"msg_{uuid.uuid4().hex[:12]}"
    message_doc = {
        "message_id": message_id,
        **data.model_dump(),
        "created_by": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "active": True
    }
    
    await db.admin_messages.insert_one(message_doc)
    return {k: v for k, v in message_doc.items() if k != "_id"}

@api_router.get("/admin/messages")
async def get_admin_messages():
    messages = await db.admin_messages.find(
        {"active": True}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(10)
    return messages

# ======================== LEADS (Public Program Interest) ========================

@api_router.post("/leads")
async def create_lead(data: LeadRegistration):
    """Public endpoint for program interest registration"""
    lead_id = f"lead_{uuid.uuid4().hex[:12]}"
    lead_doc = {
        "lead_id": lead_id,
        **data.model_dump(),
        "status": "new",  # new, contacted, converted, dropped
        "created_at": datetime.now(timezone.utc).isoformat(),
        "notes": ""
    }
    
    await db.leads.insert_one(lead_doc)
    return {"message": "Lead registered", "lead_id": lead_id}

@api_router.get("/admin/leads")
async def get_leads(user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    leads = await db.leads.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return leads

@api_router.put("/admin/leads/{lead_id}")
async def update_lead(lead_id: str, data: dict, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data.pop("lead_id", None)
    data.pop("created_at", None)
    
    result = await db.leads.update_one(
        {"lead_id": lead_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    return {"message": "Lead updated"}

@api_router.get("/admin/leads/stats")
async def get_lead_stats(user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    total = await db.leads.count_documents({})
    new_leads = await db.leads.count_documents({"status": "new"})
    contacted = await db.leads.count_documents({"status": "contacted"})
    converted = await db.leads.count_documents({"status": "converted"})
    
    return {
        "total": total,
        "new": new_leads,
        "contacted": contacted,
        "converted": converted
    }

@api_router.get("/admin/leads/export")
async def export_leads_excel(user: dict = Depends(get_current_user)):
    """Export all leads to Excel file"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    leads = await db.leads.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Headers
    headers = ["Lead ID", "Name", "Location", "Mobile", "Program Interest", "Fee Type", "Status", "Created At", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
    
    # Data rows
    for row, lead in enumerate(leads, 2):
        ws.cell(row=row, column=1, value=lead.get("lead_id", ""))
        ws.cell(row=row, column=2, value=lead.get("name", ""))
        ws.cell(row=row, column=3, value=lead.get("location", ""))
        ws.cell(row=row, column=4, value=lead.get("mobile", ""))
        ws.cell(row=row, column=5, value=lead.get("program_interest", ""))
        ws.cell(row=row, column=6, value=lead.get("fee_type", ""))
        ws.cell(row=row, column=7, value=lead.get("status", ""))
        ws.cell(row=row, column=8, value=lead.get("created_at", ""))
        ws.cell(row=row, column=9, value=lead.get("notes", ""))
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"kotlerx_leads_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ======================== MISSED SESSIONS ========================

@api_router.get("/sessions/missed/{student_id}")
async def get_missed_sessions(student_id: str, user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Get batch schedule and compare with attendance
    batch = await db.batches.find_one({"batch_id": student.get("batch_id")}, {"_id": 0})
    if not batch:
        return {"missed": [], "makeup_available": []}
    
    attendance = await db.attendance.find({"student_id": student_id}, {"_id": 0}).to_list(500)
    attended_dates = set(a["timestamp"][:10] for a in attendance if a.get("timestamp"))
    
    # For demo, return sample missed sessions
    return {
        "missed": [],
        "makeup_available": [],
        "total_missed": 0
    }

# ======================== SOP & SAFETY ========================

@api_router.get("/sop")
async def get_sop():
    return {
        "safety_rules": [
            "Always wear protective gear during practical sessions",
            "Follow trainer instructions at all times",
            "Report any equipment malfunction immediately",
            "No mobile phones during active training",
            "Emergency contacts must be updated"
        ],
        "emergency_protocols": [
            "In case of injury, alert trainer immediately",
            "Know the location of first aid kits",
            "Emergency contact: 112 (India)",
            "Fire exits are marked with green signs"
        ],
        "sop_documents": [
            {"name": "Track Safety Guidelines", "version": "2.1"},
            {"name": "Equipment Handling", "version": "1.5"},
            {"name": "Emergency Response", "version": "3.0"}
        ]
    }

# ======================== UNIVERSITY ASSESSMENT REPORTS ========================

async def generate_assessment_report_data(batch_id: Optional[str] = None, report_type: str = "batch"):
    """Generate assessment report data for university/college"""
    
    # Get students based on filter
    query = {"status": "active"}
    if batch_id:
        query["batch_id"] = batch_id
    
    students = await db.students.find(query, {"_id": 0}).to_list(1000)
    
    report_data = []
    for student in students:
        # Get attendance records
        attendance_records = await db.attendance.find(
            {"student_id": student["student_id"]}, {"_id": 0}
        ).to_list(500)
        
        # Get assessments
        assessments = await db.assessments.find(
            {"student_id": student["student_id"]}, {"_id": 0}
        ).to_list(100)
        
        # Calculate attendance percentage (assume 20 sessions per program)
        total_sessions = 20  # Can be made dynamic based on batch
        attendance_count = len(attendance_records)
        attendance_percent = round((attendance_count / total_sessions) * 100, 1) if total_sessions > 0 else 0
        
        # Calculate assessment averages
        if assessments:
            skill_avg = sum(a.get("skill_control", 0) for a in assessments) / len(assessments)
            discipline_avg = sum(a.get("discipline", 0) for a in assessments) / len(assessments)
            safety_avg = sum(a.get("safety_awareness", 0) for a in assessments) / len(assessments)
            execution_avg = sum(a.get("execution", 0) for a in assessments) / len(assessments)
            teamwork_avg = sum(a.get("teamwork", 0) for a in assessments) / len(assessments)
            overall_avg = (skill_avg + discipline_avg + safety_avg + execution_avg + teamwork_avg) / 5
        else:
            skill_avg = discipline_avg = safety_avg = execution_avg = teamwork_avg = overall_avg = 0
        
        # Get certificate status
        certificate = await db.certificates.find_one(
            {"student_id": student["student_id"]}, {"_id": 0}
        )
        cert_status = certificate.get("status", "Not Issued") if certificate else "Not Issued"
        
        # Get user email
        user = await db.users.find_one({"user_id": student.get("user_id")}, {"_id": 0, "email": 1})
        
        report_data.append({
            "student_id": student["student_id"],
            "full_name": student.get("full_name", ""),
            "email": user.get("email", "") if user else "",
            "mobile": student.get("mobile", ""),
            "nfc_card_id": student.get("nfc_card_id", ""),
            "batch_id": student.get("batch_id", ""),
            "attendance_sessions": attendance_count,
            "attendance_percent": attendance_percent,
            "skill_control": round(skill_avg, 2),
            "discipline": round(discipline_avg, 2),
            "safety_awareness": round(safety_avg, 2),
            "execution": round(execution_avg, 2),
            "teamwork": round(teamwork_avg, 2),
            "overall_rating": round(overall_avg, 2),
            "certificate_status": cert_status,
            "blood_group": student.get("blood_group", ""),
            "emergency_contact": student.get("emergency_contact", ""),
            "medical_conditions": ", ".join(student.get("medical_conditions", [])) or "None"
        })
    
    return report_data

def create_university_report_excel(report_data: list, report_title: str):
    """Create Excel workbook for university report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Assessment Report"
    
    # Title row
    ws.merge_cells('A1:Q1')
    ws['A1'] = f"KotlerX University Assessment Report - {report_title}"
    ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
    
    # Generated date
    ws['A2'] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    
    # Headers
    headers = [
        "Student ID", "Full Name", "Email", "Mobile", "NFC Card ID", "Batch ID",
        "Attendance Sessions", "Attendance %", "Skill Control", "Discipline",
        "Safety Awareness", "Execution", "Teamwork", "Overall Rating",
        "Certificate Status", "Blood Group", "Emergency Contact", "Medical Conditions"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = cell.font.copy(bold=True)
    
    # Data rows
    for row, data in enumerate(report_data, 5):
        ws.cell(row=row, column=1, value=data.get("student_id", ""))
        ws.cell(row=row, column=2, value=data.get("full_name", ""))
        ws.cell(row=row, column=3, value=data.get("email", ""))
        ws.cell(row=row, column=4, value=data.get("mobile", ""))
        ws.cell(row=row, column=5, value=data.get("nfc_card_id", ""))
        ws.cell(row=row, column=6, value=data.get("batch_id", ""))
        ws.cell(row=row, column=7, value=data.get("attendance_sessions", 0))
        ws.cell(row=row, column=8, value=data.get("attendance_percent", 0))
        ws.cell(row=row, column=9, value=data.get("skill_control", 0))
        ws.cell(row=row, column=10, value=data.get("discipline", 0))
        ws.cell(row=row, column=11, value=data.get("safety_awareness", 0))
        ws.cell(row=row, column=12, value=data.get("execution", 0))
        ws.cell(row=row, column=13, value=data.get("teamwork", 0))
        ws.cell(row=row, column=14, value=data.get("overall_rating", 0))
        ws.cell(row=row, column=15, value=data.get("certificate_status", ""))
        ws.cell(row=row, column=16, value=data.get("blood_group", ""))
        ws.cell(row=row, column=17, value=data.get("emergency_contact", ""))
        ws.cell(row=row, column=18, value=data.get("medical_conditions", ""))
    
    # Summary row
    summary_row = len(report_data) + 6
    ws.cell(row=summary_row, column=1, value="SUMMARY")
    ws.cell(row=summary_row, column=1).font = ws.cell(row=summary_row, column=1).font.copy(bold=True)
    ws.cell(row=summary_row, column=2, value=f"Total Students: {len(report_data)}")
    
    if report_data:
        avg_attendance = sum(d["attendance_percent"] for d in report_data) / len(report_data)
        avg_rating = sum(d["overall_rating"] for d in report_data) / len(report_data)
        ws.cell(row=summary_row, column=8, value=f"Avg: {avg_attendance:.1f}%")
        ws.cell(row=summary_row, column=14, value=f"Avg: {avg_rating:.2f}")
    
    # Adjust column widths
    column_widths = [15, 20, 25, 15, 18, 18, 12, 12, 12, 12, 15, 12, 12, 12, 15, 12, 18, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width
    
    return wb

@api_router.get("/admin/reports/batch/{batch_id}")
async def get_batch_report(batch_id: str, user: dict = Depends(get_current_user)):
    """Generate assessment report for a specific batch (for university)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    batch = await db.batches.find_one({"batch_id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    program = await db.programs.find_one({"program_id": batch.get("program_id")}, {"_id": 0})
    program_name = program.get("name", "Unknown") if program else "Unknown"
    
    report_data = await generate_assessment_report_data(batch_id=batch_id, report_type="batch")
    
    wb = create_university_report_excel(report_data, f"Batch: {batch_id} | Program: {program_name}")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"university_report_batch_{batch_id}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/admin/reports/weekly")
async def get_weekly_report(user: dict = Depends(get_current_user)):
    """Generate weekly assessment report for all active students (for university)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Get all students' data
    report_data = await generate_assessment_report_data(report_type="weekly")
    
    # Get week range
    today = datetime.now(timezone.utc)
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    
    wb = create_university_report_excel(
        report_data, 
        f"Weekly Report ({week_start.strftime('%Y-%m-%d')} to {week_end.strftime('%Y-%m-%d')})"
    )
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"university_weekly_report_{today.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/admin/reports/monthly")
async def get_monthly_report(user: dict = Depends(get_current_user)):
    """Generate monthly assessment report for all active students (for university)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    report_data = await generate_assessment_report_data(report_type="monthly")
    
    today = datetime.now(timezone.utc)
    month_name = today.strftime('%B %Y')
    
    wb = create_university_report_excel(report_data, f"Monthly Report - {month_name}")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"university_monthly_report_{today.strftime('%Y%m')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

class BatchCompletion(BaseModel):
    batch_id: str
    completion_date: Optional[str] = None

@api_router.post("/admin/reports/completion")
async def generate_completion_report(data: BatchCompletion, user: dict = Depends(get_current_user)):
    """Generate program completion report for a batch (final report for university)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    batch = await db.batches.find_one({"batch_id": data.batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    program = await db.programs.find_one({"program_id": batch.get("program_id")}, {"_id": 0})
    program_name = program.get("name", "Unknown") if program else "Unknown"
    
    report_data = await generate_assessment_report_data(batch_id=data.batch_id, report_type="completion")
    
    completion_date = data.completion_date or datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    wb = create_university_report_excel(
        report_data, 
        f"PROGRAM COMPLETION | Batch: {data.batch_id} | Program: {program_name} | Completed: {completion_date}"
    )
    
    # Mark batch as completed
    await db.batches.update_one(
        {"batch_id": data.batch_id},
        {"$set": {
            "status": "completed",
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "completion_report_generated": True
        }}
    )
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"university_completion_report_{data.batch_id}_{completion_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ======================== CMS - SITE SETTINGS ========================

class SiteSettings(BaseModel):
    university_email: Optional[str] = None
    auto_email_enabled: bool = False
    auto_email_weekly: bool = False
    auto_email_monthly: bool = False
    logo_text_1: str = "KX"
    logo_text_2: str = "GRID"
    site_title: str = "KXGRID"
    site_tagline: str = "Unified Operating Platform"
    hero_headline_1: str = "Unified Operating Platform for the KotlerX Ecosystem"
    hero_headline_2: str = "Connecting Brands, Programmes, Students, Crew & Partners"
    hero_headline_3: str = "NFC + AI-powered Skill Tracking Platform"
    hero_description: str = "GRID enables programme execution, department coordination, attendance & assessment tracking, content delivery, and brand visibility across the ecosystem."
    footer_text: str = "KXGRID - Unified Operating Platform"
    contact_email: str = ""
    contact_phone: str = ""

class LandingPageContent(BaseModel):
    hero_headline_1: Optional[str] = None
    hero_headline_2: Optional[str] = None
    hero_headline_3: Optional[str] = None
    hero_description: Optional[str] = None
    features: Optional[List[Dict[str, str]]] = None
    stats: Optional[Dict[str, Any]] = None

@api_router.get("/cms/settings")
async def get_site_settings():
    """Get public site settings (no auth required)"""
    settings = await db.site_settings.find_one({"setting_id": "main"}, {"_id": 0})
    if not settings:
        # Return defaults
        return SiteSettings().model_dump()
    return settings

@api_router.put("/cms/settings")
async def update_site_settings(data: dict, user: dict = Depends(get_current_user)):
    """Update site settings (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data["setting_id"] = "main"
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = user["user_id"]
    
    await db.site_settings.update_one(
        {"setting_id": "main"},
        {"$set": data},
        upsert=True
    )
    
    return {"message": "Settings updated"}

@api_router.post("/cms/logo/upload")
async def upload_logo(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload logo image (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Validate file type
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    # Read and encode as base64
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:  # 2MB limit
        raise HTTPException(status_code=400, detail="Image too large (max 2MB)")
    
    base64_image = base64.b64encode(content).decode('utf-8')
    data_url = f"data:{file.content_type};base64,{base64_image}"
    
    # Save to settings
    await db.site_settings.update_one(
        {"setting_id": "main"},
        {"$set": {
            "logo_image": data_url,
            "logo_updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"message": "Logo uploaded", "logo_url": data_url}

class Base64ImageUpload(BaseModel):
    image_base64: str

@api_router.post("/cms/logo/upload-base64")
async def upload_logo_base64(data: Base64ImageUpload, user: dict = Depends(get_current_user)):
    """Upload logo image as base64 (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Validate base64 image
    if not data.image_base64.startswith('data:image/'):
        raise HTTPException(status_code=400, detail="Invalid image format")
    
    # Check size (base64 is ~33% larger than binary)
    if len(data.image_base64) > 3 * 1024 * 1024:  # ~2MB original
        raise HTTPException(status_code=400, detail="Image too large (max 2MB)")
    
    # Save to settings
    await db.site_settings.update_one(
        {"setting_id": "main"},
        {"$set": {
            "logo_image": data.image_base64,
            "logo_updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"message": "Logo uploaded", "logo_url": data.image_base64}

@api_router.delete("/cms/logo")
async def delete_logo(user: dict = Depends(get_current_user)):
    """Delete uploaded logo and revert to text (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    await db.site_settings.update_one(
        {"setting_id": "main"},
        {"$unset": {"logo_image": ""}}
    )
    
    return {"message": "Logo deleted, reverted to text"}

@api_router.get("/cms/landing")
async def get_landing_content():
    """Get landing page content (public)"""
    content = await db.cms_content.find_one({"$or": [{"page": "landing"}, {"section": "hero"}]}, {"_id": 0})
    if not content:
        # Return defaults
        return {
            "page": "landing",
            "hero_headline_1": "Unified Operating Platform for the KotlerX Ecosystem",
            "hero_headline_2": "Connecting Brands, Programmes, Students, Crew & Partners",
            "hero_headline_3": "NFC + AI-powered Skill Tracking Platform",
            "hero_description": "GRID enables programme execution, department coordination, attendance & assessment tracking, content delivery, and brand visibility across the ecosystem.",
            "features": [
                {"title": "NFC Identity", "description": "Secure NFC-based student identification"},
                {"title": "AI Analytics", "description": "AI-powered performance gap analysis"},
                {"title": "Multi-Brand Ops", "description": "Unified platform for all departments"},
                {"title": "Certifications", "description": "Industry-recognized certifications"}
            ],
            "stats": {
                "students_trained": "500+",
                "programs": "10+",
                "placement_rate": "95%",
                "industry_partners": "20+"
            }
        }
    return content

@api_router.put("/cms/landing")
async def update_landing_content(data: dict, user: dict = Depends(get_current_user)):
    """Update landing page content (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data["page"] = "landing"
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = user["user_id"]
    
    await db.cms_content.update_one(
        {"page": "landing"},
        {"$set": data},
        upsert=True
    )
    
    return {"message": "Landing page updated"}

@api_router.get("/cms/programs-page")
async def get_programs_page_content():
    """Get programs page content (public)"""
    content = await db.cms_content.find_one({"page": "programs"}, {"_id": 0})
    if not content:
        return {
            "page": "programs",
            "title": "OUR PROGRAMS",
            "subtitle": "Choose your path to motorsport excellence",
            "cta_title": "Already a Student?",
            "cta_subtitle": "Tap your NFC card or enter your NFC ID to access your dashboard"
        }
    return content

@api_router.put("/cms/programs-page")
async def update_programs_page_content(data: dict, user: dict = Depends(get_current_user)):
    """Update programs page content (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data["page"] = "programs"
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.cms_content.update_one(
        {"page": "programs"},
        {"$set": data},
        upsert=True
    )
    
    return {"message": "Programs page updated"}

# ======================== THEME CUSTOMIZATION ========================

@api_router.get("/cms/theme")
async def get_theme():
    """Get theme settings (public)"""
    theme = await db.cms_content.find_one({"type": "theme"}, {"_id": 0})
    if not theme:
        return {
            "type": "theme",
            "primary_color": "#00f0ff",
            "secondary_color": "#f59e0b",
            "accent_color": "#ef4444",
            "background_color": "#0a0a0f",
            "surface_color": "#111118",
            "heading_font": "Unbounded",
            "body_font": "Inter",
            "border_radius": "0.75rem"
        }
    return theme

@api_router.put("/cms/theme")
async def update_theme(data: dict, user: dict = Depends(get_current_user)):
    """Update theme settings (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data["type"] = "theme"
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = user["user_id"]
    
    await db.cms_content.update_one(
        {"type": "theme"},
        {"$set": data},
        upsert=True
    )
    
    return {"message": "Theme updated"}

# ======================== MEDIA GALLERY ========================

class MediaItem(BaseModel):
    title: str
    description: Optional[str] = ""
    media_type: str  # image, video
    url: Optional[str] = None
    media_base64: Optional[str] = None  # For direct uploads
    thumbnail_url: Optional[str] = None
    thumbnail_base64: Optional[str] = None
    category: str  # public, student
    is_visible: bool = True
    order: int = 0

@api_router.get("/media/gallery/{category}")
async def get_media_gallery(category: str):
    """Get media gallery by category (public only shows public, student shows both)"""
    if category == "student":
        # Students can see both public and student content
        items = await db.media_gallery.find(
            {"is_visible": True},
            {"_id": 0}
        ).sort("order", 1).to_list(50)
    else:
        # Public only sees public content
        items = await db.media_gallery.find(
            {"category": "public", "is_visible": True},
            {"_id": 0}
        ).sort("order", 1).to_list(50)
    return items

@api_router.get("/admin/media/gallery")
async def get_all_media(user: dict = Depends(get_current_user)):
    """Get all media items (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    items = await db.media_gallery.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    return items

@api_router.post("/admin/media/gallery")
async def add_media_item(data: MediaItem, user: dict = Depends(get_current_user)):
    """Add media item (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Validate that either url or base64 is provided
    if not data.url and not data.media_base64:
        raise HTTPException(status_code=400, detail="Either URL or uploaded file is required")
    
    media_id = f"media_{uuid.uuid4().hex[:12]}"
    media_doc = {
        "media_id": media_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    await db.media_gallery.insert_one(media_doc)
    return {k: v for k, v in media_doc.items() if k != "_id"}

@api_router.put("/admin/media/gallery/{media_id}")
async def update_media_item(media_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update media item (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data.pop("media_id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.media_gallery.update_one(
        {"media_id": media_id},
        {"$set": data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Media item not found")
    return {"message": "Media updated"}

@api_router.delete("/admin/media/gallery/{media_id}")
async def delete_media_item(media_id: str, user: dict = Depends(get_current_user)):
    """Delete media item (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.media_gallery.delete_one({"media_id": media_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Media item not found")
    return {"message": "Media deleted"}

# ======================== EMAIL SERVICE ========================

class EmailRequest(BaseModel):
    recipient_email: EmailStr
    subject: str
    html_content: str

class ReportEmailRequest(BaseModel):
    report_type: str  # weekly, monthly, batch, completion
    batch_id: Optional[str] = None

def is_email_configured():
    """Check if email is properly configured"""
    api_key = os.environ.get('RESEND_API_KEY', '')
    return api_key and not api_key.startswith('re_placeholder')

async def send_welcome_email(to_email: str, name: str, role: str, password: str, brand_name: str):
    """Send welcome email with login credentials to new user"""
    if not is_email_configured():
        logger.warning(f"Email not configured. Skipping welcome email for {to_email}")
        return None
    
    role_display = {
        "brand_head": "Brand Manager",
        "trainer": "Crew / Trainer"
    }.get(role, role.title())
    
    login_url = os.environ.get('FRONTEND_URL', 'https://kxgrid.kotlerx.com')
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background: linear-gradient(135deg, #1a1a2e, #16213e); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
            .content {{ background: #f8f9fa; padding: 30px; border-radius: 0 0 10px 10px; }}
            .credentials {{ background: white; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #00f0ff; }}
            .btn {{ display: inline-block; background: #00f0ff; color: #1a1a2e; padding: 12px 30px; text-decoration: none; border-radius: 6px; font-weight: bold; }}
            .footer {{ text-align: center; margin-top: 30px; color: #666; font-size: 12px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Welcome to KXGRID</h1>
                <p>KotlerX Ecosystem Platform</p>
            </div>
            <div class="content">
                <h2>Hello {name}!</h2>
                <p>You have been added as a <strong>{role_display}</strong> for <strong>{brand_name}</strong>.</p>
                
                <div class="credentials">
                    <h3>Your Login Credentials</h3>
                    <p><strong>Email:</strong> {to_email}</p>
                    <p><strong>Password:</strong> {password}</p>
                    <p style="color: #e74c3c; font-size: 12px;">⚠️ Please change your password after your first login.</p>
                </div>
                
                <p style="text-align: center;">
                    <a href="{login_url}/login" class="btn">Login to KXGRID</a>
                </p>
                
                <h3>What you can do:</h3>
                <ul>
                    {"<li>Manage your brand's crew and trainers</li><li>View student progress reports</li><li>Schedule re-classes</li><li>Download completion reports</li>" if role == "brand_head" else "<li>Record student attendance</li><li>Submit assessments</li><li>View student profiles</li>"}
                </ul>
                
                <p>If you have any questions, please contact your administrator.</p>
            </div>
            <div class="footer">
                <p>© 2026 KotlerX. All rights reserved.</p>
                <p>This is an automated message from KXGRID.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    try:
        params = {
            "from": SENDER_EMAIL,
            "to": [to_email],
            "subject": f"Welcome to KXGRID - Your {role_display} Account",
            "html": html_content
        }
        result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Welcome email sent to {to_email}: {result}")
        return result
    except Exception as e:
        logger.error(f"Failed to send welcome email to {to_email}: {e}")
        return None

async def send_email_async(to_email: str, subject: str, html_content: str):
    """Send email using Resend (non-blocking)"""
    if not is_email_configured():
        raise HTTPException(status_code=503, detail="Email not configured. Please add Resend API key in CMS Settings or contact admin.")
    
    try:
        params = {
            "from": SENDER_EMAIL,
            "to": [to_email],
            "subject": subject,
            "html": html_content
        }
        result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Email sent to {to_email}: {result}")
        return result
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        raise

async def send_report_email(report_type: str, recipient_email: str, batch_id: str = None):
    """Generate report and send via email"""
    if not is_email_configured():
        raise HTTPException(status_code=503, detail="Email not configured. Please add Resend API key to enable email reports.")
    
    try:
        # Generate report data
        if report_type == "weekly":
            report_data = await generate_assessment_report_data(report_type="weekly")
            today = datetime.now(timezone.utc)
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            title = f"Weekly Report ({week_start.strftime('%Y-%m-%d')} to {week_end.strftime('%Y-%m-%d')})"
            filename = f"university_weekly_report_{today.strftime('%Y%m%d')}.xlsx"
        elif report_type == "monthly":
            report_data = await generate_assessment_report_data(report_type="monthly")
            today = datetime.now(timezone.utc)
            title = f"Monthly Report - {today.strftime('%B %Y')}"
            filename = f"university_monthly_report_{today.strftime('%Y%m')}.xlsx"
        elif report_type == "batch" and batch_id:
            report_data = await generate_assessment_report_data(batch_id=batch_id, report_type="batch")
            batch = await db.batches.find_one({"batch_id": batch_id}, {"_id": 0})
            title = f"Batch Report: {batch_id}"
            filename = f"university_batch_report_{batch_id}.xlsx"
        else:
            raise ValueError("Invalid report type")
        
        wb = create_university_report_excel(report_data, title)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        report_bytes = output.getvalue()
        report_base64 = base64.b64encode(report_bytes).decode()
        
        # Create email HTML
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; background-color: #0a0a0a; color: #ffffff; padding: 40px;">
            <div style="max-width: 600px; margin: 0 auto; background-color: #1a1a1a; border-radius: 12px; padding: 40px;">
                <h1 style="color: #00f0ff; margin-bottom: 20px;">KotlerX University Report</h1>
                <h2 style="color: #ffffff; font-size: 18px;">{title}</h2>
                <p style="color: #a0a0a0; line-height: 1.6;">
                    Please find attached the {report_type} assessment report for your university records.
                </p>
                <div style="background-color: #2a2a2a; padding: 20px; border-radius: 8px; margin: 20px 0;">
                    <p style="color: #00f0ff; margin: 0;">Report Summary:</p>
                    <ul style="color: #ffffff; margin: 10px 0;">
                        <li>Total Students: {len(report_data)}</li>
                        <li>Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}</li>
                    </ul>
                </div>
                <p style="color: #a0a0a0; font-size: 12px;">
                    This is an automated report from KotlerX Skill Tracking Platform.<br>
                    For questions, contact your KotlerX administrator.
                </p>
            </div>
        </body>
        </html>
        """
        
        # Send email with attachment
        params = {
            "from": SENDER_EMAIL,
            "to": [recipient_email],
            "subject": f"KotlerX {title}",
            "html": html_content,
            "attachments": [
                {
                    "filename": filename,
                    "content": report_base64
                }
            ]
        }
        
        result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Report email sent to {recipient_email}: {result}")
        return {"status": "sent", "email_id": result.get("id")}
        
    except Exception as e:
        logger.error(f"Failed to send report email: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")

@api_router.post("/admin/reports/email")
async def email_report(data: ReportEmailRequest, user: dict = Depends(get_current_user)):
    """Send report to configured university email"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Get university email from settings
    settings = await db.site_settings.find_one({"setting_id": "main"}, {"_id": 0})
    if not settings or not settings.get("university_email"):
        raise HTTPException(status_code=400, detail="University email not configured. Please set it in CMS Settings.")
    
    recipient_email = settings["university_email"]
    
    result = await send_report_email(data.report_type, recipient_email, data.batch_id)
    return {"message": f"Report sent to {recipient_email}", **result}

@api_router.post("/admin/reports/email-custom")
async def email_report_custom(data: dict, user: dict = Depends(get_current_user)):
    """Send report to a custom email address"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    recipient_email = data.get("email")
    report_type = data.get("report_type", "weekly")
    batch_id = data.get("batch_id")
    
    if not recipient_email:
        raise HTTPException(status_code=400, detail="Email address required")
    
    result = await send_report_email(report_type, recipient_email, batch_id)
    return {"message": f"Report sent to {recipient_email}", **result}

@api_router.post("/admin/test-email")
async def test_email(user: dict = Depends(get_current_user)):
    """Send a test email to verify configuration"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    settings = await db.site_settings.find_one({"setting_id": "main"}, {"_id": 0})
    if not settings or not settings.get("university_email"):
        raise HTTPException(status_code=400, detail="University email not configured")
    
    html_content = """
    <html>
    <body style="font-family: Arial, sans-serif; background-color: #0a0a0a; color: #ffffff; padding: 40px;">
        <div style="max-width: 600px; margin: 0 auto; background-color: #1a1a1a; border-radius: 12px; padding: 40px;">
            <h1 style="color: #00f0ff;">KotlerX Email Test</h1>
            <p style="color: #ffffff;">This is a test email from KotlerX platform.</p>
            <p style="color: #a0a0a0;">If you received this email, the email configuration is working correctly.</p>
        </div>
    </body>
    </html>
    """
    
    try:
        await send_email_async(settings["university_email"], "KotlerX - Test Email", html_content)
        return {"message": f"Test email sent to {settings['university_email']}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email failed: {str(e)}")

# ======================== BRAND MANAGEMENT ========================

class BrandCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    logo_url: Optional[str] = None
    color: Optional[str] = "#00f0ff"
    order: int = 0
    is_visible: bool = True
    # Editable stats for "Why Choose" section
    stats_certifications: Optional[str] = "Industry"
    stats_success_rate: Optional[str] = "95%"
    tagline: Optional[str] = None

class BrandUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    logo_url: Optional[str] = None
    color: Optional[str] = None
    order: Optional[int] = None
    is_visible: Optional[bool] = None
    stats_certifications: Optional[str] = None
    stats_success_rate: Optional[str] = None
    tagline: Optional[str] = None

@api_router.get("/brands")
async def get_brands():
    """Get all brands (public endpoint - respects visibility)"""
    brands = await db.brands.find({"is_visible": True}, {"_id": 0}).sort("order", 1).to_list(100)
    return brands

@api_router.get("/brands/{brand_slug}")
async def get_brand_public(brand_slug: str):
    """Get a single brand by slug or ID (public endpoint)"""
    # Try to find by brand_id first
    brand = await db.brands.find_one({"brand_id": brand_slug, "is_visible": True}, {"_id": 0})
    
    # If not found, try to find by name slug (e.g., "kx-core" -> "KX CORE")
    if not brand:
        # Convert slug to possible name: "kx-core" -> "KX CORE"
        name_from_slug = brand_slug.upper().replace("-", " ")
        brand = await db.brands.find_one({"name": name_from_slug, "is_visible": True}, {"_id": 0})
    
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    # Get programs associated with this brand
    programs = await db.programs.find({"brand_id": brand.get("brand_id")}, {"_id": 0}).to_list(50)
    
    # Get trainers/crew assigned to this brand (or all if none assigned)
    trainers = await db.users.find(
        {"role": "trainer", "assigned_brand_id": brand.get("brand_id")},
        {"_id": 0, "user_id": 1, "name": 1, "picture": 1, "email": 1}
    ).to_list(20)
    
    # If no trainers assigned to brand, get all trainers
    if not trainers:
        trainers = await db.users.find(
            {"role": "trainer"},
            {"_id": 0, "user_id": 1, "name": 1, "picture": 1, "email": 1}
        ).to_list(20)
    
    return {
        **brand,
        "programs": programs,
        "trainers": trainers
    }

@api_router.get("/admin/brands")
async def get_all_brands(user: dict = Depends(get_current_user)):
    """Get all brands including hidden ones (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    brands = await db.brands.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    return brands

@api_router.post("/admin/brands")
async def create_brand(data: BrandCreate, user: dict = Depends(get_current_user)):
    """Create a new brand (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    brand_id = f"brand_{uuid.uuid4().hex[:12]}"
    brand_doc = {
        "brand_id": brand_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.brands.insert_one(brand_doc)
    return {k: v for k, v in brand_doc.items() if k != "_id"}

@api_router.get("/admin/brands/{brand_id}")
async def get_brand(brand_id: str, user: dict = Depends(get_current_user)):
    """Get a specific brand (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    return brand

@api_router.put("/admin/brands/{brand_id}")
async def update_brand(brand_id: str, data: BrandUpdate, user: dict = Depends(get_current_user)):
    """Update a brand (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = user["user_id"]
    
    result = await db.brands.update_one(
        {"brand_id": brand_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    return {"message": "Brand updated"}

@api_router.delete("/admin/brands/{brand_id}")
async def delete_brand(brand_id: str, user: dict = Depends(get_current_user)):
    """Delete a brand (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.brands.delete_one({"brand_id": brand_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    return {"message": "Brand deleted"}

@api_router.post("/admin/brands/{brand_id}/logo")
async def upload_brand_logo(brand_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload logo for a brand (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Validate file
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    # Read file content
    contents = await file.read()
    if len(contents) > 2 * 1024 * 1024:  # 2MB limit
        raise HTTPException(status_code=400, detail="File too large (max 2MB)")
    
    # Convert to base64 for storage
    base64_image = base64.b64encode(contents).decode()
    logo_url = f"data:{file.content_type};base64,{base64_image}"
    
    # Update brand
    result = await db.brands.update_one(
        {"brand_id": brand_id},
        {"$set": {"logo_url": logo_url, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    return {"message": "Logo uploaded", "logo_url": logo_url}

@api_router.delete("/admin/brands/{brand_id}/logo")
async def delete_brand_logo(brand_id: str, user: dict = Depends(get_current_user)):
    """Remove logo from a brand (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.brands.update_one(
        {"brand_id": brand_id},
        {"$set": {"logo_url": None, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    return {"message": "Logo removed"}

@api_router.post("/admin/brands/reorder")
async def reorder_brands(data: dict, user: dict = Depends(get_current_user)):
    """Reorder brands (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    brand_orders = data.get("brand_orders", [])
    for item in brand_orders:
        await db.brands.update_one(
            {"brand_id": item["brand_id"]},
            {"$set": {"order": item["order"]}}
        )
    
    return {"message": "Brands reordered"}

@api_router.post("/admin/brands/seed-defaults")
async def seed_default_brands(user: dict = Depends(get_current_user)):
    """Seed default KX brands (admin only) - call once to initialize"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Check if brands already exist
    existing = await db.brands.count_documents({})
    if existing > 0:
        return {"message": f"Brands already exist ({existing} brands). Delete all to re-seed.", "seeded": 0}
    
    default_brands = [
        {"name": "KX CORE", "description": "Foundation programs for essential skills", "color": "#00f0ff", "order": 1},
        {"name": "KX PRO", "description": "Professional-grade advanced training", "color": "#ff6b35", "order": 2},
        {"name": "KX LAB", "description": "Experimental and research programs", "color": "#7c3aed", "order": 3},
        {"name": "KX MEDIA", "description": "Motorsport journalism and broadcasting", "color": "#f59e0b", "order": 4},
        {"name": "KX TECH", "description": "Automotive technology and engineering", "color": "#10b981", "order": 5},
        {"name": "KX RACING", "description": "Professional racing and karting programs", "color": "#ef4444", "order": 6},
        {"name": "KX BUSINESS", "description": "Motorsport business and management", "color": "#6366f1", "order": 7},
        {"name": "KX ACADEMY", "description": "Foundational certification courses", "color": "#14b8a6", "order": 8},
        {"name": "KX EVENTS", "description": "Event management and coordination", "color": "#ec4899", "order": 9},
        {"name": "KX DESIGN", "description": "Automotive and livery design programs", "color": "#8b5cf6", "order": 10},
        {"name": "KX SAFETY", "description": "Safety training and marshal programs", "color": "#22c55e", "order": 11},
        {"name": "KX GLOBAL", "description": "International partnerships and programs", "color": "#0ea5e9", "order": 12},
        {"name": "KX PARTNERS", "description": "Industry collaboration programs", "color": "#f97316", "order": 13},
    ]
    
    for brand_data in default_brands:
        brand_id = f"brand_{uuid.uuid4().hex[:12]}"
        brand_doc = {
            "brand_id": brand_id,
            **brand_data,
            "logo_url": None,
            "is_visible": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user["user_id"]
        }
        await db.brands.insert_one(brand_doc)
    
    return {"message": f"Seeded {len(default_brands)} default brands", "seeded": len(default_brands)}

# ======================== BRAND HEAD MANAGEMENT ========================

class BrandHeadAssign(BaseModel):
    user_id: str
    brand_id: str

@api_router.post("/admin/brand-heads/assign")
async def assign_brand_head(data: BrandHeadAssign, user: dict = Depends(get_current_user)):
    """Assign a user as Brand Head for a specific brand (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Verify brand exists
    brand = await db.brands.find_one({"brand_id": data.brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    # Verify user exists
    target_user = await db.users.find_one({"user_id": data.user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update user role and assigned brand
    await db.users.update_one(
        {"user_id": data.user_id},
        {"$set": {
            "role": "brand_head",
            "assigned_brand_id": data.brand_id,
            "assigned_brand_name": brand.get("name"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update brand with brand_head_id
    await db.brands.update_one(
        {"brand_id": data.brand_id},
        {"$set": {
            "brand_head_id": data.user_id,
            "brand_head_name": target_user.get("name"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"User assigned as Brand Head for {brand.get('name')}"}

@api_router.delete("/admin/brand-heads/{user_id}")
async def remove_brand_head(user_id: str, user: dict = Depends(get_current_user)):
    """Remove Brand Head role from a user (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    target_user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    brand_id = target_user.get("assigned_brand_id")
    
    # Remove brand head from user
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role": "student"}, "$unset": {"assigned_brand_id": "", "assigned_brand_name": ""}}
    )
    
    # Remove brand head from brand
    if brand_id:
        await db.brands.update_one(
            {"brand_id": brand_id},
            {"$unset": {"brand_head_id": "", "brand_head_name": ""}}
        )
    
    return {"message": "Brand Head role removed"}

@api_router.post("/admin/crew/assign-brand")
async def assign_crew_to_brand(data: dict, user: dict = Depends(get_current_user)):
    """Assign a crew/trainer to a specific brand (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    user_id = data.get("user_id")
    brand_id = data.get("brand_id")
    
    if not user_id:
        raise HTTPException(status_code=400, detail="user_id required")
    
    # Verify user exists and is a trainer
    target_user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if brand_id:
        # Verify brand exists
        brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
        if not brand:
            raise HTTPException(status_code=404, detail="Brand not found")
        
        # Assign to brand
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {
                "assigned_brand_id": brand_id,
                "assigned_brand_name": brand.get("name"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        return {"message": f"Crew assigned to {brand.get('name')}"}
    else:
        # Remove brand assignment
        await db.users.update_one(
            {"user_id": user_id},
            {"$unset": {"assigned_brand_id": "", "assigned_brand_name": ""}}
        )
        return {"message": "Brand assignment removed"}

class CreateCrewRequest(BaseModel):
    name: str
    email: str
    password: str
    brand_id: Optional[str] = None

@api_router.post("/admin/crew/create")
async def create_crew_member(data: CreateCrewRequest, user: dict = Depends(get_current_user)):
    """Create a new crew/trainer member (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    password_hash = bcrypt.hashpw(data.password.encode(), bcrypt.gensalt()).decode()
    
    new_crew = {
        "user_id": user_id,
        "email": data.email.lower(),
        "name": data.name,
        "password_hash": password_hash,
        "role": "trainer",
        "assigned_brand_id": data.brand_id if data.brand_id else None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("user_id")
    }
    
    await db.users.insert_one(new_crew)
    
    return {"message": "Crew member created", "user_id": user_id}

@api_router.get("/admin/crew")
async def get_all_crew(user: dict = Depends(get_current_user)):
    """Get all crew/trainers (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    crew = await db.users.find(
        {"role": "trainer"},
        {"_id": 0, "password": 0}
    ).to_list(100)
    
    return crew

@api_router.get("/crew/dashboard")
async def get_crew_dashboard(user: dict = Depends(get_current_user)):
    """Get crew dashboard data - filtered by assigned brand if applicable"""
    if user.get("role") not in ["trainer", "admin"]:
        raise HTTPException(status_code=403, detail="Crew/Admin only")
    
    brand_id = user.get("assigned_brand_id")
    
    # Get brand info if assigned
    brand = None
    if brand_id:
        brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
    
    # Get programs - filter by brand if assigned
    if brand_id:
        programs = await db.programs.find({"brand_id": brand_id}, {"_id": 0}).to_list(50)
    else:
        programs = await db.programs.find({}, {"_id": 0}).to_list(50)
    
    # Get students - for now all students, could filter by program enrollment later
    students = await db.students.find({}, {"_id": 0}).to_list(100)
    
    # Get batches
    batches = await db.batches.find({}, {"_id": 0}).to_list(50)
    
    # Get recent assessments
    assessments = await db.assessments.find({}).sort("timestamp", -1).to_list(20)
    # Convert ObjectIds for serialization
    assessments = [{k: str(v) if k == "_id" else v for k, v in a.items()} for a in assessments]
    
    return {
        "brand": brand,
        "is_brand_locked": brand_id is not None,
        "programs": programs,
        "students": students[:50],
        "batches": batches,
        "assessments": assessments,
        "stats": {
            "total_programs": len(programs),
            "total_students": len(students),
            "recent_assessments": len(assessments)
        }
    }

@api_router.get("/admin/brand-heads")
async def get_all_brand_heads(user: dict = Depends(get_current_user)):
    """Get all brand heads (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    brand_heads = await db.users.find(
        {"role": "brand_head"},
        {"_id": 0, "password": 0}
    ).to_list(100)
    
    return brand_heads

# ======================== USER ACCOUNT MANAGEMENT ========================

class CreateUserAccount(BaseModel):
    email: str
    password: str
    name: str
    brand_id: Optional[str] = None  # Required for brand_head and crew

@api_router.post("/admin/users/brand-manager")
async def create_brand_manager_account(data: CreateUserAccount, user: dict = Depends(get_current_user)):
    """Admin creates a Brand Manager (Brand Head) account with credentials"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    if not data.brand_id:
        raise HTTPException(status_code=400, detail="Brand ID is required for Brand Manager")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Verify brand exists
    brand = await db.brands.find_one({"brand_id": data.brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    # Create user account
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    hashed_password = bcrypt.hashpw(data.password.encode('utf-8'), bcrypt.gensalt())
    
    user_doc = {
        "user_id": user_id,
        "email": data.email,
        "password": hashed_password.decode('utf-8'),
        "name": data.name,
        "role": "brand_head",
        "assigned_brand_id": data.brand_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.users.insert_one(user_doc)
    
    # Update brand with brand_head info
    await db.brands.update_one(
        {"brand_id": data.brand_id},
        {"$set": {"brand_head_id": user_id, "brand_head_name": data.name}}
    )
    
    # Send welcome email with credentials (non-blocking, doesn't fail if email not configured)
    email_sent = False
    try:
        result = await send_welcome_email(data.email, data.name, "brand_head", data.password, brand.get("name"))
        email_sent = result is not None
    except Exception as e:
        logger.warning(f"Could not send welcome email: {e}")
    
    return {
        "message": "Brand Manager created successfully",
        "user_id": user_id,
        "email": data.email,
        "name": data.name,
        "role": "brand_head",
        "assigned_brand_id": data.brand_id,
        "brand_name": brand.get("name"),
        "email_sent": email_sent
    }

@api_router.post("/brand-head/crew")
async def create_crew_account(data: CreateUserAccount, user: dict = Depends(get_current_user)):
    """Brand Head creates a Crew (Trainer) account for their own brand"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned to you")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Get brand info
    brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    # Create crew account
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    hashed_password = bcrypt.hashpw(data.password.encode('utf-8'), bcrypt.gensalt())
    
    user_doc = {
        "user_id": user_id,
        "email": data.email,
        "password": hashed_password.decode('utf-8'),
        "name": data.name,
        "role": "trainer",
        "assigned_brand_id": brand_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.users.insert_one(user_doc)
    
    # Send welcome email with credentials (non-blocking)
    email_sent = False
    try:
        result = await send_welcome_email(data.email, data.name, "trainer", data.password, brand.get("name"))
        email_sent = result is not None
    except Exception as e:
        logger.warning(f"Could not send welcome email: {e}")
    
    return {
        "message": "Crew member created successfully",
        "user_id": user_id,
        "email": data.email,
        "name": data.name,
        "role": "trainer",
        "assigned_brand_id": brand_id,
        "brand_name": brand.get("name"),
        "email_sent": email_sent
    }

@api_router.get("/brand-head/crew")
async def get_brand_head_crew(user: dict = Depends(get_current_user)):
    """Get all crew members for the Brand Head's brand"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    crew = await db.users.find(
        {"role": "trainer", "assigned_brand_id": brand_id},
        {"_id": 0, "password": 0}
    ).to_list(100)
    
    return crew

@api_router.delete("/brand-head/crew/{user_id}")
async def delete_crew_account(user_id: str, user: dict = Depends(get_current_user)):
    """Brand Head can delete crew members from their brand"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Verify the crew member belongs to this brand
    crew_member = await db.users.find_one({"user_id": user_id, "role": "trainer"})
    if not crew_member:
        raise HTTPException(status_code=404, detail="Crew member not found")
    
    if crew_member.get("assigned_brand_id") != brand_id:
        raise HTTPException(status_code=403, detail="This crew member does not belong to your brand")
    
    await db.users.delete_one({"user_id": user_id})
    
    return {"message": "Crew member deleted successfully"}

@api_router.get("/brand-head/dashboard")
async def get_brand_head_dashboard(user: dict = Depends(get_current_user)):
    """Get Brand Head dashboard data"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    # Check for brand_id (supports both field names)
    brand_id = user.get("assigned_brand_id") or user.get("brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Get brand details
    brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    # Get programs for this brand (or all programs if brand_id not set on programs)
    programs = await db.programs.find(
        {"$or": [{"brand_id": brand_id}, {"brand_id": {"$exists": False}}]},
        {"_id": 0}
    ).to_list(50)
    
    # Get crew/trainers assigned to this brand
    crew = await db.users.find(
        {"role": "trainer", "assigned_brand_id": brand_id},
        {"_id": 0, "password": 0}
    ).to_list(50)
    
    # If no crew assigned to brand, get all trainers
    if not crew:
        crew = await db.users.find(
            {"role": "trainer"},
            {"_id": 0, "password": 0}
        ).to_list(20)
    
    # Get students enrolled in this brand's programs
    # For now, get all students
    students = await db.students.find(
        {"status": "active"},
        {"_id": 0}
    ).to_list(100)
    
    # Get recent assessments
    assessments = await db.assessments.find(
        {},
        {"_id": 0}
    ).sort("timestamp", -1).to_list(20)
    
    return {
        "brand": brand,
        "programs": programs,
        "crew": crew,
        "students": students[:50],
        "assessments": assessments,
        "stats": {
            "total_programs": len(programs),
            "total_crew": len(crew),
            "total_students": len(students),
            "recent_assessments": len(assessments)
        }
    }

@api_router.get("/brand-head/brand")
async def get_brand_head_brand(user: dict = Depends(get_current_user)):
    """Get the brand assigned to the current brand head"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    return brand

# ======================== BRAND HEAD REPORTS ========================

@api_router.get("/brand-head/reports/incomplete-students")
async def get_incomplete_students_report(user: dict = Depends(get_current_user)):
    """Get students with incomplete units for the brand head's brand"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Get all units assigned to this brand
    units = await db.program_units.find(
        {"brand_id": brand_id},
        {"_id": 0}
    ).to_list(100)
    
    if not units:
        return {"students": [], "units": [], "message": "No units assigned to this brand"}
    
    unit_ids = [u["unit_id"] for u in units]
    
    # Get all students
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    
    # Get all progress records for these units
    progress_records = await db.student_progress.find(
        {"unit_id": {"$in": unit_ids}},
        {"_id": 0}
    ).to_list(5000)
    
    # Group progress by student
    student_progress = {}
    for record in progress_records:
        sid = record["student_id"]
        if sid not in student_progress:
            student_progress[sid] = {}
        student_progress[sid][record["unit_id"]] = record
    
    # Build report
    incomplete_students = []
    for student in students:
        sid = student["student_id"]
        prog = student_progress.get(sid, {})
        
        student_units = []
        incomplete_count = 0
        for unit in units:
            unit_status = prog.get(unit["unit_id"], {}).get("status", "not_started")
            if unit_status != "completed":
                incomplete_count += 1
            student_units.append({
                "unit_id": unit["unit_id"],
                "unit_name": unit["name"],
                "status": unit_status,
                "completion_date": prog.get(unit["unit_id"], {}).get("completion_date"),
                "score": prog.get(unit["unit_id"], {}).get("score")
            })
        
        if incomplete_count > 0:
            incomplete_students.append({
                "student_id": sid,
                "student_name": student.get("full_name", "Unknown"),
                "mobile": student.get("mobile", ""),
                "email": student.get("email", ""),
                "city": student.get("city", ""),
                "total_units": len(units),
                "completed_units": len(units) - incomplete_count,
                "incomplete_units": incomplete_count,
                "progress_percent": round(((len(units) - incomplete_count) / len(units)) * 100),
                "units": student_units
            })
    
    # Sort by incomplete count (most incomplete first)
    incomplete_students.sort(key=lambda x: x["incomplete_units"], reverse=True)
    
    return {
        "brand_id": brand_id,
        "total_students": len(students),
        "students_with_incomplete": len(incomplete_students),
        "units": [{"unit_id": u["unit_id"], "name": u["name"], "program_id": u["program_id"]} for u in units],
        "students": incomplete_students
    }

@api_router.get("/brand-head/reports/export-excel")
async def export_incomplete_students_excel(user: dict = Depends(get_current_user)):
    """Export incomplete students report as Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Get brand info
    brand = await db.brands.find_one({"brand_id": brand_id}, {"_id": 0})
    brand_name = brand.get("name", "Unknown") if brand else "Unknown"
    
    # Get all units assigned to this brand
    units = await db.program_units.find(
        {"brand_id": brand_id},
        {"_id": 0}
    ).to_list(100)
    
    unit_ids = [u["unit_id"] for u in units]
    
    # Get all students
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    
    # Get all progress records
    progress_records = await db.student_progress.find(
        {"unit_id": {"$in": unit_ids}},
        {"_id": 0}
    ).to_list(5000)
    
    # Group progress by student
    student_progress = {}
    for record in progress_records:
        sid = record["student_id"]
        if sid not in student_progress:
            student_progress[sid] = {}
        student_progress[sid][record["unit_id"]] = record
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Incomplete Students"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    status_complete = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    status_progress = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
    status_not_started = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = f"{brand_name} - Incomplete Students Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Student Name", "Mobile", "Email", "City", "Progress %", "Completed", "Incomplete"]
    for unit in units:
        headers.append(unit["name"][:20])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 5
    for student in students:
        sid = student["student_id"]
        prog = student_progress.get(sid, {})
        
        completed = sum(1 for u in units if prog.get(u["unit_id"], {}).get("status") == "completed")
        incomplete = len(units) - completed
        
        if incomplete == 0:
            continue  # Skip students who completed all units
        
        progress_pct = round((completed / len(units)) * 100) if units else 0
        
        ws.cell(row=row, column=1, value=student.get("full_name", "")).border = thin_border
        ws.cell(row=row, column=2, value=student.get("mobile", "")).border = thin_border
        ws.cell(row=row, column=3, value=student.get("email", "")).border = thin_border
        ws.cell(row=row, column=4, value=student.get("city", "")).border = thin_border
        ws.cell(row=row, column=5, value=f"{progress_pct}%").border = thin_border
        ws.cell(row=row, column=6, value=completed).border = thin_border
        ws.cell(row=row, column=7, value=incomplete).border = thin_border
        
        # Unit status columns
        for col, unit in enumerate(units, 8):
            status = prog.get(unit["unit_id"], {}).get("status", "not_started")
            cell = ws.cell(row=row, column=col, value=status.replace("_", " ").title())
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if status == "completed":
                cell.fill = status_complete
                cell.font = Font(color="FFFFFF")
            elif status == "in_progress":
                cell.fill = status_progress
                cell.font = Font(color="FFFFFF")
            else:
                cell.fill = status_not_started
                cell.font = Font(color="FFFFFF")
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    for col in range(8, 8 + len(units)):
        ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{brand_name.replace(' ', '_')}_Incomplete_Students_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Re-class Management
class ReClassCreate(BaseModel):
    unit_id: str
    scheduled_date: str
    scheduled_time: str
    location: Optional[str] = ""
    notes: Optional[str] = ""
    student_ids: List[str] = []

@api_router.post("/brand-head/reclass")
async def create_reclass(data: ReClassCreate, user: dict = Depends(get_current_user)):
    """Create a re-class session for students with incomplete units"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Verify the unit belongs to this brand
    unit = await db.program_units.find_one({"unit_id": data.unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    if unit.get("brand_id") != brand_id:
        raise HTTPException(status_code=403, detail="This unit does not belong to your brand")
    
    reclass_id = f"reclass_{uuid.uuid4().hex[:12]}"
    reclass_doc = {
        "reclass_id": reclass_id,
        "unit_id": data.unit_id,
        "unit_name": unit["name"],
        "brand_id": brand_id,
        "scheduled_date": data.scheduled_date,
        "scheduled_time": data.scheduled_time,
        "location": data.location,
        "notes": data.notes,
        "student_ids": data.student_ids,
        "status": "scheduled",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.reclasses.insert_one(reclass_doc)
    
    return {k: v for k, v in reclass_doc.items() if k != "_id"}

@api_router.get("/brand-head/reclasses")
async def get_brand_reclasses(user: dict = Depends(get_current_user)):
    """Get all re-classes for the brand head's brand"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    reclasses = await db.reclasses.find(
        {"brand_id": brand_id},
        {"_id": 0}
    ).sort("scheduled_date", -1).to_list(100)
    
    return reclasses

@api_router.put("/brand-head/reclass/{reclass_id}")
async def update_reclass(reclass_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update a re-class (status, add/remove students)"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    # Verify the reclass belongs to this brand
    reclass = await db.reclasses.find_one({"reclass_id": reclass_id})
    if not reclass:
        raise HTTPException(status_code=404, detail="Re-class not found")
    if reclass.get("brand_id") != brand_id:
        raise HTTPException(status_code=403, detail="This re-class does not belong to your brand")
    
    data.pop("reclass_id", None)
    data.pop("brand_id", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.reclasses.update_one(
        {"reclass_id": reclass_id},
        {"$set": data}
    )
    
    return {"message": "Re-class updated"}

@api_router.delete("/brand-head/reclass/{reclass_id}")
async def delete_reclass(reclass_id: str, user: dict = Depends(get_current_user)):
    """Delete a re-class"""
    if user.get("role") != "brand_head":
        raise HTTPException(status_code=403, detail="Brand Head only")
    
    brand_id = user.get("assigned_brand_id")
    if not brand_id:
        raise HTTPException(status_code=400, detail="No brand assigned")
    
    reclass = await db.reclasses.find_one({"reclass_id": reclass_id})
    if not reclass:
        raise HTTPException(status_code=404, detail="Re-class not found")
    if reclass.get("brand_id") != brand_id:
        raise HTTPException(status_code=403, detail="This re-class does not belong to your brand")
    
    await db.reclasses.delete_one({"reclass_id": reclass_id})
    
    return {"message": "Re-class deleted"}

# ======================== PARTNERS & SPONSORS ========================

@api_router.get("/partners")
async def get_partners():
    """Get all visible partners/sponsors (public)"""
    partners = await db.partners.find(
        {"is_visible": True},
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    return partners

@api_router.get("/admin/partners")
async def get_all_partners(user: dict = Depends(get_current_user)):
    """Get all partners including hidden (admin/super_admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    partners = await db.partners.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    return partners

@api_router.post("/admin/partners")
async def create_partner(data: PartnerCreate, user: dict = Depends(get_current_user)):
    """Create a new partner/sponsor"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    partner_id = f"partner_{uuid.uuid4().hex[:12]}"
    partner_doc = {
        "partner_id": partner_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.partners.insert_one(partner_doc)
    return {k: v for k, v in partner_doc.items() if k != "_id"}

@api_router.put("/admin/partners/{partner_id}")
async def update_partner(partner_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update a partner/sponsor"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data.pop("partner_id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.partners.update_one(
        {"partner_id": partner_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Partner not found")
    
    return {"message": "Partner updated"}

@api_router.delete("/admin/partners/{partner_id}")
async def delete_partner(partner_id: str, user: dict = Depends(get_current_user)):
    """Delete a partner/sponsor"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.partners.delete_one({"partner_id": partner_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Partner not found")
    
    return {"message": "Partner deleted"}

# ======================== PROGRAMME DIRECTOR ========================

@api_router.get("/cms/programme-director")
async def get_programme_director():
    """Get Programme Director info (public)"""
    director = await db.cms_content.find_one({"$or": [{"page": "programme_director"}, {"section": "programme_director"}]}, {"_id": 0})
    if not director:
        return {
            "page": "programme_director",
            "name": "Programme Director",
            "designation": "Director of Programmes",
            "message": "Welcome to KXGRID. Our mission is to nurture the next generation of motorsport professionals through world-class training and industry partnerships.",
            "photo_url": None,
            "photo_base64": None
        }
    return director

@api_router.put("/cms/programme-director")
async def update_programme_director(data: ProgrammeDirectorUpdate, user: dict = Depends(get_current_user)):
    """Update Programme Director info (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    director_doc = {
        "page": "programme_director",
        **data.model_dump(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": user["user_id"]
    }
    
    await db.cms_content.update_one(
        {"page": "programme_director"},
        {"$set": director_doc},
        upsert=True
    )
    
    return {"message": "Programme Director updated"}

# ======================== CONTACT INFO ========================

@api_router.get("/cms/contact-info")
async def get_contact_info():
    """Get contact info (public)"""
    contact = await db.cms_content.find_one({"$or": [{"page": "contact_info"}, {"section": "contact"}]}, {"_id": 0})
    if not contact:
        return {
            "page": "contact_info",
            "email": "admissions@kotlerx.com",
            "phone": "+91 98765 43210",
            "whatsapp_number": "+919876543210",
            "location_address": "KotlerX Academy, Motorsport City, India",
            "location_maps_url": "https://maps.google.com",
            "heading_text": "Questions? Please get in touch",
            "subheading_text": "Our admission team will be happy to discuss your options"
        }
    return contact

@api_router.put("/cms/contact-info")
async def update_contact_info(data: ContactInfoUpdate, user: dict = Depends(get_current_user)):
    """Update contact info (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    contact_doc = {
        "page": "contact_info",
        **data.model_dump(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": user["user_id"]
    }
    
    await db.cms_content.update_one(
        {"page": "contact_info"},
        {"$set": contact_doc},
        upsert=True
    )
    
    return {"message": "Contact info updated"}

# ======================== TEAM MEMBERS ========================

class TeamMemberCreate(BaseModel):
    name: str
    role: str
    category: str = "instructor"  # instructor, staff
    photo_url: Optional[str] = None
    photo_base64: Optional[str] = None
    bio: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    display_order: int = 0
    is_active: bool = True

@api_router.get("/team")
async def get_team_members():
    """Get all active team members (public)"""
    members = await db.team_members.find(
        {"is_active": True},
        {"_id": 0}
    ).sort("display_order", 1).to_list(100)
    return members

@api_router.get("/team/category/{category}")
async def get_team_by_category(category: str):
    """Get team members by category (public)"""
    members = await db.team_members.find(
        {"is_active": True, "category": category},
        {"_id": 0}
    ).sort("display_order", 1).to_list(100)
    return members

@api_router.get("/admin/team")
async def get_all_team_members(user: dict = Depends(get_current_user)):
    """Get all team members including inactive (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    members = await db.team_members.find({}, {"_id": 0}).sort("display_order", 1).to_list(100)
    return members

@api_router.post("/admin/team")
async def create_team_member(data: TeamMemberCreate, user: dict = Depends(get_current_user)):
    """Create a new team member"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    member_id = f"team_{uuid.uuid4().hex[:12]}"
    member_doc = {
        "member_id": member_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.team_members.insert_one(member_doc)
    return {k: v for k, v in member_doc.items() if k != "_id"}

@api_router.put("/admin/team/{member_id}")
async def update_team_member(member_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update a team member"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data.pop("member_id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.team_members.update_one({"member_id": member_id}, {"$set": data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Team member not found")
    
    return {"message": "Team member updated"}

@api_router.delete("/admin/team/{member_id}")
async def delete_team_member(member_id: str, user: dict = Depends(get_current_user)):
    """Delete a team member"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.team_members.delete_one({"member_id": member_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Team member not found")
    
    return {"message": "Team member deleted"}

@api_router.post("/admin/team/bulk-import")
async def bulk_import_team_members(payload: dict, user: dict = Depends(get_current_user)):
    """Bulk-create team members from a JSON list. Skips entries whose name already exists.
    Body: {"members": [{name, role, category, bio?, photo_url?, photo_base64?, email?, phone?, display_order?, is_active?}, ...]}
    """
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")

    members = payload.get("members") or []
    if not isinstance(members, list) or not members:
        raise HTTPException(status_code=400, detail="members must be a non-empty list")

    created, skipped = [], []
    for m in members:
        name = (m.get("name") or "").strip()
        if not name:
            skipped.append({"reason": "missing name", "entry": m})
            continue
        existing = await db.team_members.find_one({"name": name}, {"_id": 0, "member_id": 1})
        if existing:
            skipped.append({"reason": "duplicate name", "name": name, "member_id": existing["member_id"]})
            continue
        doc = {
            "member_id": f"team_{uuid.uuid4().hex[:12]}",
            "name": name,
            "role": m.get("role", ""),
            "category": m.get("category", "instructor"),
            "bio": m.get("bio", ""),
            "photo_url": m.get("photo_url", ""),
            "photo_base64": m.get("photo_base64", ""),
            "email": m.get("email"),
            "phone": m.get("phone"),
            "display_order": int(m.get("display_order", 0) or 0),
            "is_active": bool(m.get("is_active", True)),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user["user_id"],
        }
        await db.team_members.insert_one(doc)
        created.append({"member_id": doc["member_id"], "name": doc["name"]})
    return {"created": created, "skipped": skipped, "created_count": len(created), "skipped_count": len(skipped)}

@api_router.get("/admin/team/export")
async def export_team_members(user: dict = Depends(get_current_user)):
    """Export all team members as JSON for safekeeping/backup."""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    members = await db.team_members.find({}, {"_id": 0}).sort("display_order", 1).to_list(1000)
    return {"count": len(members), "exported_at": datetime.now(timezone.utc).isoformat(), "members": members}

# ======================== CALLBACK REQUESTS (Leads) ========================

@api_router.post("/callback-request")
async def create_callback_request(data: CallbackRequestCreate):
    """Create a callback request (public - converts to lead)"""
    request_id = f"callback_{uuid.uuid4().hex[:12]}"
    
    # Create callback request
    callback_doc = {
        "request_id": request_id,
        **data.model_dump(),
        "status": "new",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.callback_requests.insert_one(callback_doc)
    
    # Also create a lead entry
    lead_id = f"lead_{uuid.uuid4().hex[:12]}"
    lead_doc = {
        "lead_id": lead_id,
        "name": data.name,
        "mobile": data.phone,
        "email": data.email,
        "location": "Callback Request",
        "program_interest": data.message or "General Inquiry",
        "fee_type": "unknown",
        "status": "new",
        "source": "callback_request",
        "callback_request_id": request_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.leads.insert_one(lead_doc)
    
    return {"message": "Callback request received. Our team will contact you soon.", "request_id": request_id}

@api_router.get("/admin/callback-requests")
async def get_callback_requests(user: dict = Depends(get_current_user)):
    """Get all callback requests (admin only)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    requests = await db.callback_requests.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return requests

@api_router.put("/admin/callback-requests/{request_id}")
async def update_callback_request(request_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update callback request status"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = user["user_id"]
    
    await db.callback_requests.update_one(
        {"request_id": request_id},
        {"$set": data}
    )
    
    return {"message": "Callback request updated"}

# ======================== SUPER ADMIN (KX ROOT) ========================

SUPER_ADMIN_EMAIL = "root@kotlerx.com"

@api_router.post("/super-admin/create-admin")
async def create_admin_account(data: AdminCreate, user: dict = Depends(get_current_user)):
    """Create an admin account (Super Admin only)"""
    if user.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Super Admin only")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = f"admin_{uuid.uuid4().hex[:12]}"
    admin_doc = {
        "user_id": user_id,
        "email": data.email,
        "name": data.name,
        "password_hash": hash_password(data.password),
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"],
        "picture": None
    }
    
    await db.users.insert_one(admin_doc)
    
    # Try to send welcome email
    try:
        if resend.api_key and resend.api_key != 're_placeholder':
            email_body = f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background: #1a1a2e; color: #ffffff;">
                <h1 style="color: #00f0ff; text-align: center;">Welcome to KXGRID Admin</h1>
                <p>Hello {data.name},</p>
                <p>Your Admin account has been created. Here are your login credentials:</p>
                <div style="background: #16213e; padding: 20px; border-radius: 8px; margin: 20px 0;">
                    <p><strong>Email:</strong> {data.email}</p>
                    <p><strong>Password:</strong> {data.password}</p>
                </div>
                <p>Please change your password after first login.</p>
                <p style="margin-top: 30px;">Best regards,<br>KX ROOT - Super Admin</p>
            </div>
            """
            resend.Emails.send({
                "from": SENDER_EMAIL,
                "to": data.email,
                "subject": "KXGRID Admin Account Created",
                "html": email_body
            })
    except Exception as e:
        logger.warning(f"Failed to send admin welcome email: {e}")
    
    return {"message": f"Admin account created for {data.email}", "user_id": user_id}

@api_router.get("/super-admin/admins")
async def get_all_admins(user: dict = Depends(get_current_user)):
    """Get all admin accounts (Super Admin only)"""
    if user.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Super Admin only")
    
    admins = await db.users.find(
        {"role": "admin"},
        {"_id": 0, "password_hash": 0}
    ).to_list(100)
    
    return admins

@api_router.delete("/super-admin/admins/{user_id}")
async def delete_admin_account(user_id: str, user: dict = Depends(get_current_user)):
    """Delete an admin account (Super Admin only)"""
    if user.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Super Admin only")
    
    # Prevent deleting super admin
    admin = await db.users.find_one({"user_id": user_id})
    if admin and admin.get("role") == "super_admin":
        raise HTTPException(status_code=400, detail="Cannot delete Super Admin account")
    
    result = await db.users.delete_one({"user_id": user_id, "role": "admin"})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Admin not found")
    
    return {"message": "Admin account deleted"}

@api_router.get("/super-admin/dashboard")
async def get_super_admin_dashboard(user: dict = Depends(get_current_user)):
    """Get Super Admin dashboard stats"""
    if user.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Super Admin only")
    
    # Count various entities
    admin_count = await db.users.count_documents({"role": "admin"})
    brand_head_count = await db.users.count_documents({"role": "brand_head"})
    crew_count = await db.users.count_documents({"role": "trainer"})
    student_count = await db.students.count_documents({})
    program_count = await db.programs.count_documents({})
    brand_count = await db.brands.count_documents({})
    lead_count = await db.leads.count_documents({})
    callback_count = await db.callback_requests.count_documents({"status": "new"})
    
    return {
        "admins": admin_count,
        "brand_heads": brand_head_count,
        "crew": crew_count,
        "students": student_count,
        "programs": program_count,
        "brands": brand_count,
        "total_leads": lead_count,
        "pending_callbacks": callback_count
    }

# Seed Super Admin on startup (if not exists)
async def seed_super_admin():
    """Create the KX ROOT Super Admin if it doesn't exist"""
    try:
        existing = await db.users.find_one({"email": SUPER_ADMIN_EMAIL})
        if not existing:
            super_admin_doc = {
                "user_id": "kx_root_superadmin",
                "email": SUPER_ADMIN_EMAIL,
                "name": "KX ROOT",
                "password_hash": hash_password("KXRoot@2024"),
                "role": "super_admin",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "picture": None
            }
            await db.users.insert_one(super_admin_doc)
            logger.info("KX ROOT Super Admin created")
    except Exception as e:
        logger.warning(f"Could not seed super admin (will retry on next startup): {e}")

# ==================== KXCraft Products ====================

class KXCraftProductCreate(BaseModel):
    name: str
    category: str = "motorsport"
    price: str = "0"
    description: str = ""
    image_url: str = ""
    image_base64: Optional[str] = None
    badge: Optional[str] = None
    rating: float = 5.0
    is_visible: bool = True
    order: int = 0
    external_link: Optional[str] = None

@api_router.get("/kxcraft/products")
async def get_kxcraft_products():
    products = await db.kxcraft_products.find({"is_visible": True}, {"_id": 0}).sort("order", 1).to_list(100)
    return products

@api_router.get("/kxcraft/products/all")
async def get_all_kxcraft_products():
    products = await db.kxcraft_products.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    return products

@api_router.post("/admin/kxcraft/products")
async def create_kxcraft_product(product: KXCraftProductCreate, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "root"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    product_data = product.dict()
    product_data["product_id"] = f"kxcraft_{uuid.uuid4().hex[:8]}"
    product_data["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.kxcraft_products.insert_one(product_data)
    product_data.pop("_id", None)
    return product_data

@api_router.put("/admin/kxcraft/products/{product_id}")
async def update_kxcraft_product(product_id: str, product: KXCraftProductCreate, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "root"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    update_data = product.dict()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.kxcraft_products.update_one(
        {"product_id": product_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product updated"}

@api_router.delete("/admin/kxcraft/products/{product_id}")
async def delete_kxcraft_product(product_id: str, user: dict = Depends(get_current_user)):
    if user.get("role") not in ["admin", "root"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    result = await db.kxcraft_products.delete_one({"product_id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted"}

@api_router.get("/kxcraft/categories")
async def get_kxcraft_categories():
    categories = await db.kxcraft_products.distinct("category")
    return categories



@app.on_event("startup")
async def startup_event():
    """Application startup - seed initial data only.
    NOTE: Auto-sync from production has been permanently disabled. The previous
    sync routine performed destructive delete_many() operations and was the root
    cause of team_members data loss. Never re-enable without an explicit, manual,
    audited migration plan and a backup taken first.
    """
    logger.info("KXGRID API starting up...")
    await seed_super_admin()
    logger.info("KXGRID API startup complete")

async def seed_promo_banners():
    """Seed initial promotional banners"""
    existing = await db.promo_banners.count_documents({})
    if existing > 0:
        return
    
    banners = [
        {
            "banner_id": "banner_workshop001",
            "title": "Workshop on Custom Painting",
            "description": "Master the art of custom vehicle painting. Register for workshops and one-day programmes!",
            "button_text": "Register Now",
            "link_url": None,
            "link_type": "registration",
            "background_color": "#1a1a2e",
            "gradient_from": "#f59e0b",
            "gradient_to": "#ef4444",
            "icon": "Paintbrush",
            "is_active": True,
            "display_order": 1,
            "registration_enabled": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "banner_id": "banner_fitness002",
            "title": "Online Class for Motorsport Fitness",
            "description": "Get race-ready with professional fitness training designed for motorsport athletes.",
            "button_text": "Join Classes",
            "link_url": "https://www.kotlerxiron.com",
            "link_type": "external",
            "background_color": "#1a1a2e",
            "gradient_from": "#10b981",
            "gradient_to": "#06b6d4",
            "icon": "Dumbbell",
            "is_active": True,
            "display_order": 2,
            "registration_enabled": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "banner_id": "banner_connect003",
            "title": "Connect. Create. Earn.",
            "description": "Join the community! A business portal for visual creators and influencers.",
            "button_text": "Join Community",
            "link_url": "https://www.kotlerxconnect.com",
            "link_type": "external",
            "background_color": "#1a1a2e",
            "gradient_from": "#8b5cf6",
            "gradient_to": "#ec4899",
            "icon": "Users",
            "is_active": True,
            "display_order": 3,
            "registration_enabled": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "banner_id": "banner_ride004",
            "title": "Join the Ride",
            "description": "Be part of the riding community. Adventures, meetups, and more await!",
            "button_text": "Start Riding",
            "link_url": "https://www.kotlerxcore.com",
            "link_type": "external",
            "background_color": "#1a1a2e",
            "gradient_from": "#00f0ff",
            "gradient_to": "#0066ff",
            "icon": "Bike",
            "is_active": True,
            "display_order": 4,
            "registration_enabled": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
    ]
    
    await db.promo_banners.insert_many(banners)
    logger.info("Seeded initial promotional banners")

# ======================== CAREERS ========================

class CareerCreate(BaseModel):
    title: str
    company: str
    location: str
    job_type: str = "full-time"  # full-time, part-time, internship, contract
    description: str
    requirements: List[str] = []
    brand_id: Optional[str] = None
    salary_range: Optional[str] = None
    application_url: Optional[str] = None
    is_active: bool = True

@api_router.get("/careers")
async def get_careers():
    """Get all active career opportunities (public)"""
    careers = await db.careers.find(
        {"is_active": True},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    # Enrich with brand info
    for career in careers:
        if career.get("brand_id"):
            brand = await db.brands.find_one({"brand_id": career["brand_id"]}, {"_id": 0, "name": 1, "color": 1})
            if brand:
                career["brand_name"] = brand.get("name")
                career["brand_color"] = brand.get("color", "#00f0ff")
    
    return careers

@api_router.get("/admin/careers")
async def get_all_careers(user: dict = Depends(get_current_user)):
    """Get all careers including inactive (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    careers = await db.careers.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return careers

@api_router.post("/admin/careers")
async def create_career(data: CareerCreate, user: dict = Depends(get_current_user)):
    """Create a new career opportunity"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    career_id = f"career_{uuid.uuid4().hex[:12]}"
    career_doc = {
        "career_id": career_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.careers.insert_one(career_doc)
    return {k: v for k, v in career_doc.items() if k != "_id"}

@api_router.put("/admin/careers/{career_id}")
async def update_career(career_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update a career opportunity"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data.pop("career_id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.careers.update_one({"career_id": career_id}, {"$set": data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Career not found")
    
    return {"message": "Career updated"}

@api_router.delete("/admin/careers/{career_id}")
async def delete_career(career_id: str, user: dict = Depends(get_current_user)):
    """Delete a career opportunity"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.careers.delete_one({"career_id": career_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Career not found")
    
    return {"message": "Career deleted"}

# ======================== PROMOTIONAL BANNERS ========================

class PromoBannerCreate(BaseModel):
    title: str
    description: str
    button_text: str = "Learn More"
    link_url: Optional[str] = None
    link_type: str = "external"  # external, registration, internal
    background_color: str = "#1a1a2e"
    gradient_from: str = "#00f0ff"
    gradient_to: str = "#ff00ff"
    icon: str = "Zap"  # Lucide icon name
    logo_url: Optional[str] = None  # Custom logo image URL
    is_active: bool = True
    display_order: int = 0
    registration_enabled: bool = False  # For workshop registrations

@api_router.get("/promo-banners")
async def get_promo_banners():
    """Get all active promotional banners (public)"""
    banners = await db.promo_banners.find(
        {"is_active": True},
        {"_id": 0}
    ).sort("display_order", 1).to_list(50)
    return banners

@api_router.get("/admin/promo-banners")
async def get_all_promo_banners(user: dict = Depends(get_current_user)):
    """Get all promotional banners including inactive (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    banners = await db.promo_banners.find({}, {"_id": 0}).sort("display_order", 1).to_list(50)
    return banners

@api_router.post("/admin/promo-banners")
async def create_promo_banner(data: PromoBannerCreate, user: dict = Depends(get_current_user)):
    """Create a new promotional banner"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    banner_id = f"banner_{uuid.uuid4().hex[:12]}"
    banner_doc = {
        "banner_id": banner_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.promo_banners.insert_one(banner_doc)
    return {k: v for k, v in banner_doc.items() if k != "_id"}

@api_router.put("/admin/promo-banners/{banner_id}")
async def update_promo_banner(banner_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update a promotional banner"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    data.pop("banner_id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.promo_banners.update_one({"banner_id": banner_id}, {"$set": data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Banner not found")
    
    return {"message": "Banner updated"}

@api_router.delete("/admin/promo-banners/{banner_id}")
async def delete_promo_banner(banner_id: str, user: dict = Depends(get_current_user)):
    """Delete a promotional banner"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    result = await db.promo_banners.delete_one({"banner_id": banner_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Banner not found")
    
    return {"message": "Banner deleted"}

# ======================== WORKSHOP/EVENT REGISTRATIONS ========================

class WorkshopRegistration(BaseModel):
    banner_id: str
    name: str
    email: str
    phone: str
    message: Optional[str] = None

@api_router.post("/workshop-register")
async def register_for_workshop(data: WorkshopRegistration):
    """Register for a workshop/event (public)"""
    # Check if banner exists and has registration enabled
    banner = await db.promo_banners.find_one({"banner_id": data.banner_id}, {"_id": 0})
    if not banner:
        raise HTTPException(status_code=404, detail="Event not found")
    
    if not banner.get("registration_enabled"):
        raise HTTPException(status_code=400, detail="Registration not available for this event")
    
    registration_id = f"reg_{uuid.uuid4().hex[:12]}"
    reg_doc = {
        "registration_id": registration_id,
        "banner_id": data.banner_id,
        "event_title": banner.get("title"),
        "name": data.name,
        "email": data.email,
        "phone": data.phone,
        "message": data.message,
        "status": "pending",
        "registered_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.workshop_registrations.insert_one(reg_doc)
    return {"message": "Registration successful", "registration_id": registration_id}

@api_router.get("/admin/workshop-registrations")
async def get_workshop_registrations(user: dict = Depends(get_current_user), banner_id: Optional[str] = None):
    """Get all workshop registrations (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    query = {}
    if banner_id:
        query["banner_id"] = banner_id
    
    registrations = await db.workshop_registrations.find(query, {"_id": 0}).sort("registered_at", -1).to_list(1000)
    return registrations

@api_router.get("/admin/workshop-registrations/export")
async def export_workshop_registrations(user: dict = Depends(get_current_user), banner_id: Optional[str] = None):
    """Export workshop registrations as CSV (admin)"""
    if user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    query = {}
    if banner_id:
        query["banner_id"] = banner_id
    
    registrations = await db.workshop_registrations.find(query, {"_id": 0}).sort("registered_at", -1).to_list(1000)
    
    # Create CSV content
    import csv
    import io
    
    output = io.StringIO()
    if registrations:
        fieldnames = ["registration_id", "event_title", "name", "email", "phone", "message", "status", "registered_at"]
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(registrations)
    
    csv_content = output.getvalue()
    
    from fastapi.responses import Response
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=workshop_registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )

# Include router and middleware
app.include_router(api_router)

# Include additional routers
from routers import sms, sheets
app.include_router(sms.router)
app.include_router(sheets.router)

# Set database for sheets router
sheets.set_database(db)

# CORS Configuration - Supports all Emergent domains
CORS_ORIGINS = os.environ.get('CORS_ORIGINS', '*')

# Parse allowed origins
if CORS_ORIGINS == "*":
    # For wildcard mode, we rely on the regex pattern
    # This matches all Emergent preview/production URLs and localhost
    allowed_origins = ["*"]
    allow_creds = False  # Cannot use credentials with wildcard
else:
    allowed_origins = [o.strip() for o in CORS_ORIGINS.split(',') if o.strip()]
    allow_creds = True

app.add_middleware(
    CORSMiddleware,
    allow_credentials=allow_creds,
    allow_origins=allowed_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
