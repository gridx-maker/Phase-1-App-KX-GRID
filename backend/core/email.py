import os
import logging
import asyncio
import resend
import base64
import io
from datetime import datetime, timezone, timedelta
from typing import Optional
from fastapi import HTTPException
from openpyxl import Workbook

# Core database connection & configurations
from core.database import db
from core.config import SENDER_EMAIL

logger = logging.getLogger(__name__)

def is_email_configured() -> bool:
    """Check if email is properly configured"""
    api_key = os.environ.get('RESEND_API_KEY', '')
    return bool(api_key and not api_key.startswith('re_placeholder'))

async def send_welcome_email(to_email: str, name: str, role: str, password: str, brand_name: str) -> Optional[dict]:
    """Send welcome email with login credentials to new user"""
    if not is_email_configured():
        logger.warning(f"Email not configured. Skipping welcome email for {to_email}")
        return None
    
    role_display = {
        "brand_head": "Brand Manager",
        "trainer": "Crew / Trainer"
    }.get(role, role.title())
    
    login_url = os.environ.get('FRONTEND_URL', 'https://kxgrid.kotlerx.com')
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background: linear-gradient(135deg, #1a1a2e, #16213e); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
            .content {{ background: #f8f9fa; padding: 30px; border-radius: 0 0 10px 10px; }}
            .credentials {{ background: white; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #00f0ff; }}
            .btn {{ display: inline-block; background: #00f0ff; color: #1a1a2e; padding: 12px 30px; text-decoration: none; border-radius: 6px; font-weight: bold; }}
            .footer {{ text-align: center; margin-top: 30px; color: #666; font-size: 12px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Welcome to KXGRID</h1>
                <p>KotlerX Ecosystem Platform</p>
            </div>
            <div class="content">
                <h2>Hello {name}!</h2>
                <p>You have been added as a <strong>{role_display}</strong> for <strong>{brand_name}</strong>.</p>
                
                <div class="credentials">
                    <h3>Your Login Credentials</h3>
                    <p><strong>Email:</strong> {to_email}</p>
                    <p><strong>Password:</strong> {password}</p>
                    <p style="color: #e74c3c; font-size: 12px;">⚠️ Please change your password after your first login.</p>
                </div>
                
                <p style="text-align: center;">
                    <a href="{login_url}/login" class="btn">Login to KXGRID</a>
                </p>
                
                <h3>What you can do:</h3>
                <ul>
                    {"<li>Manage your brand's crew and trainers</li><li>View student progress reports</li><li>Schedule re-classes</li><li>Download completion reports</li>" if role == "brand_head" else "<li>Record student attendance</li><li>Submit assessments</li><li>View student profiles</li>"}
                </ul>
                
                <p>If you have any questions, please contact your administrator.</p>
            </div>
            <div class="footer">
                <p>© 2026 KotlerX. All rights reserved.</p>
                <p>This is an automated message from KXGRID.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    try:
        params = {
            "from": SENDER_EMAIL,
            "to": [to_email],
            "subject": f"Welcome to KXGRID - Your {role_display} Account",
            "html": html_content
        }
        result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Welcome email sent to {to_email}: {result}")
        return result
    except Exception as e:
        logger.error(f"Failed to send welcome email to {to_email}: {e}")
        return None

async def send_email_async(to_email: str, subject: str, html_content: str) -> dict:
    """Send email using Resend (non-blocking)"""
    if not is_email_configured():
        raise HTTPException(status_code=503, detail="Email not configured. Please add Resend API key in CMS Settings or contact admin.")
    
    try:
        params = {
            "from": SENDER_EMAIL,
            "to": [to_email],
            "subject": subject,
            "html": html_content
        }
        result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Email sent to {to_email}: {result}")
        return result
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        raise

async def generate_assessment_report_data(batch_id: Optional[str] = None, report_type: str = "batch") -> list:
    """Generate assessment report data for university/college"""
    query = {"status": "active"}
    if batch_id:
        query["batch_id"] = batch_id
    
    students = await db.students.find(query, {"_id": 0}).to_list(1000)
    
    report_data = []
    for student in students:
        # Get attendance records
        attendance_records = await db.attendance.find(
            {"student_id": student["student_id"]}, {"_id": 0}
        ).to_list(500)
        
        # Get assessments
        assessments = await db.assessments.find(
            {"student_id": student["student_id"]}, {"_id": 0}
        ).to_list(100)
        
        # Calculate attendance percentage (assume 20 sessions per program)
        total_sessions = 20
        attendance_count = len(attendance_records)
        attendance_percent = round((attendance_count / total_sessions) * 100, 1) if total_sessions > 0 else 0
        
        # Calculate assessment averages
        if assessments:
            skill_avg = sum(a.get("skill_control", 0) for a in assessments) / len(assessments)
            discipline_avg = sum(a.get("discipline", 0) for a in assessments) / len(assessments)
            safety_avg = sum(a.get("safety_awareness", 0) for a in assessments) / len(assessments)
            execution_avg = sum(a.get("execution", 0) for a in assessments) / len(assessments)
            teamwork_avg = sum(a.get("teamwork", 0) for a in assessments) / len(assessments)
            overall_avg = (skill_avg + discipline_avg + safety_avg + execution_avg + teamwork_avg) / 5
        else:
            skill_avg = discipline_avg = safety_avg = execution_avg = teamwork_avg = overall_avg = 0
        
        # Get certificate status
        certificate = await db.certificates.find_one(
            {"student_id": student["student_id"]}, {"_id": 0}
        )
        cert_status = certificate.get("status", "Not Issued") if certificate else "Not Issued"
        
        # Get user email
        user = await db.users.find_one({"user_id": student.get("user_id")}, {"_id": 0, "email": 1})
        
        report_data.append({
            "student_id": student["student_id"],
            "full_name": student.get("full_name", ""),
            "email": user.get("email", "") if user else "",
            "mobile": student.get("mobile", ""),
            "nfc_card_id": student.get("nfc_card_id", ""),
            "batch_id": student.get("batch_id", ""),
            "attendance_sessions": attendance_count,
            "attendance_percent": attendance_percent,
            "skill_control": round(skill_avg, 2),
            "discipline": round(discipline_avg, 2),
            "safety_awareness": round(safety_avg, 2),
            "execution": round(execution_avg, 2),
            "teamwork": round(teamwork_avg, 2),
            "overall_rating": round(overall_avg, 2),
            "certificate_status": cert_status,
            "blood_group": student.get("blood_group", ""),
            "emergency_contact": student.get("emergency_contact", ""),
            "medical_conditions": ", ".join(student.get("medical_conditions", [])) or "None"
        })
    
    return report_data

def create_university_report_excel(report_data: list, report_title: str) -> Workbook:
    """Create Excel workbook for university report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Assessment Report"
    
    # Title row
    ws.merge_cells('A1:Q1')
    ws['A1'] = f"KotlerX University Assessment Report - {report_title}"
    ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
    
    # Generated date
    ws['A2'] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    
    # Headers
    headers = [
        "Student ID", "Full Name", "Email", "Mobile", "NFC Card ID", "Batch ID",
        "Attendance Sessions", "Attendance %", "Skill Control", "Discipline",
        "Safety Awareness", "Execution", "Teamwork", "Overall Rating",
        "Certificate Status", "Blood Group", "Emergency Contact", "Medical Conditions"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = cell.font.copy(bold=True)
    
    # Data rows
    for row, data in enumerate(report_data, 5):
        ws.cell(row=row, column=1, value=data.get("student_id", ""))
        ws.cell(row=row, column=2, value=data.get("full_name", ""))
        ws.cell(row=row, column=3, value=data.get("email", ""))
        ws.cell(row=row, column=4, value=data.get("mobile", ""))
        ws.cell(row=row, column=5, value=data.get("nfc_card_id", ""))
        ws.cell(row=row, column=6, value=data.get("batch_id", ""))
        ws.cell(row=row, column=7, value=data.get("attendance_sessions", 0))
        ws.cell(row=row, column=8, value=data.get("attendance_percent", 0))
        ws.cell(row=row, column=9, value=data.get("skill_control", 0))
        ws.cell(row=row, column=10, value=data.get("discipline", 0))
        ws.cell(row=row, column=11, value=data.get("safety_awareness", 0))
        ws.cell(row=row, column=12, value=data.get("execution", 0))
        ws.cell(row=row, column=13, value=data.get("teamwork", 0))
        ws.cell(row=row, column=14, value=data.get("overall_rating", 0))
        ws.cell(row=row, column=15, value=data.get("certificate_status", ""))
        ws.cell(row=row, column=16, value=data.get("blood_group", ""))
        ws.cell(row=row, column=17, value=data.get("emergency_contact", ""))
        ws.cell(row=row, column=18, value=data.get("medical_conditions", ""))
    
    # Summary row
    summary_row = len(report_data) + 6
    ws.cell(row=summary_row, column=1, value="SUMMARY")
    ws.cell(row=summary_row, column=1).font = ws.cell(row=summary_row, column=1).font.copy(bold=True)
    ws.cell(row=summary_row, column=2, value=f"Total Students: {len(report_data)}")
    
    if report_data:
        avg_attendance = sum(d["attendance_percent"] for d in report_data) / len(report_data)
        avg_rating = sum(d["overall_rating"] for d in report_data) / len(report_data)
        ws.cell(row=summary_row, column=8, value=f"Avg: {avg_attendance:.1f}%")
        ws.cell(row=summary_row, column=14, value=f"Avg: {avg_rating:.2f}")
    
    # Adjust column widths
    column_widths = [15, 20, 25, 15, 18, 18, 12, 12, 12, 12, 15, 12, 12, 12, 15, 12, 18, 25]
    for i, width in enumerate(column_widths, 1):
        # Handle Excel columns
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = width
    
    return wb

async def send_report_email(report_type: str, recipient_email: str, batch_id: str = None) -> dict:
    """Generate report and send via email"""
    if not is_email_configured():
        raise HTTPException(status_code=503, detail="Email not configured. Please add Resend API key to enable email reports.")
    
    try:
        # Generate report data
        if report_type == "weekly":
            report_data = await generate_assessment_report_data(report_type="weekly")
            today = datetime.now(timezone.utc)
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            title = f"Weekly Report ({week_start.strftime('%Y-%m-%d')} to {week_end.strftime('%Y-%m-%d')})"
            filename = f"university_weekly_report_{today.strftime('%Y%m%d')}.xlsx"
        elif report_type == "monthly":
            report_data = await generate_assessment_report_data(report_type="monthly")
            today = datetime.now(timezone.utc)
            title = f"Monthly Report - {today.strftime('%B %Y')}"
            filename = f"university_monthly_report_{today.strftime('%Y%m')}.xlsx"
        elif report_type == "batch" and batch_id:
            report_data = await generate_assessment_report_data(batch_id=batch_id, report_type="batch")
            title = f"Batch Report: {batch_id}"
            filename = f"university_batch_report_{batch_id}.xlsx"
        else:
            raise ValueError("Invalid report type")
        
        wb = create_university_report_excel(report_data, title)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        report_bytes = output.getvalue()
        report_base64 = base64.b64encode(report_bytes).decode()
        
        # Create email HTML
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; background-color: #0a0a0a; color: #ffffff; padding: 40px;">
            <div style="max-width: 600px; margin: 0 auto; background-color: #1a1a1a; border-radius: 12px; padding: 40px;">
                <h1 style="color: #00f0ff; margin-bottom: 20px;">KotlerX University Report</h1>
                <h2 style="color: #ffffff; font-size: 18px;">{title}</h2>
                <p style="color: #a0a0a0; line-height: 1.6;">
                    Please find attached the {report_type} assessment report for your university records.
                </p>
                <div style="background-color: #2a2a2a; padding: 20px; border-radius: 8px; margin: 20px 0;">
                    <p style="color: #00f0ff; margin: 0;">Report Summary:</p>
                    <ul style="color: #ffffff; margin: 10px 0;">
                        <li>Total Students: {len(report_data)}</li>
                        <li>Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}</li>
                    </ul>
                </div>
                <p style="color: #a0a0a0; font-size: 12px;">
                    This is an automated report from KotlerX Skill Tracking Platform.<br>
                    For questions, contact your KotlerX administrator.
                </p>
            </div>
        </body>
        </html>
        """
        
        # Send email with attachment
        params = {
            "from": SENDER_EMAIL,
            "to": [recipient_email],
            "subject": f"KotlerX {title}",
            "html": html_content,
            "attachments": [
                {
                    "filename": filename,
                    "content": report_base64
                }
            ]
        }
        
        result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Report email sent to {recipient_email}: {result}")
        return {"status": "sent", "email_id": result.get("id")}
        
    except Exception as e:
        logger.error(f"Failed to send report email: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")
